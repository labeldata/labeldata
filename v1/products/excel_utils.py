"""
Excel 업로드/다운로드 유틸리티
openpyxl을 사용하여 제품 데이터를 Excel 파일로 변환
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
import json


class ProductExcelHandler:
    """제품 데이터 Excel 처리 클래스"""
    
    # Excel 헤더 정의
    HEADERS = [
        '제품코드', '제품명', '카테고리', '설명', '태그', 
        '원료 사용 가능', '상태', '등록일', '수정일'
    ]
    
    # 버전 헤더 정의
    VERSION_HEADERS = [
        '버전번호', '버전명', '변경내용', '식품군', '식품유형', '품목유형',
        '제품명(버전)', '품목보고번호', '제조원', '내용량', '원재료명',
        '알레르기물질', '보관방법', '소비기한', '포장재질', '원산지',
        '유통원', '수입원', '영양성분(JSON)', '활성', '등록일'
    ]
    
    @staticmethod
    def export_products(products, include_versions=False):
        """
        제품 목록을 Excel 파일로 export
        
        Args:
            products: QuerySet of Product objects
            include_versions: 버전 정보 포함 여부
        
        Returns:
            HttpResponse with Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "제품 목록"
        
        # 스타일 정의
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 헤더 작성
        for col_num, header in enumerate(ProductExcelHandler.HEADERS, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 데이터 작성
        for row_num, product in enumerate(products, 2):
            ws.cell(row=row_num, column=1, value=product.product_code)
            ws.cell(row=row_num, column=2, value=product.product_name)
            ws.cell(row=row_num, column=3, value=product.product_category or '')
            ws.cell(row=row_num, column=4, value=product.description or '')
            
            # 태그를 문자열로 변환
            tags = ', '.join([f'#{tag}' for tag in product.tags]) if product.tags else ''
            ws.cell(row=row_num, column=5, value=tags)
            
            ws.cell(row=row_num, column=6, value='가능' if product.is_raw_material else '불가')
            ws.cell(row=row_num, column=7, value='활성' if product.is_active else '비활성')
            ws.cell(row=row_num, column=8, value=product.created_datetime.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=9, value=product.updated_datetime.strftime('%Y-%m-%d %H:%M'))
            
            # 테두리 적용
            for col_num in range(1, len(ProductExcelHandler.HEADERS) + 1):
                ws.cell(row=row_num, column=col_num).border = border
        
        # 열 너비 자동 조정
        for col_num in range(1, len(ProductExcelHandler.HEADERS) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 15
        
        # 버전 정보 포함
        if include_versions:
            ws_versions = wb.create_sheet(title="버전 상세")
            
            # 버전 헤더 작성
            for col_num, header in enumerate(ProductExcelHandler.VERSION_HEADERS, 1):
                cell = ws_versions.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # 버전 데이터 작성
            row_num = 2
            for product in products:
                versions = product.versions.all().order_by('-version_number')
                for version in versions:
                    ws_versions.cell(row=row_num, column=1, value=version.version_number)
                    ws_versions.cell(row=row_num, column=2, value=version.version_name or '')
                    ws_versions.cell(row=row_num, column=3, value=version.change_description or '')
                    ws_versions.cell(row=row_num, column=4, value=version.food_category or '')
                    ws_versions.cell(row=row_num, column=5, value=version.food_type or '')
                    ws_versions.cell(row=row_num, column=6, value=version.item_type or '')
                    ws_versions.cell(row=row_num, column=7, value=version.product_name or '')
                    ws_versions.cell(row=row_num, column=8, value=version.item_report_number or '')
                    ws_versions.cell(row=row_num, column=9, value=version.manufacturer or '')
                    ws_versions.cell(row=row_num, column=10, value=version.content_amount or '')
                    ws_versions.cell(row=row_num, column=11, value=version.ingredients_list or '')
                    ws_versions.cell(row=row_num, column=12, value=version.allergen_info or '')
                    ws_versions.cell(row=row_num, column=13, value=version.storage_method or '')
                    ws_versions.cell(row=row_num, column=14, value=version.expiration_date_info or '')
                    ws_versions.cell(row=row_num, column=15, value=version.packaging_material or '')
                    ws_versions.cell(row=row_num, column=16, value=version.country_of_origin or '')
                    ws_versions.cell(row=row_num, column=17, value=version.distributor or '')
                    ws_versions.cell(row=row_num, column=18, value=version.importer or '')
                    
                    # 영양성분 JSON 문자열로 변환
                    nutrition_str = json.dumps(version.nutrition_facts, ensure_ascii=False) if version.nutrition_facts else ''
                    ws_versions.cell(row=row_num, column=19, value=nutrition_str)
                    
                    ws_versions.cell(row=row_num, column=20, value='활성' if version.is_active else '비활성')
                    ws_versions.cell(row=row_num, column=21, value=version.created_at.strftime('%Y-%m-%d %H:%M'))
                    
                    # 테두리 적용
                    for col_num in range(1, len(ProductExcelHandler.VERSION_HEADERS) + 1):
                        ws_versions.cell(row=row_num, column=col_num).border = border
                    
                    row_num += 1
            
            # 열 너비 자동 조정
            for col_num in range(1, len(ProductExcelHandler.VERSION_HEADERS) + 1):
                column_letter = get_column_letter(col_num)
                ws_versions.column_dimensions[column_letter].width = 15
        
        # Excel 파일 생성
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # HttpResponse 생성
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f'products_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @staticmethod
    def import_products(file, user):
        """
        Excel 파일에서 제품 데이터 import
        
        Args:
            file: UploadedFile object
            user: 등록할 사용자
        
        Returns:
            dict: {'success': bool, 'message': str, 'created': int, 'errors': list}
        """
        try:
            wb = load_workbook(file)
            ws = wb.active
            
            created_count = 0
            errors = []
            
            # 헤더 확인 (2번째 행부터 데이터)
            for row_num in range(2, ws.max_row + 1):
                try:
                    product_code = ws.cell(row=row_num, column=1).value
                    product_name = ws.cell(row=row_num, column=2).value
                    product_category = ws.cell(row=row_num, column=3).value
                    description = ws.cell(row=row_num, column=4).value
                    tags_str = ws.cell(row=row_num, column=5).value
                    is_raw_material_str = ws.cell(row=row_num, column=6).value
                    is_active_str = ws.cell(row=row_num, column=7).value
                    
                    # 필수 필드 확인
                    if not product_name:
                        errors.append(f'행 {row_num}: 제품명이 비어있습니다.')
                        continue
                    
                    # 태그 파싱
                    tags = []
                    if tags_str:
                        # '#태그1, #태그2' 형식을 ['태그1', '태그2']로 변환
                        tags = [tag.strip().replace('#', '') for tag in str(tags_str).split(',') if tag.strip()]
                    
                    # Boolean 변환
                    is_raw_material = str(is_raw_material_str).strip() == '가능'
                    is_active = str(is_active_str).strip() != '비활성'
                    
                    # Product import 로직 (실제 구현 시 모델 import 필요)
                    from v1.products.models import Product
                    
                    product, created = Product.objects.update_or_create(
                        product_code=product_code,
                        defaults={
                            'product_name': product_name,
                            'product_category': product_category,
                            'description': description,
                            'tags': tags,
                            'is_raw_material': is_raw_material,
                            'is_active': is_active,
                            'user': user,
                        }
                    )
                    
                    if created:
                        created_count += 1
                    
                except Exception as e:
                    errors.append(f'행 {row_num}: {str(e)}')
                    continue
            
            return {
                'success': True,
                'message': f'{created_count}개 제품이 성공적으로 등록되었습니다.',
                'created': created_count,
                'errors': errors
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f'파일 처리 중 오류 발생: {str(e)}',
                'created': 0,
                'errors': [str(e)]
            }
