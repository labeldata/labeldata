# ==================== 제품 관리 Views (V2) ====================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.urls import reverse
from django.contrib import messages
from v1.common.media_access import (
    downloadable_label_ids, user_can_download_label_files, visible_documents)
from django.http import JsonResponse
from django.conf import settings
from django.db.models import Q, Count, Prefetch, Sum, Case, When, IntegerField
from django.db import models
from django.core.paginator import Paginator
from django.views.decorators.http import require_GET, require_POST, require_http_methods
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.utils import timezone
from django.http import FileResponse, Http404, HttpResponse
from datetime import timedelta, datetime
from urllib.parse import quote
import zipfile
import io
import os
import json
import logging
import time as _time
import hmac

logger = logging.getLogger('django')

from .models import Product, ProductFolder, ProductAccessLog, ProductMetadata, FoodType, CountryList

# 앱 통합 뷰를 위한 추가 import
from v1.bom.models import ProductBOM
from .models import ProductDocument, ProductComment, ProductShare, SharedProductReceipt, DocumentType, DocumentSlot, SharePermission, ProductNotification, UserContact
from v1.label.models import MyLabel, FoodItem
from v1.label.services.label_naming import next_temp_label_name

from .forms import ProductForm
from v1.activity_log.utils import log_activity


# ==================== 공통 알림·이메일 헬퍼 ====================

def _send_email_safe(subject, body, to_email, from_email=None, attachment=None, html_body=None):
    """
    이메일 안전 발송 헬퍼.
    - html_body 제공 시 HTML + 텍스트 멀티파트로 발송
    - 발송 성공 여부(bool) 반환, 실패 시 Django logger에 기록
    """
    import logging
    from django.core.mail import EmailMultiAlternatives
    from django.conf import settings
    logger = logging.getLogger('django')

    _from = from_email or getattr(settings, 'EMAIL_FROM_DISPLAY', None) or settings.DEFAULT_FROM_EMAIL
    try:
        msg = EmailMultiAlternatives(subject, body, _from, [to_email])
        if html_body:
            msg.attach_alternative(html_body, 'text/html')
        if attachment:
            if isinstance(attachment, tuple):
                msg.attach(*attachment)
            else:
                try:
                    attachment.seek(0)
                    msg.attach(attachment.name, attachment.read(),
                                getattr(attachment, 'content_type', 'application/octet-stream'))
                except Exception:
                    pass
        msg.send(fail_silently=False)
        return True
    except Exception as e:
        logger.error(f'[Email] Send failed to {to_email}: {type(e).__name__}: {e}')
        return False


def _create_notification(label, recipient_user, message, status_code=''):
    """
    recipient_user가 실제 User 객체일 때만 인앱 알림 생성 (None-safe).
    """
    if recipient_user is None:
        return None
    return ProductNotification.objects.create(
        label=label,
        recipient=recipient_user,
        message=message,
        status_code=status_code,
    )


def _get_sender_info(user):
    """발신자 이름·회사명 반환. profile.company_name 없으면 빈 문자열."""
    name = user.get_full_name() or user.username
    try:
        company = (user.profile.company_name or '').strip()
    except Exception:
        company = ''
    return name, company


def _render_email(template_name, context):
    """
    HTML 이메일 렌더링 + 텍스트 폴백(strip_tags) 반환.
    returns (text_body, html_body)
    """
    from django.template.loader import render_to_string
    from django.utils.html import strip_tags
    from django.conf import settings as _cfg
    ctx = {
        'site_url': getattr(_cfg, 'SITE_URL', 'https://www.ezlabeling.com'),
    }
    ctx.update(context)
    html_body = render_to_string(template_name, ctx)
    text_body = strip_tags(html_body)
    return text_body, html_body




# ==================== Google Drive 스타일 탐색기 ====================

@login_required
def product_explorer(request, folder_id=None):
    """Google Drive 스타일 제품 탐색기"""
    user = request.user
    
    # 시스템 폴더 자동 생성 (최초 접속)
    ProductFolder.get_or_create_system_folders(user)
    
    # 현재 폴더
    current_folder = None
    if folder_id:
        current_folder = get_object_or_404(ProductFolder, folder_id=folder_id, owner=user)
    
    # 하위 폴더 목록
    subfolders = ProductFolder.objects.filter(
        owner=user,
        parent=current_folder
    ).order_by('sort_order', 'name')
    
    # ─── filter_type: 전체(ALL) / 내 제품(MINE) / 참여 중(COLLAB) ───
    filter_type = request.GET.get('filter', 'ALL')

    # 나에게 공유된 label_id 목록
    shared_to_me_ids = list(
        ProductShare.objects.filter(
            recipient_user=user,
            active_yn=True,
        ).filter(
            Q(share_end_date__isnull=True) | Q(share_end_date__gt=timezone.now())
        ).values_list('label_id', flat=True)
    )
    collab_count = len(set(shared_to_me_ids))

    # 현재 폴더의 제품 목록 (MyLabel 기반)
    if filter_type == 'MINE':
        labels = MyLabel.objects.filter(user_id=user, delete_YN='N')
    elif filter_type == 'COLLAB':
        labels = MyLabel.objects.filter(
            my_label_id__in=shared_to_me_ids, delete_YN='N'
        ).exclude(user_id=user)
    else:  # ALL
        labels = MyLabel.objects.filter(
            Q(user_id=user) | Q(my_label_id__in=shared_to_me_ids),
            delete_YN='N',
        ).distinct()
    labels = labels.order_by('-update_datetime')
    
    # 뷰 타입(grid/list)
    view_type = request.GET.get('view', 'grid')
    
    # 정렬
    sort_by = request.GET.get('sort', '-update_datetime')
    if sort_by in ['my_label_name', '-my_label_name', 'create_datetime', '-create_datetime', 
                   'update_datetime', '-update_datetime']:
        labels = labels.order_by(sort_by)
    
    # 검색
    search_query = request.GET.get('q', '')
    search_field = request.GET.get('search_field', 'all').strip() or 'all'
    _product_field_map = {
        'my_label_name': 'my_label_name',
        'prdlst_nm': 'prdlst_nm',
        'prdlst_dcnm': 'prdlst_dcnm',
        'prdlst_report_no': 'prdlst_report_no',
    }
    if search_query:
        if search_field != 'all' and search_field in _product_field_map:
            labels = labels.filter(
                Q(**{f"{_product_field_map[search_field]}__icontains": search_query})
            )
        else:
            labels = labels.filter(
                Q(my_label_name__icontains=search_query) |
                Q(prdlst_nm__icontains=search_query) |
                Q(prdlst_dcnm__icontains=search_query) |
                Q(prdlst_report_no__icontains=search_query) |
                Q(v2_metadata__search_tags__icontains=search_query)
            )

    # 원료 연결 필터: ingredient_id 파라미터가 있으면 해당 원료와 연결된 제품만 표시
    ingredient_id_filter = request.GET.get('ingredient_id')
    ingredient_name_filter = None
    if ingredient_id_filter:
        try:
            from v1.label.models import LabelIngredientRelation, MyIngredient
            linked_label_ids = LabelIngredientRelation.objects.filter(
                ingredient_id=ingredient_id_filter
            ).values_list('label_id', flat=True)
            labels = labels.filter(my_label_id__in=linked_label_ids)
            try:
                ingredient_name_filter = MyIngredient.objects.get(
                    my_ingredient_id=ingredient_id_filter
                ).prdlst_nm or f'원료 #{ingredient_id_filter}'
            except Exception:
                ingredient_name_filter = f'원료 #{ingredient_id_filter}'
        except Exception:
            ingredient_id_filter = None

    # 정확 일치 우선 정렬 (전체 검색 시)
    if search_query and search_field == 'all':
        from django.db.models import Case, When, Value, IntegerField
        labels = labels.annotate(
            _match_priority=Case(
                When(my_label_name__iexact=search_query, then=Value(0)),
                When(prdlst_nm__iexact=search_query, then=Value(0)),
                When(prdlst_report_no__iexact=search_query, then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            )
        ).order_by('_match_priority', '-update_datetime')

    # 페이지네이션 적용
    per_page = request.GET.get('per_page', '50')
    try:
        per_page = int(per_page)
        if per_page not in [25, 50, 100, 200]:
            per_page = 50
    except (ValueError, TypeError):
        per_page = 50
    
    paginator = Paginator(labels, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # 현재 페이지의 제품만 처리
    current_page_labels = page_obj.object_list
    
    # ProductMetadata에서 즐겨찾기 상태 가져오기
    products_data = []
    metadata_dict = {}
    
    # 한번에 모든 metadata 가져오기
    label_ids = [label.my_label_id for label in current_page_labels]
    metadatas = ProductMetadata.objects.filter(
        label__my_label_id__in=label_ids
    ).select_related('label')
    
    for meta in metadatas:
        metadata_dict[meta.label.my_label_id] = meta
    
    # 각 제품의 문서 개수 가져오기 (구분별 1개로 카운팅 - 여러 버전이 있어도 1개)
    doc_type_rows = (
        ProductDocument.objects.filter(
            label__my_label_id__in=label_ids,
            active_yn=True,
        )
        .values('label__my_label_id', 'document_type_id')
        .distinct()
    )
    document_counts = {}  # {label_id: distinct doc_type 수}
    for row in doc_type_rows:
        lid = row['label__my_label_id']
        document_counts[lid] = document_counts.get(lid, 0) + 1

    # 준수율 통계: DocumentSlot 기반 (문서함 탭과 동일한 로직)
    # total_slots = 숨겨지지 않은 슬롯 수, filled = status != EMPTY
    from .models import DocumentSlot
    slot_rows = (
        DocumentSlot.objects.filter(
            label__my_label_id__in=label_ids,
            hidden_yn=False,
        )
        .values('label__my_label_id', 'status')
    )
    slot_totals = {}   # {label_id: 전체 슬롯 수}
    slot_filled = {}   # {label_id: 채워진 슬롯 수}
    for row in slot_rows:
        lid = row['label__my_label_id']
        slot_totals[lid] = slot_totals.get(lid, 0) + 1
        if row['status'] != DocumentSlot.SlotStatus.EMPTY:
            slot_filled[lid] = slot_filled.get(lid, 0) + 1

    document_stats = {}
    for label_id in label_ids:
        doc_count = document_counts.get(label_id, 0)
        total = slot_totals.get(label_id, 0)
        filled = slot_filled.get(label_id, 0)
        document_stats[label_id] = {
            'total': doc_count,       # 등록 문서 수 (구분별 1개)
            'required': total,        # 전체 슬롯 수 (분모)
            'filled': filled,         # 채워진 슬롯 수 (분자)
            'rate': (filled / total * 100) if total > 0 else 0,
        }
    
    # BOM 통계 가져오기 (level=1인 직접 원료 행만 카운팅)
    bom_stats = {}
    from v1.bom.models import ProductBOM
    from django.db.models import Sum as BomSum, Count as BomCount

    bom_agg = (
        ProductBOM.objects.filter(
            parent_label__my_label_id__in=label_ids,
            level=1,
            active_yn=True,
        )
        .values('parent_label__my_label_id')
        .annotate(
            row_count=BomCount('bom_id'),
            total_ratio=BomSum('usage_ratio'),
        )
    )
    for row in bom_agg:
        lid = row['parent_label__my_label_id']
        ratio = float(row['total_ratio'] or 0)
        bom_stats[lid] = {
            'count': row['row_count'],
            'ratio': ratio,
            'complete': ratio >= 99.9,
        }
    # BOM 미등록 제품은 기본값
    for label_id in label_ids:
        if label_id not in bom_stats:
            bom_stats[label_id] = {'count': 0, 'ratio': 0, 'complete': False}
    
    # 영양성분 데이터 입력 여부 (dict 조회로 O(N²) → O(N))
    _label_obj_map = {l.my_label_id: l for l in current_page_labels}
    nutrition_stats = {}
    for label_id in label_ids:
        label_obj = _label_obj_map.get(label_id)
        nutrition_stats[label_id] = bool(label_obj and label_obj.calories)

    # 표시사항 체크: PDF 저장(LABEL_DESIGN 문서 존재) 또는 규정 검증(label_create_YN='Y') 완료 여부
    #
    # **판독에 쓴 사진은 빼야 한다.** 그 사진도 같은 문서 종류로 들어가지만
    # (사용자가 문서함의 '한글표시사항도안' 자리에서 찾기 때문), 사진을 한 장
    # 올린 것과 도안을 만든 것은 다른 일이다. 빼지 않으면 사진만 올려도 목록에
    # 표시사항 완료 표시가 켜진다.
    label_design_ids = set(
        ProductDocument.objects.filter(
            label__my_label_id__in=label_ids,
            document_type__type_code='LABEL_DESIGN',
            active_yn=True,
        ).exclude(metadata__source='ocr_import')
        .values_list('label__my_label_id', flat=True)
    )
    label_checked_stats = {}
    for label_id in label_ids:
        label_obj = _label_obj_map.get(label_id)
        pdf_saved = label_id in label_design_ids
        verified = label_obj.label_create_YN == 'Y' if label_obj else False
        label_checked_stats[label_id] = pdf_saved or verified
    
    # 권한 부여 인원 통계 (배치 쿼리 - 기존 N+1 방지)
    permission_stats = {}
    _perm_counts = (
        ProductShare.objects.filter(
            label__my_label_id__in=label_ids,
            active_yn=True,
        )
        .values('label__my_label_id', 'recipient_user')
        .distinct()
        .values('label__my_label_id')
        .annotate(_cnt=Count('share_id'))
    )
    for row in _perm_counts:
        permission_stats[row['label__my_label_id']] = row['_cnt']
    
    # 현재 페이지 라벨에서 내가 받은 역할 코드 조회
    my_role_map = {}  # label_id → role_code
    for share in ProductShare.objects.filter(
        recipient_user=user,
        label_id__in=label_ids,
        active_yn=True,
    ).select_related('permission'):
        try:
            my_role_map[share.label_id] = share.permission.role_code
        except Exception:
            my_role_map[share.label_id] = 'VIEWER'

    for label in current_page_labels:
        metadata = metadata_dict.get(label.my_label_id)
        is_owned = (label.user_id_id == user.id)
        my_role = None if is_owned else my_role_map.get(label.my_label_id, 'VIEWER')
        products_data.append({
            'label': label,
            'metadata': metadata,
            'starred_yn': metadata.starred_yn if metadata else False,
            'document_count': document_counts.get(label.my_label_id, 0),
            'document_stats': document_stats.get(label.my_label_id, {'total': 0, 'required': 0, 'filled': 0, 'rate': 0}),
            'bom_stats': bom_stats.get(label.my_label_id, {'count': 0, 'ratio': 0, 'complete': False}),
            'has_nutrition': nutrition_stats.get(label.my_label_id, False),
            'permission_count': permission_stats.get(label.my_label_id, 0),
            'is_owned': is_owned,
            'my_role': my_role,
            'label_checked': label_checked_stats.get(label.my_label_id, False),
        })
    
    # 브레드크럼 경로
    breadcrumb = []
    temp_folder = current_folder
    while temp_folder:
        breadcrumb.insert(0, temp_folder)
        temp_folder = temp_folder.parent
    
    # 접근 로그 (최근 열어본 제품에 표시)
    recent_product_ids = ProductAccessLog.objects.filter(
        user=user
    ).values_list('product_id', flat=True)[:50]
    
    # 즐겨찾기 총 개수 (현재 filter_type + 검색어 기준, 페이지네이션 무관)
    starred_count = ProductMetadata.objects.filter(
        label__in=labels,
        starred_yn=True,
    ).count()

    # 즐겨찾기 제품 (루트 폴더에서만, 사이드바용)
    starred_items = []
    if not current_folder:
        # ProductMetadata에서 즐겨찾기된 제품의 label 가져오기
        starred_metadata = ProductMetadata.objects.filter(
            label__in=labels,
            starred_yn=True
        ).select_related('label').order_by('-starred_datetime')[:10]
        starred_items = [meta.label for meta in starred_metadata]

    # 공유 문서함 요약 (최근 수신)
    shared_receipts = SharedProductReceipt.objects.filter(
        receiver=user
    ).select_related('share__label', 'share__created_by').order_by('-received_datetime')[:5]

    # 만료 예정 문서 요약 (30일 이내)
    today = timezone.now().date()
    alert_date = today + timedelta(days=30)
    expiring_documents = ProductDocument.objects.filter(
        label__user_id=user,
        active_yn=True,
        expiry_date__isnull=False,
        expiry_date__gte=today,
        expiry_date__lte=alert_date
    ).select_related('label', 'document_type').order_by('expiry_date')[:5]
    
    # 승인 대기 및 검토 필요 통계
    approval_pending_count = ProductMetadata.objects.filter(
        label__user_id=user,
        label__delete_YN='N',
        status=ProductMetadata.Status.REQUESTING
    ).count()
    
    review_needed_count = ProductMetadata.objects.filter(
        label__user_id=user,
        label__delete_YN='N',
        status=ProductMetadata.Status.REVIEW
    ).count()
    
    context = {
        'current_folder': current_folder,
        'subfolders': subfolders,
        'products_data': products_data,
        'breadcrumb': breadcrumb,
        'view_type': view_type,
        'search_query': search_query,
        'search_field': search_field,
        'recent_product_ids': list(recent_product_ids),
        'starred_items': starred_items,
        'shared_receipts': shared_receipts,
        'expiring_documents': expiring_documents,
        'recent_products': products_data[:4],
        'folders': subfolders,
        'files': products_data,
        'approval_pending_count': approval_pending_count,
        'review_needed_count': review_needed_count,
        'page_obj': page_obj,
        'per_page': per_page,
        'total_products_count': paginator.count,
        'filter_type': filter_type,
        'collab_count': collab_count,
        'starred_count': starred_count,
        'ingredient_id_filter': ingredient_id_filter,
        'ingredient_name_filter': ingredient_name_filter,
    }
    return render(request, 'products/product_explorer.html', context)


# ==================== 제품 상세 ====================

def _version_root(doc, by_id):
    """
    이 판이 어느 문서의 판인가. **끝까지 거슬러 올라간다.**

    새 판은 바로 앞 판을 parent_document 로 가리킨다. 뿌리를 가리키는 것이
    아니라 **줄로 이어지는** 구조다.

        v1 <- v2 <- v3 <- v4

    한 칸만 보고 묶으면 v1·v2 만 한 묶음이 되고 v3 과 v4 는 저마다 새 문서로
    선다. 실제로 그렇게 나왔다 - 다섯 판짜리 도안이 목록에 네 줄로 있었다.

    사라진 조상을 만나면 그 id 를 뿌리로 삼는다. 형제들이 같은 조상을 가리키고
    있으면 그 조상이 목록에 없어도 한 묶음이어야 한다.
    """
    seen = set()
    node = doc
    while True:
        parent_id = node.parent_document_id
        if not parent_id or parent_id in seen:
            return node.document_id
        seen.add(parent_id)
        nxt = by_id.get(parent_id)
        if nxt is None:
            return parent_id
        node = nxt


def _mend_broken_chains(docs, roots):
    """
    사슬이 끊긴 판을 판 번호로 이어 붙인다. **roots 를 제자리에서 고친다.**

    판 번호는 (제품, 문서 종류) 안에서 **하나의 줄**로 매겨진다 - 올리는 쪽이
    언제나 "그 종류의 가장 높은 판 + 1" 로 정한다. 입구가 셋인데(직접 업로드,
    표시사항 사진 저장, 시안 대조) 셋 다 그 규칙을 쓴다.

    그러니 **판 2 이상인데 위로 이어진 데가 없는 것**은 홀로 선 문서가 아니라
    사슬이 끊긴 판이다. 그 종류의 가장 낮은 판에 이어 붙인다. 옛 자료나, 이어
    달기 전에 들어온 것이 그렇게 남아 있다.

    판 1 은 건드리지 않는다. 그것은 정말 새 문서일 수 있다.
    """
    lowest = {}
    for doc in sorted(docs, key=lambda d: ((d.version or 1), d.document_id)):
        lowest.setdefault(doc.document_type_id, roots[doc.document_id])

    for doc in docs:
        if doc.parent_document_id is None and (doc.version or 1) > 1:
            roots[doc.document_id] = lowest.get(doc.document_type_id,
                                                roots[doc.document_id])

    # 뿌리가 다시 이어졌으면 그 끝까지 따라간다
    for doc in docs:
        seen, node = set(), roots[doc.document_id]
        while node in roots and roots[node] != node and node not in seen:
            seen.add(node)
            node = roots[node]
        roots[doc.document_id] = node


def version_stacks(documents):
    """
    같은 문서의 여러 판을 한 줄로 묶는다.

    새 판을 올려도 예전 판은 지워지지 않고 parent_document 로 이어 달린다.
    그런데 목록은 그 판들을 **각각 한 줄씩** 그렸다. 문서 세 종류를 가진
    제품이 여섯 줄로 보이고 같은 이름이 연달아 세 번 나온다 - 정작 알고 싶은
    "무슨 문서를 갖고 있는가" 가 판 수에 묻힌다.

    한 줄에 최신 판을 놓고 예전 판은 그 아래에 접어 둔다. 예전 판을 감추는
    것이 아니다 - 문서 하나에 이력이 쌓이는 것이 원래 모습이고, 목록은 문서를
    세는 자리다.

    documents 의 순서를 지킨다. 최신 판이 나왔을 자리에 그 묶음이 선다.

    Returns: [{'root', 'latest', 'older', 'count'}…]
    """
    docs = list(documents)
    by_id = {d.document_id: d for d in docs}
    roots = {d.document_id: _version_root(d, by_id) for d in docs}
    _mend_broken_chains(docs, roots)

    order, stacks = [], {}
    for doc in docs:
        root = roots[doc.document_id]
        if root not in stacks:
            stacks[root] = []
            order.append(root)
        stacks[root].append(doc)

    groups = []
    for root in order:
        # 판 번호가 같거나 없는 옛 자료가 있어 올린 시각까지 함께 본다.
        docs = sorted(stacks[root],
                      key=lambda d: (d.version or 1, d.uploaded_datetime or 0),
                      reverse=True)
        groups.append({'root': root, 'latest': docs[0],
                       'older': docs[1:], 'count': len(docs)})
    return groups


@login_required
def product_detail(request, product_id):
    """제품 상세 보기 - V2 스타일 (BOM, 문서 등록 등)"""
    # V2에서는 label_id를 product_id로 받음
    # 먼저 직접 MyLabel 조회 시도 (오너)
    shared_share = None
    is_owner = False
    try:
        label = MyLabel.objects.get(my_label_id=product_id, user_id=request.user)
        is_owner = True
    except MyLabel.DoesNotExist:
        # 공유 사용자 접근 허용
        shared_share = ProductShare.objects.filter(
            label__my_label_id=product_id,
            active_yn=True
        ).filter(
            Q(recipient_user=request.user) | Q(recipient_email__iexact=request.user.email)
        ).filter(
            Q(share_end_date__isnull=True) | Q(share_end_date__gt=timezone.now())
        ).select_related('label').first()

        if shared_share:
            label = shared_share.label
        else:
            # MyLabel이 없으면 ProductMetadata를 통해 조회 시도 (오너만 가능)
            try:
                metadata = ProductMetadata.objects.select_related('label').get(
                    metadata_id=product_id,
                    label__user_id=request.user
                )
                label = metadata.label
                is_owner = True
            except ProductMetadata.DoesNotExist:
                # 둘 다 없으면 404
                messages.error(request, '제품을 찾을 수 없습니다.')
                return redirect('products:product_explorer')
    
    # ProductMetadata 자동 생성 (없는 경우 - 오너만 가능)
    try:
        metadata = ProductMetadata.objects.get(label=label)
    except ProductMetadata.DoesNotExist:
        if not is_owner:
            messages.error(request, '제품 메타데이터를 찾을 수 없습니다.')
            return redirect('products:product_explorer')

        # 고유한 product_code 생성
        user_product_count = ProductMetadata.objects.filter(
            label__user_id=request.user
        ).count()

        # 중복되지 않는 고유 코드 찾기
        product_code = None
        max_attempts = 100
        for i in range(max_attempts):
            candidate_code = f"PRD-{request.user.id}-{user_product_count + i + 1:04d}"
            if not ProductMetadata.objects.filter(product_code=candidate_code).exists():
                product_code = candidate_code
                break

        # 고유 코드를 찾지 못한 경우 타임스탬프 추가
        if not product_code:
            import time
            product_code = f"PRD-{request.user.id}-{int(time.time())}"

        # ProductMetadata 생성
        metadata = ProductMetadata.objects.create(
            label=label,
            product_code=product_code,
            starred_yn=False
        )
    
    # BOM 정보 조회
    bom_items = ProductBOM.objects.filter(
        parent_label=label,
        active_yn=True
    ).select_related('child_label', 'shared_receipt').order_by('level', 'sort_order')
    
    # 문서 정보 조회
    documents = ProductDocument.objects.filter(
        label=label,
        active_yn=True
    ).select_related('document_type', 'uploaded_by', 'uploaded_by__profile').order_by('document_type__display_order', '-uploaded_datetime')
    
    # 남의 서류가 목록에 보이면 파일을 못 받아도 거래처와 서류 종류가 드러난다
    documents = visible_documents(request.user, label, documents)

    # 같은 문서의 여러 판을 한 줄로
    document_groups = version_stacks(documents)

    # 문서 타입별 그룹화
    document_types = DocumentType.objects.filter(active_yn=True).order_by('display_order', 'type_name')
    documents_by_type = {}
    for dtype in document_types:
        docs = documents.filter(document_type=dtype)
        if docs.exists():
            documents_by_type[dtype] = docs
    
    # 타입 미지정 문서
    untyped_docs = documents.filter(document_type__isnull=True)
    if untyped_docs.exists():
        documents_by_type['기타'] = untyped_docs
    
    # 공유 정보 조회 (현재 제품)
    shares = ProductShare.objects.filter(
        label=label,
        active_yn=True
    ).select_related('recipient_user').order_by('-created_datetime')
    share_permissions = SharePermission.objects.filter(share__in=shares)
    permission_map = {permission.share_id: permission for permission in share_permissions}
    for share in shares:
        share.permission_record = permission_map.get(share.share_id)
        share.display_name = (
            share.recipient_name
            or (share.recipient_user.username if share.recipient_user else None)
            or share.recipient_email
            or '미가입'
        )

    # 각 역할 담당자 존재 여부 (추가 DB 쿼리 없이 이미 조회된 permission_map 활용)
    has_uploader = any(p.role_code == 'UPLOADER' for p in permission_map.values())
    has_reviewer = any(p.role_code == 'REVIEWER' for p in permission_map.values())
    has_approver = any(p.role_code == 'APPROVER' for p in permission_map.values())
    
    # 내가 owner인 모든 제품의 공유자 목록 조회 (팔레트용)
    all_shared_users = []
    if is_owner:
        # 현재 사용자가 owner인 모든 제품 찾기
        my_labels = MyLabel.objects.filter(user_id=request.user)
        
        # 이 제품들에 대한 모든 공유 정보
        all_shares = ProductShare.objects.filter(
            label__in=my_labels,
            active_yn=True
        ).select_related('recipient_user').order_by('-created_datetime')
        
        # 중복 제거를 위해 email 기준으로 unique한 사용자만
        seen_emails = set()
        for share in all_shares:
            if share.recipient_email not in seen_emails:
                seen_emails.add(share.recipient_email)
                
                # 현재 제품에서의 권한 정보 확인
                current_share = shares.filter(recipient_email=share.recipient_email).first()
                if current_share:
                    share.permission_record = permission_map.get(current_share.share_id)
                    share.share_id = current_share.share_id
                else:
                    share.permission_record = None
                    share.share_id = None
                
                all_shared_users.append(share)
                # 템플릿에서 None.username 조회 방지: display_name 미리 계산
                share.display_name = (
                    share.recipient_name
                    or (share.recipient_user.username if share.recipient_user else None)
                    or share.recipient_email
                    or '미가입'
                )
    elif shared_share:
        # EDITOR 등 공유받은 사용자: 현재 제품의 공유자만 팔레트에 표시
        for share in shares:
            if not hasattr(share, 'display_name'):
                share.display_name = (
                    share.recipient_name
                    or (share.recipient_user.username if share.recipient_user else None)
                    or share.recipient_email
                    or '미가입'
                )
            all_shared_users.append(share)

    # 상태 기반 권한 계산 (역할 기반)
    status = metadata.status

    # ── 단방향 전이 맵 (각 상태에서 이동 가능한 action 키) ──
    # action 키 → Status 값은 product_update_status 뷰에서 정의
    status_actions_map = {
        ProductMetadata.Status.DRAFT:      ['requesting'],
        ProductMetadata.Status.REQUESTING: ['submitted', 'draft'],
        ProductMetadata.Status.SUBMITTED:  ['review', 'requesting'],
        ProductMetadata.Status.REVIEW:     ['pending', 'submitted'],
        ProductMetadata.Status.PENDING:    ['confirmed', 'review'],
        ProductMetadata.Status.CONFIRMED:  ['draft'],   # → 새 버전으로 돌아감
    }

    user_role = 'OWNER' if is_owner else 'VIEWER'
    if shared_share and not is_owner:
        share_permission = SharePermission.objects.filter(share=shared_share).first()
        if share_permission:
            user_role = share_permission.role_code

    role_labels = {
        'OWNER':    '관리자',
        'UPLOADER': '자료 제출',
        'EDITOR':   '공동 작성',
        'REVIEWER': '검토/QA',
        'APPROVER': '최종 승인',
        'VIEWER':   '뷰어',
    }

    available_actions = []
    can_edit = False
    can_upload_documents = False
    can_comment = False
    can_delete_product = False

    # ── 상태별 편집 가능 역할 ──
    # DRAFT           : OWNER, EDITOR
    # REQUESTING      : OWNER, EDITOR, UPLOADER
    # SUBMITTED/REVIEW: OWNER, EDITOR, REVIEWER
    # PENDING         : OWNER, EDITOR, APPROVER
    # CONFIRMED       : OWNER, EDITOR (다른 역할은 수정 불가)

    edit_roles_by_status = {
        ProductMetadata.Status.DRAFT:      {'OWNER', 'EDITOR'},
        ProductMetadata.Status.REQUESTING: {'OWNER', 'EDITOR', 'UPLOADER'},
        ProductMetadata.Status.SUBMITTED:  {'OWNER', 'EDITOR'},
        ProductMetadata.Status.REVIEW:     {'OWNER', 'EDITOR'},
        ProductMetadata.Status.PENDING:    {'OWNER', 'EDITOR', 'APPROVER'},
        ProductMetadata.Status.CONFIRMED:  {'OWNER', 'EDITOR'},
    }
    can_edit = user_role in edit_roles_by_status.get(status, set())

    def _skip_aware_actions(cur_status):
        """담당자 미지정 단계를 건너뛰는 available_actions를 반환합니다."""
        if cur_status == ProductMetadata.Status.DRAFT:
            if has_uploader:
                return ['requesting']
            elif has_reviewer:
                return ['review']       # REQUESTING/SUBMITTED 건너뜀
            elif has_approver:
                return ['pending']      # REQUESTING~REVIEW 건너뜀
            else:
                return ['confirmed']    # 전 단계 건너뜀
        elif cur_status == ProductMetadata.Status.SUBMITTED:
            if has_reviewer:
                return ['review', 'requesting']
            elif has_approver:
                return ['pending', 'requesting']    # REVIEW 건너뜀
            else:
                return ['confirmed', 'requesting']  # REVIEW+PENDING 건너뜀
        return status_actions_map.get(cur_status, [])

    if user_role == 'OWNER':
        can_upload_documents = True
        can_comment = True
        can_delete_product = True
        available_actions = _skip_aware_actions(status)
    elif user_role == 'EDITOR':
        can_upload_documents = True
        can_comment = True
        available_actions = _skip_aware_actions(status)
    elif user_role == 'UPLOADER':
        can_upload_documents = True
        can_comment = True
        if status == ProductMetadata.Status.REQUESTING:
            available_actions = ['submitted']
    elif user_role == 'REVIEWER':
        can_comment = True
        if status == ProductMetadata.Status.REVIEW:
            available_actions = ['pending']
    elif user_role == 'APPROVER':
        can_comment = True
        if status == ProductMetadata.Status.PENDING:
            available_actions = ['confirmed']
    else:  # VIEWER
        can_comment = True
    
    # 댓글 정보 조회
    comments = ProductComment.objects.filter(
        label=label,
        parent__isnull=True
    ).select_related('author').prefetch_related(
        Prefetch('replies', queryset=ProductComment.objects.select_related('author').order_by('created_at'))
    ).order_by('-created_at')

    # 댓글 작성자들의 역할 정보 조회
    author_ids = set()
    for comment in comments:
        if comment.author_id:
            author_ids.add(comment.author_id)
        for reply in comment.replies.all():
            if reply.author_id:
                author_ids.add(reply.author_id)

    author_roles = {}
    if author_ids:
        owner_user_id = label.user_id.id if label.user_id else None
        
        shared_users_permissions = SharePermission.objects.filter(
            share__label=label,
            share__recipient_user_id__in=author_ids,
            share__active_yn=True
        ).select_related('share')

        share_roles = {
            p.share.recipient_user_id: role_labels.get(p.role_code, '참여자')
            for p in shared_users_permissions
        }

        for author_id in author_ids:
            if author_id == owner_user_id:
                author_roles[author_id] = role_labels.get('OWNER', '관리자')
            else:
                author_roles[author_id] = share_roles.get(author_id, '참여자')

    # 필드별 댓글 존재 여부
    comment_fields = set(comments.values_list('field_name', flat=True))
    
    # 활동 로그 조회
    from .models import ProductActivityLog
    activity_logs = ProductActivityLog.objects.filter(
        label=label
    ).select_related('user').order_by('-created_at')[:50]  # 최근 50개
    
    # 활동 로그 작성자들의 권한 정보 조회
    activity_author_ids = set()
    for log in activity_logs:
        if log.user_id:
            activity_author_ids.add(log.user_id)
    
    activity_author_roles = {}
    if activity_author_ids:
        owner_user_id = label.user_id.id if label.user_id else None
        
        activity_shared_permissions = SharePermission.objects.filter(
            share__label=label,
            share__recipient_user_id__in=activity_author_ids,
            share__active_yn=True
        ).select_related('share')
        
        activity_share_roles = {
            p.share.recipient_user_id: role_labels.get(p.role_code, '참여자')
            for p in activity_shared_permissions
        }
        
        for author_id in activity_author_ids:
            if author_id == owner_user_id:
                activity_author_roles[author_id] = role_labels.get('OWNER', '관리자')
            else:
                activity_author_roles[author_id] = activity_share_roles.get(author_id, '참여자')
    
    # 식품유형과 원산지 목록 추가
    food_types = FoodType.objects.all().order_by('food_group', 'food_type')
    food_groups = FoodType.objects.values_list('food_group', flat=True).distinct().order_by('food_group')
    countries = CountryList.objects.all().order_by('country_name_ko')
    
    # 문서 슬롯 정보 조회
    from .models import DocumentSlot

    # 전체 슬롯 타입 (숨김 포함) — 중복 생성 방지용
    all_slot_type_ids = set(
        DocumentSlot.objects.filter(label=label).values_list('document_type_id', flat=True)
    )
    # 보이는 슬롯 타입 (숨김 제외) — "필수 문서 추가" 드롭다운용
    visible_slot_type_ids = set(
        DocumentSlot.objects.filter(label=label, hidden_yn=False).values_list('document_type_id', flat=True)
    )

    # 필수 문서 슬롯 자동 생성 (최초 접근 시)
    if is_owner:  # 오너만 슬롯 생성 가능
        
        # 모든 필수 문서 타입에 대해 슬롯 생성 (전체 목록 기준 — 숨김 포함하여 중복 방지)
        required_types = DocumentType.objects.filter(required_yn=True, active_yn=True)
        slots_to_create = []
        
        for doc_type in required_types:
            if doc_type.type_id not in all_slot_type_ids:
                slots_to_create.append(
                    DocumentSlot(
                        label=label,
                        document_type=doc_type
                        # status는 default 값(EMPTY) 사용
                    )
                )
        
        if slots_to_create:
            DocumentSlot.objects.bulk_create(slots_to_create)
    
    document_slots = DocumentSlot.objects.filter(
        label=label,
        hidden_yn=False  # 숨겨지지 않은 슬롯만
    ).select_related('document_type', 'current_document').order_by('document_type__display_order')
    
    available_doc_types = DocumentType.objects.filter(
        active_yn=True
    ).exclude(type_id__in=visible_slot_type_ids).order_by('display_order', 'type_name')
    
    # 슬롯 상태 업데이트 (만료일 기준)
    for slot in document_slots:
        slot.update_status()
    
    # 슬롯 통계 계산 (숨겨지지 않은 슬롯만)
    total_slots = document_slots.count()
    filled_slots = document_slots.exclude(status=DocumentSlot.SlotStatus.EMPTY).count()
    compliance_rate = (filled_slots / total_slots * 100) if total_slots > 0 else 0
    empty_count = document_slots.filter(status=DocumentSlot.SlotStatus.EMPTY).count()
    expiring_count = document_slots.filter(status=DocumentSlot.SlotStatus.EXPIRING).count()
    expired_count = document_slots.filter(status=DocumentSlot.SlotStatus.EXPIRED).count()
    
    context = {
        'product': label,  # 템플릿에서 product로 참조
        'label': label,
        'latest_version': label,  # V1/V2 호환성을 위해 추가
        'metadata': metadata,
        'bom_items': bom_items,
        'documents': documents,
        'document_groups': document_groups,
        'food_types': food_types,
        'food_groups': food_groups,
        'countries': countries,
        'documents_by_type': documents_by_type,
        'document_types': document_types,
        'shares': shares,
        'all_shared_users': all_shared_users,  # 모든 공유자 목록
        'product_status': status,
        'status_choices': ProductMetadata.Status.choices,
        'can_edit': can_edit,
        'can_upload_documents': can_upload_documents,
        'can_comment': can_comment,
        'can_delete_product': can_delete_product,
        'available_actions': available_actions,
        'user_role': user_role,
        'user_role_label': role_labels.get(user_role, '뷰어'),
        'label_owner': label.user_id,  # 실제 라벨 소유자 (EDITOR 팔레트용)
        'has_uploader': has_uploader,
        'has_reviewer': has_reviewer,
        'has_approver': has_approver,
        'comments': comments,
        'comment_fields': comment_fields,
        'author_roles': author_roles,
        'activity_logs': activity_logs,  # 활동 로그 추가
        'activity_author_roles': activity_author_roles,  # 활동 로그 작성자 권한
        'today': timezone.now().date(),
        'warning_date': timezone.now().date() + timedelta(days=30),  # 30일 이내 만료 경고
        # 문서 슬롯 정보
        'document_slots': document_slots,
        'total_slots': total_slots,
        'filled_slots': filled_slots,
        'compliance_rate': compliance_rate,
        'empty_count': empty_count,
        'expiring_count': expiring_count,
        'expired_count': expired_count,
        'available_doc_types': available_doc_types,
        'from_source': request.GET.get('from', ''),
        'custom_fields_json': json.dumps(label.custom_fields or [], ensure_ascii=False),
        # 표시 항목(chckd_*). 지금까지 이 화면에는 이걸 볼 수도 바꿀 수도 없었는데,
        # 필수 입력 검사가 chckd_* 를 근거로 삼으면서 "해당하지 않으면 체크를
        # 해제하세요" 라는 안내를 따를 방법이 없었다.
        'display_items': _build_display_items(label),
        'workflow_steps': _build_workflow_steps(label),
        'preservation_choices': PRESERVATION_CHOICES,
        'processing_choices': PROCESSING_CHOICES,
    }

    return render(request, 'products/product_detail.html', context)


# 장기보존식품·제조방법 선택지.
#
# 지금까지 템플릿이 preservation_choices 로 돌리고 {% empty %} 에 같은 목록을
# 손으로 또 적어 뒀는데, 그 변수를 넘기는 뷰가 없어서 **항상 폴백만** 그려졌다.
# 값(value)은 저장되는 문자열이라 바꾸면 기존 데이터와 어긋난다.
PRESERVATION_CHOICES = [
    ('frozen_heated',    '냉동(가열)'),
    ('frozen_nonheated', '냉동(비가열)'),
    ('canned',           '통·병조림'),
    ('retort',           '레토르트'),
]
PROCESSING_CHOICES = [
    ('sanitized',   '살균'),
    ('aseptic',     '멸균'),
    ('yutang',      '유탕/유처리'),
    ('unsanitized', '비살균'),
]


# 오른쪽 패널에 뿌릴 순서. 인쇄되는 순서와 대체로 맞춘다.
#
# chckd_weight_calorie(내용량(열량))는 뺐다. 별도로 입력하는 칸이 아니라 내용량에
# 병기하는 값이라("250 g (100 kcal)") 켜고 끌 대상이 아니다. 표시 여부는 식품유형이
# 정하고, 값이 적혔는지는 validation_service 가 내용량의 kcal 표기로 판정한다.
#
# 순서는 기본정보 탭의 입력칸 순서와 같게 맞춘다. 목록이 곧 목차 역할을 하므로
# 화면을 훑는 순서와 어긋나면 찾기 어려워진다.
#
# 영양성분만 예외로 맨 끝이다 — 이 탭에 칸이 없어 누르면 다른 탭으로 넘어가고,
# 흐름이 끊기는 항목이라 입력을 다 마친 뒤 보이는 게 낫다.
_DISPLAY_ITEM_ORDER = [
    'chckd_prdlst_nm', 'chckd_ingredient_info', 'chckd_prdlst_dcnm',
    'chckd_prdlst_report_no', 'chckd_content_weight', 'chckd_country_of_origin',
    'chckd_storage_method', 'chckd_frmlc_mtrqlt',
    'chckd_bssh_nm', 'chckd_distributor_address', 'chckd_repacker_address',
    'chckd_importer_address', 'chckd_pog_daycnt', 'chckd_rawmtrl_nm_display',
    'chckd_cautions', 'chckd_additional_info',
    'chckd_nutrition_text',
]

# 항목 이름을 눌렀을 때 갈 곳. 기본은 'field-<필드명을 하이픈으로>' 인데 둘이 다르다.
_DISPLAY_ITEM_ANCHORS = {
    # 칸의 id 는 예전 이름 그대로다(참조하는 JS 가 여럿). 값은 표시 필드를 쓴다.
    'rawmtrl_nm_display': 'field-rawmtrl-nm',
}
# 이 탭에 칸이 없어 다른 탭으로 보내야 하는 항목. (탭 id, 탭 이름)
_DISPLAY_ITEM_TABS = {
    'nutrition_text': ('tab-nutrition', '영양성분'),
}


# 모델에 있는 표시 항목 체크박스 전부.
#
# 저장이 반영할 대상을 여기서 한 번만 정한다. 예전에는 식품유형 규칙표를
# 대신 썼는데 그 표에 없는 넷(유통전문판매원·소분원·수입원·기타표시사항)이
# 조용히 빠졌다. 모델을 근거로 삼으면 칸이 늘어도 같은 일이 되풀이되지 않는다.
_CHECKBOX_FIELDS = tuple(
    f.name for f in MyLabel._meta.get_fields()
    if getattr(f, 'name', '').startswith('chckd_')
)


def _checkbox_label(checkbox):
    """'chckd_nutrition_text' -> '영양성분'. 화면에 그대로 내보내는 이름."""
    field = checkbox[len('chckd_'):]
    try:
        return str(MyLabel._meta.get_field(field).verbose_name)
    except Exception:
        return field


def _display_item_sources(field):
    """
    그 항목이 "채워졌다" 를 판정할 때 볼 입력칸 id 전부.

    화면은 저장하기 전에도 미입력 표시를 갱신해야 한다. 서버가 보는 자리
    (validation_service.content_sources)와 같은 목록을 그대로 넘겨 주지 않으면,
    주의사항/기타표시사항처럼 한쪽에만 적어도 되는 항목에서 화면과 검증이
    서로 다른 말을 하게 된다.

    이 탭에 칸이 없는 항목(영양성분)은 존재하지 않는 id 만 나오는데, 그때는
    화면이 서버가 계산해 준 값을 그대로 쓴다.
    """
    from v1.label.services.validation_service import content_sources

    ids = []
    for src in content_sources(field):
        element_id = _DISPLAY_ITEM_ANCHORS.get(src, 'field-' + src.replace('_', '-'))
        if element_id not in ids:
            ids.append(element_id)
    return ids


# 제품 화면의 탭은 **업무 순서**다. 그런데 화면에는 나란한 네 개로만 보여서,
# 처음 쓰는 사람은 어디부터 손대야 하는지 알 수 없었다. 순서와 뜻을 여기 적어
# 화면이 번호·화살표로 드러내게 한다.
#
# 문서함·권한 설정은 단계가 아니라 곁에 두는 것이라 여기 없다.
_WORKFLOW_STEPS = (
    ('tab-info', '기본 정보',
     '제품명·내용량·보관방법처럼 라벨에 인쇄될 값을 채웁니다.'),
    ('tab-bom', 'BOM',
     '원료와 배합비를 넣습니다. 원재료명과 알레르기 표시가 여기서 나옵니다.'),
    ('tab-nutrition', '영양성분',
     '영양성분표의 값을 넣습니다. 표시사항의 영양정보 표로 함께 그려집니다.'),
    ('tab-label', '표시사항',
     '앞의 셋을 모아 표로 그립니다. 규정 검증·내보내기·시안 대조를 여기서 합니다.'),
)


def _build_workflow_steps(label):
    """
    각 단계를 마쳤는가.

    "마쳤다" 를 엄격하게 보지 않는다 — 이 표시는 **어디까지 왔는지 눈으로 알게
    하려는 것**이지 판정이 아니다. 판정은 규정 검증이 한다. 여기서 빡빡하게
    굴면 멀쩡히 진행 중인 제품이 계속 "안 함" 으로 보여 오히려 헷갈린다.
    """
    from v1.bom.models import ProductBOM

    done = {
        'tab-info': bool((label.prdlst_nm or '').strip()
                         and (label.content_weight or '').strip()),
        'tab-bom': ProductBOM.objects.filter(parent_label=label).exists(),
        'tab-nutrition': bool((label.calories or '').strip()),
        'tab-label': bool((label.rawmtrl_nm_display or label.rawmtrl_nm or '').strip()),
    }

    return [
        {'no': index, 'tab': tab, 'name': name, 'hint': hint, 'done': done.get(tab, False)}
        for index, (tab, name, hint) in enumerate(_WORKFLOW_STEPS, start=1)
    ]


def _build_display_items(label):
    """
    표시 항목 체크박스 목록. 식품유형이 정하는 값(Y/D/N)을 함께 실어 보내
    화면이 "이 유형에서는 필수" / "해당 없음" 을 표시할 수 있게 한다.
    """
    from v1.label.services import food_type_settings as fts
    from v1.label.services.validation_service import _has_content

    try:
        rule = fts.resolve_settings(label.food_group or '', label.food_type or '')
    except Exception:
        logger.exception('[표시 항목] 식품유형 규칙 조회 실패 (label=%s)', label.pk)
        rule = {'settings': {}, 'found': False}

    by_checkbox = {fts.FIELD_TO_CHECKBOX[f]: v
                   for f, v in rule['settings'].items()
                   if f in fts.FIELD_TO_CHECKBOX}

    items = []
    for checkbox in _DISPLAY_ITEM_ORDER:
        field = checkbox[len('chckd_'):]
        items.append({
            'checkbox': checkbox,
            'field': field,
            'label': _checkbox_label(checkbox),
            'checked': (getattr(label, checkbox, '') or '') == 'Y',
            'rule': by_checkbox.get(checkbox, ''),   # 'Y' 필수 / 'D' 해당없음 / '' 규칙없음
            # 다른 탭이 채우는 자리까지 본다 — 영양성분은 영양성분 탭이,
            # 원재료명(표시)은 BOM/기본정보의 rawmtrl_nm 이 채운다.
            'filled': _has_content(label, field),
            # 화면이 저장 전에도 같은 판정을 할 수 있게 볼 자리를 알려 준다
            'sources': ','.join(_display_item_sources(field)),
            'anchor': _DISPLAY_ITEM_ANCHORS.get(
                field, 'field-' + field.replace('_', '-')),
            'tab': _DISPLAY_ITEM_TABS.get(field, ('', ''))[0],
            'tab_label': _DISPLAY_ITEM_TABS.get(field, ('', ''))[1],
        })
    return items


# ==================== 최근 항목 / 즐겨찾기 ====================

@login_required
def product_recent(request):
    """최근 열어본 항목"""
    recent_logs = ProductAccessLog.objects.filter(
        user=request.user,
        product__delete_YN='N'
    ).select_related('product')[:50]
    
    context = {
        'recent_logs': recent_logs,
        'title': '최근 항목'
    }
    return render(request, 'products/product_recent.html', context)


@login_required
def product_favorite(request):
    """즐겨찾기 항목"""
    # ProductMetadata에서 즐겨찾기된 제품 가져오기
    starred_metadata = ProductMetadata.objects.filter(
        label__user_id=request.user,
        label__delete_YN='N',
        starred_yn=True
    ).select_related('label').order_by('-starred_datetime')

    products = []
    for meta in starred_metadata:
        label = meta.label
        products.append({
            'product_id': label.my_label_id,
            'product_name': label.my_label_name or label.prdlst_nm,
            'product_code': meta.product_code,
            'description': label.prdlst_dcnm,
            'starred_datetime': meta.starred_datetime,
        })

    context = {
        'products': products,
        'title': '즐겨찾기',
        'is_favorite_view': True,
    }
    return render(request, 'products/product_starred.html', context)


# ==================== 제품 생성/수정/삭제 ====================

@login_required
def product_create(request):
    """
    새 제품을 즉시 만들고 워크스페이스로 보낸다.

    예전에는 별도의 등록 폼(product_form.html)을 먼저 보여 주고, 저장해야 제품이
    생겼다. 그래서 화면이 두 벌이 됐고 - 등록 폼과 제품 상세가 생김새도 저장
    방식도 달랐다 - 등록 폼에서는 사진 불러오기·BOM·문서함이 아무것도 되지
    않았다. 붙일 제품이 아직 없었기 때문이다.

    표시사항 작성(label:create_new_label)은 원래 이 방식이었다. 제품 관리만
    달라서 두 벌이 된 것이라, 그쪽에 맞춘다.

    빈 제품이 쌓이는 것이 이 방식의 대가다. 손대지 않은 것은
    `manage.py cleanup_temp_labels` 가 치운다(지우지 않고 숨김 처리).
    """
    label = MyLabel.objects.create(
        user_id=request.user,
        my_label_name=next_temp_label_name(request.user),
        delete_YN='N',
    )

    # 제품 코드는 사용자 안에서 겹치지 않으면 된다. 번호가 비어 있어도(지운 제품)
    # 다시 쓰지 않고 뒤로 민다 - 옛 제품의 코드가 되살아나면 이력이 헷갈린다.
    used = set(
        ProductMetadata.objects
        .filter(label__user_id=request.user)
        .values_list('product_code', flat=True)
    )
    seq = len(used) + 1
    while f'PRD-{request.user.id}-{seq:04d}' in used:
        seq += 1
    product_code = f'PRD-{request.user.id}-{seq:04d}'

    ProductMetadata.objects.get_or_create(
        label=label,
        defaults={'product_code': product_code, 'starred_yn': False},
    )

    from .models import ProductActivityLog
    ProductActivityLog.objects.create(
        label=label,
        user=request.user,
        action='CREATED',
        details={'product_code': product_code, 'product_name': label.my_label_name},
    )

    log_activity(request, 'product', 'product_create', label.my_label_id)

    # 홈의 "사진으로 시작" 처럼 곧바로 사진을 읽으러 온 경우, 워크스페이스가
    # 열리자마자 불러오기 창을 띄운다. 이 표시를 안 넘기면 사용자는 빈 제품
    # 화면에 떨어져서 어느 버튼이 사진 읽기인지 다시 찾아야 한다.
    target = reverse('products:product_detail_new', args=[label.my_label_id])
    if request.GET.get('import') == '1':
        target += '?import=1'
    return redirect(target)


@login_required
def product_update(request, product_id):
    """제품 수정"""
    # V2에서는 label_id를 product_id로 받음
    # 먼저 직접 MyLabel 조회 시도
    try:
        label = MyLabel.objects.get(my_label_id=product_id, user_id=request.user)
    except MyLabel.DoesNotExist:
        # MyLabel이 없으면 ProductMetadata를 통해 조회 시도
        try:
            metadata = ProductMetadata.objects.select_related('label').get(
                metadata_id=product_id,
                label__user_id=request.user
            )
            label = metadata.label
        except ProductMetadata.DoesNotExist:
            # 둘 다 없으면 404
            messages.error(request, '제품을 찾을 수 없습니다.')
            return redirect('products:product_explorer')
    
    if request.method == 'POST':
        # MyLabel 필드 업데이트
        label.my_label_name = request.POST.get('my_label_name', label.my_label_name)
        label.prdlst_nm = request.POST.get('prdlst_nm', label.prdlst_nm)
        label.ingredient_info = request.POST.get('ingredient_info', label.ingredient_info)
        label.prdlst_dcnm = request.POST.get('prdlst_dcnm', label.prdlst_dcnm)
        label.prdlst_report_no = request.POST.get('prdlst_report_no', label.prdlst_report_no)
        label.content_weight = request.POST.get('content_weight', label.content_weight)
        label.country_of_origin = request.POST.get('country_of_origin', label.country_of_origin)
        label.storage_method = request.POST.get('storage_method', label.storage_method)
        label.frmlc_mtrqlt = request.POST.get('frmlc_mtrqlt', label.frmlc_mtrqlt)
        label.bssh_nm = request.POST.get('bssh_nm', label.bssh_nm)
        label.pog_daycnt = request.POST.get('pog_daycnt', label.pog_daycnt)
        label.rawmtrl_nm_display = request.POST.get(
            'rawmtrl_nm_display', label.rawmtrl_nm_display)
        label.cautions = request.POST.get('cautions', label.cautions)
        label.additional_info = request.POST.get('additional_info', label.additional_info)
        label.food_group = request.POST.get('food_group', label.food_group)
        label.food_type = request.POST.get('food_type', label.food_type)
        label.processing_method = request.POST.get('processing_method', label.processing_method)
        label.processing_condition = request.POST.get('processing_condition', label.processing_condition)
        label.preservation_type = request.POST.get('preservation_type', label.preservation_type)
        label.distributor_address = request.POST.get('distributor_address', label.distributor_address)
        label.repacker_address = request.POST.get('repacker_address', label.repacker_address)
        label.importer_address = request.POST.get('importer_address', label.importer_address)
        import json as _json
        raw_cf = request.POST.get('custom_fields_json', None)
        if raw_cf is not None:
            try:
                label.custom_fields = _json.loads(raw_cf)
            except Exception:
                label.custom_fields = []
        label.save()

        # raw_material_yn / search_tags 저장
        meta = ProductMetadata.objects.filter(label=label).first()
        if meta:
            raw_yn = request.POST.get('raw_material_yn') == 'on'
            search_tags = request.POST.get('search_tags', '').strip()
            meta.raw_material_yn = raw_yn
            meta.search_tags = search_tags
            meta.save(update_fields=['raw_material_yn', 'search_tags'])

            # 원료로 사용 체크 시 MyIngredient 자동 등록 + LabelIngredientRelation 연동
            if raw_yn:
                from v1.label.models import MyIngredient as _MyIngredient, LabelIngredientRelation as _LIR
                product_name = label.prdlst_nm or label.my_label_name
                ingredient = _MyIngredient.objects.filter(
                    user_id=request.user,
                    prdlst_nm=product_name,
                    delete_YN='N',
                ).first()
                if not ingredient:
                    ingredient = _MyIngredient.objects.create(
                        user_id=request.user,
                        prdlst_nm=product_name,
                        bssh_nm=label.bssh_nm or '',
                        prdlst_dcnm=label.prdlst_dcnm or '',
                        pog_daycnt=label.pog_daycnt or '',
                        rawmtrl_nm=label.rawmtrl_nm or '',
                        ingredient_display_name=label.rawmtrl_nm or product_name,
                        food_category='processed',
                        delete_YN='N',
                    )
                _LIR.objects.get_or_create(
                    label=label,
                    ingredient=ingredient,
                    defaults={'relation_sequence': 1},
                )

        messages.success(request, '제품이 수정되었습니다.')
        return redirect('products:product_detail_new', product_id=label.my_label_id)

    # GET 은 워크스페이스로 보낸다.
    #
    # 예전에는 별도의 수정 폼(product_form.html)을 그렸다. 제품 상세의 기본
    # 정보 탭이 같은 일을 더 잘 하고(부분 저장·자동저장·불러오기), 어느 화면도
    # 이 주소로 링크하지 않는다. 폼을 두 벌로 두면 한쪽만 고쳐지는 날이 온다.
    #
    # POST 는 그대로 둔다 - 밖에서 부르고 있을지 모른다.
    return redirect('products:product_detail_new', product_id=label.my_label_id)


@login_required
@require_POST
def product_update_fields(request, product_id):
    """제품 정보 필드 업데이트 (AJAX)"""
    label = MyLabel.objects.filter(my_label_id=product_id, user_id=request.user).first()
    shared_share = None
    is_owner = False
    if label:
        is_owner = True
    else:
        shared_share = ProductShare.objects.filter(
            label__my_label_id=product_id,
            active_yn=True
        ).filter(
            Q(recipient_user=request.user) | Q(recipient_email__iexact=request.user.email)
        ).filter(
            Q(share_end_date__isnull=True) | Q(share_end_date__gt=timezone.now())
        ).select_related('label').first()

        if shared_share:
            label = shared_share.label
        else:
            return JsonResponse({'success': False, 'error': '제품을 찾을 수 없습니다'}, status=404)

    metadata = ProductMetadata.objects.filter(label=label).first()
    user_role = 'OWNER' if is_owner else 'VIEWER'
    if shared_share and not is_owner:
        share_permission = SharePermission.objects.filter(share=shared_share).first()
        if share_permission:
            user_role = share_permission.role_code

    if user_role not in ['OWNER', 'EDITOR']:
        return JsonResponse({'success': False, 'error': '수정 권한이 없습니다.'}, status=403)

    if user_role == 'EDITOR' and metadata and metadata.status not in [
        ProductMetadata.Status.DRAFT,
        ProductMetadata.Status.REQUESTING,
    ]:
        return JsonResponse({'success': False, 'error': '현재 상태에서는 수정할 수 없습니다.'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        # 허용된 필드만 업데이트
        allowed_fields = [
            'my_label_name', 'prdlst_dcnm', 'prdlst_nm', 'content_weight',
            'country_of_origin', 'pog_daycnt', 'storage_method', 'bssh_nm',
            'rawmtrl_nm', 'rawmtrl_nm_display', 'cautions', 'additional_info', 'ingredient_info',
            'prdlst_report_no', 'frmlc_mtrqlt',
            'food_group', 'food_type', 'processing_method', 'processing_condition',
            'preservation_type', 'distributor_address', 'repacker_address', 'importer_address',
            'allergens',
        ]
        
        # 변경된 필드 추적
        changed_fields = []
        field_labels = {
            'my_label_name': '라벨명',
            'prdlst_dcnm': '식품유형',
            'prdlst_nm': '제품명',
            'content_weight': '내용량',
            'country_of_origin': '원산지',
            'pog_daycnt': '소비기한',
            'storage_method': '보관방법',
            'bssh_nm': '제조원',
            'rawmtrl_nm': '원재료명(참고)',
            'rawmtrl_nm_display': '원재료명(표시)',
            'cautions': '주의사항',
            'additional_info': '기타표시사항',
            'ingredient_info': '특정성분 함량',
            'prdlst_report_no': '품목보고번호',
            'frmlc_mtrqlt': '용기·포장재질',
            'food_group': '식품유형(대분류)',
            'food_type': '식품유형(소분류)',
            'processing_method': '제조방법',
            'processing_condition': '제조방법 상세',
            'preservation_type': '장기보존식품',
            'distributor_address': '유통전문판매원',
            'repacker_address': '소분원',
            'importer_address': '수입원',
            'allergens': '알레르기 성분',
        }
        
        food_type_changed = (
            'food_type' in data
            and str(data['food_type'] or '') != str(label.food_type or '')
        )

        for field_name in allowed_fields:
            if field_name in data:
                old_value = getattr(label, field_name, '')
                new_value = data[field_name]
                if str(old_value) != str(new_value):
                    changed_fields.append(field_labels.get(field_name, field_name))
                setattr(label, field_name, data[field_name])

        # ── 표시 항목(chckd_*) ─────────────────────────────────────────────
        # 화면이 보낸 체크 상태를 그대로 반영한 뒤, 식품유형이 바뀐 경우에만
        # 그 유형의 규칙을 덧씌운다. 규칙은 필수('Y')를 켜고 해당 없음('D')을
        # 끄기만 하므로, 사용자 재량('N') 항목의 선택은 그대로 남는다.
        #
        # 반영 대상은 **모델에 있는 chckd_* 전부**다. 예전에는 식품유형 규칙표
        # (FIELD_TO_CHECKBOX)에 있는 것만 봤는데, 그 표에는 유통전문판매원·
        # 소분원·수입원·기타표시사항이 없다. 넷 다 오른쪽 패널에 있고 규정
        # 검증의 근거이기도 해서, 화면에서 껐다 켜도 저장이 안 되고 검증만
        # "표시하기로 선택했는데 비어 있습니다" 라고 말하는 상태였다.
        from v1.label.services import food_type_settings as fts

        checkbox_changed = []
        for item_checkbox in _CHECKBOX_FIELDS:
            if item_checkbox not in data:
                continue
            new_state = 'Y' if data[item_checkbox] in (True, 'Y', 'true', 1) else 'N'
            if (getattr(label, item_checkbox, '') or '') != new_state:
                checkbox_changed.append(item_checkbox)
            setattr(label, item_checkbox, new_state)

        # 규칙이 손댄 항목은 따로 모아 화면에 알린다. 조용히 켜 두면 사용자는
        # 켠 적 없는 체크를 근거로 한 지적을 받고, 무엇을 고쳐야 하는지 알 수 없다.
        rule_applied = {'turned_on': [], 'turned_off': []}
        if food_type_changed:
            rule = fts.resolve_settings(label.food_group or '', label.food_type or '')
            if rule['found']:
                applied = fts.apply_to_label(label, rule['settings'])
                checkbox_changed.extend(applied['turned_on'] + applied['turned_off'])
                rule_applied = {
                    'turned_on': [{'checkbox': cb, 'label': _checkbox_label(cb)}
                                  for cb in applied['turned_on']],
                    'turned_off': [{'checkbox': cb, 'label': _checkbox_label(cb)}
                                   for cb in applied['turned_off']],
                }

        if checkbox_changed:
            changed_fields.append('표시 항목')

        # 맞춤항목 JSON 처리
        if 'custom_fields_json' in data:
            import json as _json
            raw_cf = data['custom_fields_json']
            try:
                label.custom_fields = _json.loads(raw_cf) if isinstance(raw_cf, str) else raw_cf
            except Exception:
                label.custom_fields = []
            changed_fields.append('맞춤항목')
        
        label.save()
        
        # 변경사항이 있으면 활동 로그 생성
        if changed_fields:
            from .models import ProductActivityLog
            ProductActivityLog.objects.create(
                label=label,
                user=request.user,
                action='INFO_UPDATED',
                details={
                    'changed_fields': changed_fields
                }
            )
        
        # 저장된 값으로 표시 항목 목록을 다시 계산해서 돌려준다.
        #
        # 이 목록은 지금까지 페이지를 그릴 때 한 번만 만들어졌다. 그런데 저장
        # 뒤의 상태는 화면이 보낸 것과 다를 수 있다 — 식품유형이 바뀌면 위에서
        # 그 유형의 규칙이 체크를 켜고 끄고(apply_to_label), 미입력 표시는
        # 다른 탭이 채운 자리까지 봐야 한다. 돌려주지 않으면 오른쪽 패널이
        # **검증이 보는 값과 다른 화면**을 계속 보여 준다.
        return JsonResponse({
            'success': True,
            'message': '저장되었습니다',
            'display_items': _build_display_items(label),
            'rule_applied': rule_applied,
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def product_update_status(request, product_id):
    """제품 상태 업데이트 (워크플로우)"""
    label = MyLabel.objects.filter(my_label_id=product_id, user_id=request.user).first()
    shared_share = None
    is_owner = False
    if label:
        is_owner = True
    else:
        shared_share = ProductShare.objects.filter(
            label__my_label_id=product_id,
            active_yn=True
        ).filter(
            Q(recipient_user=request.user) | Q(recipient_email__iexact=request.user.email)
        ).filter(
            Q(share_end_date__isnull=True) | Q(share_end_date__gt=timezone.now())
        ).select_related('label').first()

        if shared_share:
            label = shared_share.label
        else:
            return JsonResponse({'success': False, 'error': '제품을 찾을 수 없습니다'}, status=404)

    metadata = get_object_or_404(ProductMetadata, label=label)
    user_role = 'OWNER' if is_owner else 'VIEWER'
    if shared_share and not is_owner:
        share_permission = SharePermission.objects.filter(share=shared_share).first()
        if share_permission:
            user_role = share_permission.role_code

    new_status = request.POST.get('status', '').strip()
    valid_statuses = {choice[0] for choice in ProductMetadata.Status.choices}
    if new_status not in valid_statuses:
        return JsonResponse({'success': False, 'error': '잘못된 상태입니다.'}, status=400)

    transitions = {
        ProductMetadata.Status.DRAFT:      {ProductMetadata.Status.REQUESTING},
        ProductMetadata.Status.REQUESTING: {ProductMetadata.Status.SUBMITTED, ProductMetadata.Status.DRAFT},
        ProductMetadata.Status.SUBMITTED:  {ProductMetadata.Status.REVIEW, ProductMetadata.Status.REQUESTING},
        ProductMetadata.Status.REVIEW:     {ProductMetadata.Status.PENDING, ProductMetadata.Status.SUBMITTED},
        ProductMetadata.Status.PENDING:    {ProductMetadata.Status.CONFIRMED, ProductMetadata.Status.REVIEW},
        ProductMetadata.Status.CONFIRMED:  {ProductMetadata.Status.DRAFT},
    }

    # 담당자 미지정 단계 스킵 전환 (OWNER/EDITOR만) — 스킵 허용 action 코드 목록
    skip_action_codes = []
    if user_role in ('OWNER', 'EDITOR'):
        _active_roles = set(SharePermission.objects.filter(
            share__label=label,
            share__active_yn=True,
        ).values_list('role_code', flat=True))
        _has_up = 'UPLOADER' in _active_roles
        _has_rv = 'REVIEWER' in _active_roles
        _has_ap = 'APPROVER' in _active_roles

        if metadata.status == ProductMetadata.Status.DRAFT and not _has_up:
            # REQUESTING/SUBMITTED 단계 건너뜀
            if _has_rv:
                transitions[ProductMetadata.Status.DRAFT].add(ProductMetadata.Status.REVIEW)
                skip_action_codes.append('review')
            if _has_ap:
                transitions[ProductMetadata.Status.DRAFT].add(ProductMetadata.Status.PENDING)
                skip_action_codes.append('pending')
            transitions[ProductMetadata.Status.DRAFT].add(ProductMetadata.Status.CONFIRMED)
            skip_action_codes.append('confirmed')
        elif metadata.status == ProductMetadata.Status.SUBMITTED and not _has_rv:
            # REVIEW 단계 건너뜀
            if _has_ap:
                transitions[ProductMetadata.Status.SUBMITTED].add(ProductMetadata.Status.PENDING)
                skip_action_codes.append('pending')
            transitions[ProductMetadata.Status.SUBMITTED].add(ProductMetadata.Status.CONFIRMED)
            skip_action_codes.append('confirmed')

    allowed_targets = transitions.get(metadata.status, set())
    if new_status not in allowed_targets:
        return JsonResponse({'success': False, 'error': '현재 상태에서 변경할 수 없습니다.'}, status=400)

    # 해당 단계 전이에 필요한 담당자 확인
    # (새 상태에 해당하는 권한자가 없으면 일부 전이 차단)
    status_required_roles = {
        ProductMetadata.Status.REQUESTING: ['UPLOADER'],
        ProductMetadata.Status.SUBMITTED:  [],  # 누구나 제출 가능
        ProductMetadata.Status.REVIEW:     ['REVIEWER'],
        ProductMetadata.Status.PENDING:    ['APPROVER'],
        ProductMetadata.Status.CONFIRMED:  [],
    }
    required_roles = status_required_roles.get(new_status, [])
    if required_roles:
        has_required = SharePermission.objects.filter(
            share__label=label,
            share__active_yn=True,
            role_code__in=required_roles
        ).filter(
            Q(share__share_end_date__isnull=True) | Q(share__share_end_date__gt=timezone.now())
        ).exists()
        if not has_required:
            role_names = {'UPLOADER': '자료 제출', 'REVIEWER': '검토자', 'APPROVER': '승인자'}
            required_label = ', '.join(role_names.get(r, r) for r in required_roles)
            status_label_map = {
                ProductMetadata.Status.REQUESTING: '자료 요청',
                ProductMetadata.Status.REVIEW:     '검토 중',
                ProductMetadata.Status.PENDING:    '승인 대기',
            }
            return JsonResponse({
                'success': False,
                'error': f'"{status_label_map.get(new_status, new_status)}" 단계로 이동하려면 "{required_label}" 권한을 가진 담당자가 필요합니다.',
            }, status=400)

    action_to_status = {
        'requesting': ProductMetadata.Status.REQUESTING,
        'submitted':  ProductMetadata.Status.SUBMITTED,
        'review':     ProductMetadata.Status.REVIEW,
        'pending':    ProductMetadata.Status.PENDING,
        'confirmed':  ProductMetadata.Status.CONFIRMED,
        'draft':      ProductMetadata.Status.DRAFT,
    }

    # 역할별 허용 행동
    status_actions_map_view = {
        ProductMetadata.Status.DRAFT:      ['requesting'],
        ProductMetadata.Status.REQUESTING: ['submitted', 'draft'],
        ProductMetadata.Status.SUBMITTED:  ['review', 'requesting'],
        ProductMetadata.Status.REVIEW:     ['pending', 'submitted'],
        ProductMetadata.Status.PENDING:    ['confirmed', 'review'],
        ProductMetadata.Status.CONFIRMED:  ['draft'],
    }
    if user_role == 'OWNER':
        available_actions = status_actions_map_view.get(metadata.status, []) + skip_action_codes
    elif user_role == 'EDITOR':
        available_actions = status_actions_map_view.get(metadata.status, []) + skip_action_codes
    elif user_role == 'UPLOADER':
        available_actions = ['submitted'] if metadata.status == ProductMetadata.Status.REQUESTING else []
    elif user_role == 'REVIEWER':
        available_actions = ['pending'] if metadata.status == ProductMetadata.Status.REVIEW else []
    elif user_role == 'APPROVER':
        available_actions = ['confirmed'] if metadata.status == ProductMetadata.Status.PENDING else []
    else:
        available_actions = []

    if new_status not in {action_to_status.get(a) for a in available_actions}:
        return JsonResponse({'success': False, 'error': '상태 변경 권한이 없습니다.'}, status=403)

    # 권한 플래그도 함께 확인한다.
    # 지금까지는 role_code 만 보고 can_review / can_approve 는 어디서도 쓰이지 않아,
    # 관리자 화면에서 체크를 풀어도 아무 변화가 없었다(화면이 거짓말을 하는 상태).
    if shared_share and not is_owner:
        _perm = SharePermission.objects.filter(share=shared_share).first()
        _needs = {
            ProductMetadata.Status.PENDING:   ('can_review',  '검토'),
            ProductMetadata.Status.CONFIRMED: ('can_approve', '승인'),
        }
        _need = _needs.get(new_status)
        if _need and not getattr(_perm, _need[0], False):
            return JsonResponse(
                {'success': False, 'error': f'{_need[1]} 권한이 없습니다.'}, status=403)

    # ── 승인 완료 전 표시사항 검증 ──────────────────────────────────────────
    # 실제로 잣(알레르기) 누락을 표시사항에 반영하지 못한 채 확정해 행정처분을 받은
    # 사례가 있었다. 검증 기능은 있었지만 확정 단계에서 강제되지 않아, 돌리지 않고
    # 넘어가면 그대로 통과됐다. 규칙 기반 검증(무료·무제한)을 확정 직전에 다시 돌린다.
    #
    # 오탐 여지가 있으므로 넘길 수 있게 하되, 넘긴 사실은 활동 로그에 남긴다.
    # 무엇을 요구하느냐는 이 제품에 검토·승인 역할이 배정돼 있는지에 따라 갈린다.
    #   - 배정돼 있다: 확정하는 사람과 작성한 사람이 다르다. 예외로 넘긴다는 판단을
    #     남겨야 하므로 사유를 받는다.
    #   - 배정돼 있지 않다: 혼자 쓰는 제품이다. 자기가 쓴 것을 자기에게 해명하게 만드는
    #     절차라 값이 없다. 무엇이 비었는지 보여주고 확인만 받는다.
    # 어느 쪽이든 첫 요청은 목록을 돌려주고 멈춘다 — 무엇이 빠졌는지 못 본 채로
    # 확정되는 경로가 있으면 안 된다.
    validation_override_reason = (request.POST.get('override_reason') or '').strip()
    validation_acknowledged = request.POST.get('validation_ack') == '1'
    validation_override = None
    if new_status == ProductMetadata.Status.CONFIRMED:
        from v1.label.services import validation_service as _vs
        try:
            _result = _vs.validate_label(label)
        except Exception:
            logger.exception('[승인 전 검증] 실행 실패 — 검증을 건너뛰고 진행합니다')
            _result = {'ok': True, 'issues': []}

        # 권고 항목만 남았으면 길을 막지 않는다. 확인이나 사유를 받는 무게는
        # 표시기준이 그렇게 적으라고 한 것에만 쓴다 — 그러지 않으면 절차가
        # 늘 뜨는 창이 되고, 사람은 읽지 않고 넘기는 법을 익힌다.
        _issues = [i for i in _result.get('issues', []) if not i.get('advisory')]
        if _issues:
            _missing = [i for i in _issues if i.get('category') == 'required_missing']
            # 필수 미입력은 한 건에 여러 항목이 담겨 온다(같은 문구 반복을 피하려고)
            _missing_names = [n for i in _missing for n in i.get('field_labels', [])]

            _has_workflow_roles = SharePermission.objects.filter(
                share__label=label,
                share__active_yn=True,
                role_code__in=['REVIEWER', 'APPROVER'],
            ).filter(
                Q(share__share_end_date__isnull=True) | Q(share__share_end_date__gt=timezone.now())
            ).exists()

            if _has_workflow_roles:
                _passed = bool(validation_override_reason)
            else:
                _passed = validation_acknowledged or bool(validation_override_reason)

            if not _passed:
                return JsonResponse({
                    'success': False,
                    'error': '표시사항 검증에서 확인이 필요한 항목이 있습니다.',
                    'validation_blocked': True,
                    # False 면 화면이 사유 입력 없이 "확인하고 계속"만 받는다
                    'requires_reason': _has_workflow_roles,
                    'issue_count': _result.get('issue_count', len(_issues)),
                    'issues': _issues,
                    'missing_required': _missing_names,
                }, status=400)

            validation_override = {
                'reason': validation_override_reason,
                'issue_count': len(_issues),
                'missing_required': _missing_names,
            }

    # ── CONFIRMED → DRAFT: 새 버전 번호 증가 ──
    old_status = metadata.status
    old_status_label = metadata.get_status_display()

    if old_status == ProductMetadata.Status.CONFIRMED and new_status == ProductMetadata.Status.DRAFT:
        metadata.status = new_status
        if hasattr(metadata, 'version'):
            metadata.version = (metadata.version or 1) + 1
            metadata.save(update_fields=['status', 'version', 'updated_datetime'])
        else:
            metadata.save(update_fields=['status', 'updated_datetime'])
    else:
        metadata.status = new_status
        metadata.save(update_fields=['status', 'updated_datetime'])

    # ── 활동 로그 기록 ──
    from .models import ProductActivityLog
    _log_details = {
        'old_status': old_status,
        'old_status_label': old_status_label,
        'new_status': new_status,
        'new_status_label': metadata.get_status_display(),
    }
    if validation_override:
        # 검증 이슈를 남긴 채 확정한 경우 — 누가, 무엇을 알고도 넘겼는지 남긴다.
        # 사유가 없는 건(검토·승인 역할 미배정) 확인만 받고 넘어간 경우다.
        _log_details['validation_override'] = True
        _log_details['override_issue_count'] = validation_override['issue_count']
        if validation_override['missing_required']:
            _log_details['override_missing_required'] = validation_override['missing_required']
        if validation_override['reason']:
            _log_details['override_reason'] = validation_override['reason']
        else:
            _log_details['override_acknowledged'] = True
    ProductActivityLog.objects.create(
        label=label,
        user=request.user,
        action='STATUS_CHANGED',
        details=_log_details,
    )

    # ── 새 상태에서 담당 역할을 가진 공유자에게 인앱 알림 발송 ──
    notify_roles = {
        ProductMetadata.Status.REQUESTING: ['UPLOADER'],
        ProductMetadata.Status.SUBMITTED:  ['REVIEWER', 'EDITOR', 'OWNER'],
        ProductMetadata.Status.REVIEW:     ['REVIEWER'],
        ProductMetadata.Status.PENDING:    ['APPROVER'],
        ProductMetadata.Status.CONFIRMED:  ['OWNER', 'EDITOR'],
        ProductMetadata.Status.DRAFT:      ['OWNER', 'EDITOR'],
    }
    new_status_label = metadata.get_status_display()
    product_name = label.my_label_name or label.prdlst_nm or '제품'
    changer_name, changer_company = _get_sender_info(request.user)
    _status_email_subject = f'[EzLabeling] {product_name} 상태가 변경되었습니다'
    from django.conf import settings as _ds
    _site_url = getattr(_ds, 'SITE_URL', 'https://labeldata.pythonanywhere.com')
    # 역할별 해야 할 일 안내
    _status_role_guide = {
        'UPLOADER': '요청된 자료(원료 규격서, 성분 데이터 등)를 시스템에 업로드해 주세요.',
        'REVIEWER': '제출된 자료를 검토하고 피드백을 남겨주세요.',
        'EDITOR':   '제품 정보를 확인하고 필요 시 수정·보완해 주세요.',
        'APPROVER': '검토 결과를 확인하고 최종 승인 또는 반려해 주세요.',
        'OWNER':    '제품 상태 변경 내역을 확인해 주세요.',
    }

    for nrole in notify_roles.get(new_status, []):
        if nrole in ('OWNER',):
            # 소유자 직접 알림
            owner_user = label.user_id
            if owner_user and owner_user != request.user:
                ProductNotification.objects.create(
                    label=label,
                    recipient=owner_user,
                    message=f'[{product_name}] 상태가 "{new_status_label}"으(로) 변경되었습니다.',
                    status_code=new_status,
                )
                # 소유자에게 이메일도 발송
                _txt, _html = _render_email('emails/workflow_status.html', {
                    'subject': _status_email_subject,
                    'sender_name': changer_name, 'sender_company': changer_company, 'sender_email': request.user.email,
                    'product_name': product_name, 'new_status_label': new_status_label, 'role_label': 'OWNER',
                    'task_description': _status_role_guide.get('OWNER', '시스템에서 확인해 주세요.'),
                    'inbox_url': f'{_site_url}/products/inbox/',
                })
                _send_email_safe(subject=_status_email_subject, body=_txt, to_email=owner_user.email, html_body=_html)
        else:
            perm_qs = SharePermission.objects.filter(
                share__label=label,
                share__active_yn=True,
                role_code=nrole,
            ).filter(
                Q(share__share_end_date__isnull=True) | Q(share__share_end_date__gt=timezone.now())
            ).select_related('share__recipient_user', 'share')
            for perm in perm_qs:
                recipient = perm.share.recipient_user
                notify_msg = f'[{product_name}] 상태가 "{new_status_label}"으(로) 변경되었습니다. 귀하의 작업이 필요합니다.'
                _email_ctx = {
                    'subject': _status_email_subject,
                    'sender_name': changer_name, 'sender_company': changer_company, 'sender_email': request.user.email,
                    'product_name': product_name, 'new_status_label': new_status_label, 'role_label': nrole,
                    'task_description': _status_role_guide.get(nrole, '시스템에서 확인해 주세요.'),
                    'inbox_url': f'{_site_url}/products/inbox/',
                }
                _txt, _html = _render_email('emails/workflow_status.html', _email_ctx)
                if recipient and recipient != request.user:
                    # 시스템 계정 있는 공유자: 인앱 알림 + 이메일
                    ProductNotification.objects.create(
                        label=label, recipient=recipient, message=notify_msg, status_code=new_status,
                    )
                    if recipient.email:
                        _send_email_safe(subject=_status_email_subject, body=_txt, to_email=recipient.email, html_body=_html)
                elif not recipient and perm.share.recipient_email:
                    # 이메일 전용 공유자(시스템 계정 없음): 이메일만 발송
                    _send_email_safe(subject=_status_email_subject, body=_txt, to_email=perm.share.recipient_email, html_body=_html)

    # ── 승인 완료 시 거래처(공유 대상)에 확정 통보 ──────────────────────────
    # 원산지를 바꾸고 사내·거래처 공유가 누락돼, 일부 거래처가 이전 표시사항으로
    # 계속 판매하다 행정처분을 받은 사례가 있었다. 확정 알림이 담당 역할(OWNER/EDITOR)
    # 에게만 가고 정작 물건을 파는 쪽에는 가지 않았던 것이 원인이다.
    # 확정 시점에 공유받은 전원에게 알리고, 표시사항 도안 PDF 가 있으면 첨부한다.
    if new_status == ProductMetadata.Status.CONFIRMED:
        _notify_confirmed_to_partners(request, label, product_name, changer_name, changer_company)

    log_activity(request, 'product', 'workflow_status_change', product_id)
    return JsonResponse({
        'success': True,
        'status': metadata.status,
        'status_label': metadata.get_status_display()
    })


def _latest_label_pdf(label):
    """
    문서함에 저장된 최신 '한글표시사항도안' PDF 를 (파일명, 바이트, mimetype) 로 반환.
    없으면 None. 미리보기에서 PDF 저장을 하면 이 타입으로 등록된다.
    """
    try:
        # **PDF 만 고른다.** 이 자리에는 판독에 쓴 표시사항 사진(JPG/PNG)도
        # 들어간다. 확장자를 안 보면 그 사진을 집어 application/pdf 로 붙이게
        # 되고, 받는 쪽에서는 열리지 않는 첨부가 된다.
        doc = (ProductDocument.objects
               .filter(label=label, active_yn=True,
                       document_type__type_name__contains='표시사항')
               .filter(Q(file_extension__iexact='.pdf')
                       | Q(original_filename__iendswith='.pdf'))
               .order_by('-uploaded_datetime', '-document_id')
               .first())
        if not doc or not doc.file:
            return None
        with doc.file.open('rb') as fh:
            content = fh.read()
        return (doc.original_filename or 'label.pdf', content, 'application/pdf')
    except Exception:
        logger.exception('[확정 통보] 표시사항 PDF 첨부 실패 (첨부 없이 발송)')
        return None


def _notify_confirmed_to_partners(request, label, product_name, changer_name, changer_company):
    """
    표시사항 확정 사실을 공유받은 거래처 전원에게 알린다.

    - 유효한 공유(active / 기간 내)만 대상으로 한다
    - 역할과 무관하게 보낸다. 물건을 파는 쪽은 보통 VIEWER 라서 역할 기준으로 거르면
      정작 알아야 할 사람이 빠진다(이번 사고의 원인)
    - 회원이면 인앱 알림 + 이메일, 이메일 전용 공유자면 이메일만
    """
    from django.conf import settings as _ds
    site_url = getattr(_ds, 'SITE_URL', 'https://labeldata.pythonanywhere.com')

    shares = (ProductShare.objects
              .filter(label=label, active_yn=True, share_mode='PRIVATE')
              .filter(Q(share_end_date__isnull=True) | Q(share_end_date__gt=timezone.now()))
              .select_related('recipient_user', 'permission'))
    if not shares:
        return

    attachment = _latest_label_pdf(label)
    subject = f'[EzLabeling] {product_name} 표시사항이 확정되었습니다'
    sent = 0

    for share in shares:
        recipient = share.recipient_user
        to_email = (recipient.email if recipient else '') or share.recipient_email
        if recipient and recipient == request.user:
            continue

        ctx = {
            'subject': subject,
            'sender_name': changer_name,
            'sender_company': changer_company,
            'sender_email': request.user.email,
            'product_name': product_name,
            'new_status_label': '승인 완료',
            'role_label': getattr(share.permission, 'role_code', '') if hasattr(share, 'permission') else '',
            'task_description': (
                '이 제품의 표시사항이 확정되었습니다. '
                '기존에 받으신 표시사항이 있다면 최신본으로 교체해 주세요. '
                + ('확정된 표시사항 도안을 첨부합니다.' if attachment else
                   '시스템에서 최신 표시사항을 확인하실 수 있습니다.')
            ),
            'inbox_url': f'{site_url}/products/inbox/',
        }
        txt, html = _render_email('emails/workflow_status.html', ctx)

        if recipient:
            ProductNotification.objects.create(
                label=label,
                recipient=recipient,
                message=f'[{product_name}] 표시사항이 확정되었습니다. 최신본을 확인해 주세요.',
                status_code=ProductMetadata.Status.CONFIRMED,
            )
        if to_email:
            _send_email_safe(subject=subject, body=txt, to_email=to_email,
                             html_body=html, attachment=attachment)
            sent += 1

    logger.info('[확정 통보] %s — 거래처 %d곳에 발송 (첨부 %s)',
                product_name, sent, '있음' if attachment else '없음')


@login_required
def product_delete(request, product_id):
    """제품 삭제 (휴지통으로 이동)"""
    # V2에서는 label_id를 product_id로 받음
    try:
        label = MyLabel.objects.get(my_label_id=product_id, user_id=request.user)
    except MyLabel.DoesNotExist:
        try:
            metadata = ProductMetadata.objects.select_related('label').get(
                metadata_id=product_id,
                label__user_id=request.user
            )
            label = metadata.label
        except ProductMetadata.DoesNotExist:
            messages.error(request, '제품을 찾을 수 없습니다.')
            return redirect('products:product_explorer')
    
    if request.method == 'POST':
        # 소프트 삭제
        label.delete_YN = 'Y'
        from datetime import datetime
        label.delete_datetime = datetime.now().strftime('%Y%m%d')
        label.save()
        
        messages.success(request, '제품이 삭제되었습니다.')
        return redirect('products:product_explorer')
    
    messages.info(request, '삭제는 확인 화면 없이 바로 처리됩니다.')
    return redirect('products:product_detail_new', product_id=label.my_label_id)


@login_required
@require_POST
def bulk_delete_products(request):
    """제품 일괄 삭제 (AJAX)"""
    try:
        data = json.loads(request.body)
        product_ids = data.get('product_ids', [])
        
        if not product_ids:
            return JsonResponse({'success': False, 'message': '삭제할 제품이 선택되지 않았습니다.'})
        
        # 사용자 소유의 제품만 삭제
        labels = MyLabel.objects.filter(
            my_label_id__in=product_ids,
            user_id=request.user,
            delete_YN='N'
        )
        
        deleted_count = 0
        for label in labels:
            label.delete_YN = 'Y'
            from datetime import datetime
            label.delete_datetime = datetime.now().strftime('%Y%m%d')
            label.save()
            deleted_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'{deleted_count}개 제품이 삭제되었습니다.',
            'deleted_count': deleted_count
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def bulk_copy_products(request):
    """제품 일괄 복사 (AJAX)"""
    try:
        data = json.loads(request.body)
        product_ids = data.get('product_ids', [])
        
        if not product_ids:
            return JsonResponse({'success': False, 'message': '복사할 제품이 선택되지 않았습니다.'})
        
        # 사용자 소유의 제품만 복사
        labels = MyLabel.objects.filter(
            my_label_id__in=product_ids,
            user_id=request.user,
            delete_YN='N'
        )
        
        copied_count = 0
        for original_label in labels:
            # 새 라벨 생성
            new_label = MyLabel.objects.create(
                user_id=request.user,
                my_label_name=f"{original_label.my_label_name} (복사본)",
                prdlst_nm=original_label.prdlst_nm,
                prdlst_dcnm=original_label.prdlst_dcnm,
                food_type=original_label.food_type,
                calories=original_label.calories,
                carbohydrates=original_label.carbohydrates,
                proteins=original_label.proteins,
                fats=original_label.fats,
                natriums=original_label.natriums,
                sugars=original_label.sugars,
                saturated_fats=original_label.saturated_fats,
                trans_fats=original_label.trans_fats,
                cholesterols=original_label.cholesterols,
                allergens=original_label.allergens,
                delete_YN='N'
            )
            
            # ProductMetadata 복사
            try:
                original_meta = ProductMetadata.objects.get(label=original_label)
                
                # 고유한 product_code 생성
                user_product_count = ProductMetadata.objects.filter(
                    label__user_id=request.user
                ).count()
                
                product_code = None
                for i in range(100):
                    candidate_code = f"PRD-{request.user.id}-{user_product_count + i + 1:04d}"
                    if not ProductMetadata.objects.filter(product_code=candidate_code).exists():
                        product_code = candidate_code
                        break
                
                if not product_code:
                    import time
                    product_code = f"PRD-{request.user.id}-{int(time.time())}"
                
                ProductMetadata.objects.create(
                    label=new_label,
                    product_code=product_code,
                    starred_yn=False
                )
            except ProductMetadata.DoesNotExist:
                pass
            
            # BOM 데이터 복사
            original_bom = ProductBOM.objects.filter(parent_label=original_label)
            for bom_item in original_bom:
                ProductBOM.objects.create(
                    parent_label=new_label,
                    child_label=bom_item.child_label,
                    shared_receipt=bom_item.shared_receipt,
                    ingredient_name=bom_item.ingredient_name,
                    raw_material_name=bom_item.raw_material_name,
                    sub_ingredients=bom_item.sub_ingredients,
                    food_type=bom_item.food_type,
                    usage_ratio=bom_item.usage_ratio,
                    manufacturer=bom_item.manufacturer,
                    allergens=bom_item.allergens,
                    gmo=bom_item.gmo,
                    report_no=bom_item.report_no,
                    origin=bom_item.origin,
                    origin_detail=bom_item.origin_detail,
                    additive_yn=bom_item.additive_yn,
                    additive_role=bom_item.additive_role,
                    gmo_yn=bom_item.gmo_yn,
                    notes=bom_item.notes
                )
            
            copied_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'{copied_count}개 제품이 복사되었습니다.',
            'copied_count': copied_count
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def bulk_export_products_excel(request):
    """선택된 제품 데이터를 탭별 시트로 엑셀 다운로드"""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font

        data = json.loads(request.body)
        product_ids = data.get('product_ids', [])
        tabs = data.get('tabs', ['basic'])  # 선택된 탭 목록

        if not product_ids:
            return JsonResponse({'success': False, 'error': '선택된 제품이 없습니다.'}, status=400)
        if not tabs:
            return JsonResponse({'success': False, 'error': '다운로드할 탭을 선택하세요.'}, status=400)

        labels = MyLabel.objects.filter(
            my_label_id__in=product_ids,
            delete_YN='N'
        ).filter(
            Q(user_id=request.user) |
            Q(v2_shares__recipient_user=request.user, v2_shares__active_yn=True)
        ).distinct().order_by('my_label_id')

        if not labels.exists():
            return JsonResponse({'success': False, 'error': '다운로드할 데이터가 없습니다.'}, status=400)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color='D3E4F0', end_color='D3E4F0', fill_type='solid')
        header_font = Font(bold=True, size=9)
        body_font = Font(size=9)

        def style_ws(ws):
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            ws.freeze_panes = ws.cell(row=2, column=1)

        today_str = timezone.now().strftime('%y%m%d')

        # ========== 기본정보 ==========
        if 'basic' in tabs:
            ws = wb.create_sheet('기본정보')
            ws.append(['번호', '제품코드', '제품명', '식품유형', '상태', '수정일', '즐겨찾기', '폴더'])
            for idx, label in enumerate(labels, 1):
                meta = getattr(label, 'v2_metadata', None)
                ws.append([
                    idx,
                    meta.product_code if meta else '',
                    label.my_label_name or '',
                    label.prdlst_dcnm or '',
                    meta.get_status_display() if meta else '',
                    label.update_datetime.strftime('%y-%m-%d %H:%M') if label.update_datetime else '',
                    '★' if (meta and meta.starred_yn) else '',
                    meta.folder.name if (meta and meta.folder) else '',
                ])
            style_ws(ws)
            for col, w in [('B', 14), ('C', 25), ('D', 16), ('E', 10), ('F', 14)]:
                ws.column_dimensions[col].width = w

        # ========== 표시사항 ==========
        if 'label' in tabs:
            ws = wb.create_sheet('표시사항')
            ws.append([
                '번호', '제품명', '품목보고번호', '제조사명', '성분명 및 함량',
                '내용량', '내용량(열량)', '원산지', '보관방법', '포장재질',
                '유통전문판매원', '소분원', '수입원', '소비기한',
                '원재료명(표시)', '원재료명(참고)', '주의사항', '기타표시사항', '영양성분'
            ])
            for idx, label in enumerate(labels, 1):
                ws.append([
                    idx, label.my_label_name or '', label.prdlst_report_no or '',
                    label.bssh_nm or '', label.ingredient_info or '',
                    label.content_weight or '', label.weight_calorie or '',
                    label.country_of_origin or '', label.storage_method or '',
                    label.frmlc_mtrqlt or '', label.distributor_address or '',
                    label.repacker_address or '', label.importer_address or '',
                    label.pog_daycnt or '', label.rawmtrl_nm_display or '',
                    label.rawmtrl_nm or '', label.cautions or '',
                    label.additional_info or '', label.nutrition_text or '',
                ])
            style_ws(ws)
            for col, w in [('B', 22), ('C', 14), ('D', 18), ('E', 30), ('O', 28), ('P', 28)]:
                ws.column_dimensions[col].width = w

        # ========== 영양성분 ==========
        if 'nutrition' in tabs:
            ws = wb.create_sheet('영양성분')
            ws.append(['번호', '제품명', '열량(kcal)', '탄수화물(g)', '당류(g)',
                        '단백질(g)', '지방(g)', '포화지방(g)', '트랜스지방(g)', '콜레스테롤(mg)', '나트륨(mg)'])
            for idx, label in enumerate(labels, 1):
                ws.append([
                    idx, label.my_label_name or '',
                    label.calories or '', label.carbohydrates or '', label.sugars or '',
                    label.proteins or '', label.fats or '',
                    label.saturated_fats or '', label.trans_fats or '',
                    label.cholesterols or '', label.natriums or '',
                ])
            style_ws(ws)
            ws.column_dimensions['B'].width = 22

        # ========== BOM ==========
        if 'bom' in tabs:
            from v1.bom.models import ProductBOM as BOM
            ws = wb.create_sheet('BOM')
            ws.append(['번호', '제품명', '원재료명', '함량(%)', '원산지', '알레르기', '식품첨가물', '용도', '비고'])
            row_idx = 1
            for label in labels:
                bom_items = BOM.objects.filter(parent_label=label, level=1).order_by('sort_order')
                if bom_items.exists():
                    for bom in bom_items:
                        ws.append([
                            row_idx, label.my_label_name or '',
                            bom.ingredient_name or '',
                            float(bom.usage_ratio) if bom.usage_ratio is not None else '',
                            bom.origin or '', bom.allergen or '',
                            '○' if bom.additive_yn else '',
                            bom.additive_role or '', bom.notes or '',
                        ])
                        row_idx += 1
                else:
                    ws.append([row_idx, label.my_label_name or '', '(BOM 없음)', '', '', '', '', '', ''])
                    row_idx += 1
            style_ws(ws)
            for col, w in [('B', 22), ('C', 26), ('E', 14), ('F', 20)]:
                ws.column_dimensions[col].width = w

        # ========== 문서함 ==========
        if 'documents' in tabs:
            ws = wb.create_sheet('문서함')
            ws.append(['번호', '제품명', '문서 구분', '파일명', '발행일', '만료일', '상태', '등록자', '등록일'])
            row_idx = 1
            for label in labels:
                slots = DocumentSlot.objects.filter(label=label, hidden_yn=False).select_related('document_type')
                if slots.exists():
                    for slot in slots:
                        doc = ProductDocument.objects.filter(
                            label=label, document_type=slot.document_type, active_yn=True
                        ).order_by('-uploaded_datetime').first()
                        today_date = timezone.now().date()
                        if doc:
                            if doc.expiry_date is None:
                                status_str = '유효(무기한)'
                            elif doc.expiry_date < today_date:
                                status_str = '만료'
                            elif (doc.expiry_date - today_date).days <= 30:
                                status_str = '만료임박'
                            else:
                                status_str = '유효'
                            ws.append([
                                row_idx, label.my_label_name or '',
                                slot.document_type.type_name or '',
                                doc.original_filename or '',
                                doc.issue_date.strftime('%Y-%m-%d') if doc.issue_date else '',
                                doc.expiry_date.strftime('%Y-%m-%d') if doc.expiry_date else '무기한',
                                status_str,
                                doc.uploaded_by.get_full_name() or doc.uploaded_by.username if doc.uploaded_by else '',
                                doc.uploaded_datetime.strftime('%Y-%m-%d') if doc.uploaded_datetime else '',
                            ])
                        else:
                            ws.append([row_idx, label.my_label_name or '',
                                       slot.document_type.type_name or '',
                                       '미등록', '', '', '미등록', '', ''])
                        row_idx += 1
                else:
                    ws.append([row_idx, label.my_label_name or '', '(문서 슬롯 없음)', '', '', '', '', '', ''])
                    row_idx += 1
            style_ws(ws)
            for col, w in [('B', 22), ('C', 18), ('D', 32), ('F', 12), ('G', 10), ('H', 14), ('I', 12)]:
                ws.column_dimensions[col].width = w

        # 본문 폰트 적용
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = body_font

        filename = f'LabelData_제품데이터_{today_str}.xlsx'
        resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
        wb.save(resp)
        return resp

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ==================== 폴더 관리 ====================

@login_required
@require_POST
def folder_create(request):
    """폴더 생성 (AJAX)"""
    folder_name = request.POST.get('folder_name', '').strip()
    parent_id = request.POST.get('parent_id')
    
    if not folder_name:
        return JsonResponse({'success': False, 'error': '폴더명을 입력하세요.'})
    
    parent_folder = None
    if parent_id:
        parent_folder = get_object_or_404(ProductFolder, folder_id=parent_id, owner=request.user)
    
    folder = ProductFolder.objects.create(
        owner=request.user,
        name=folder_name,
        parent=parent_folder
    )
    
    return JsonResponse({
        'success': True,
        'folder_id': folder.folder_id,
        'name': folder.name
    })


@login_required
@require_POST
def folder_rename(request, folder_id):
    """폴더 이름 변경 (AJAX)"""
    folder = get_object_or_404(ProductFolder, folder_id=folder_id, owner=request.user)
    new_name = request.POST.get('new_name', '').strip()
    
    if not new_name:
        return JsonResponse({'success': False, 'error': '폴더명을 입력하세요.'})
    
    folder.name = new_name
    folder.save()
    
    return JsonResponse({'success': True})


@login_required
@require_POST
def folder_delete(request, folder_id):
    """폴더 삭제 (AJAX)"""
    folder = get_object_or_404(ProductFolder, folder_id=folder_id, owner=request.user)
    
    # 하위 항목이 있는지 확인
    if folder.children.exists() or folder.products.exists():
        return JsonResponse({'success': False, 'error': '폴더에 항목이 있어 삭제할 수 없습니다.'})
    
    folder.delete()
    return JsonResponse({'success': True})


# ==================== 제품 이동 ====================

@login_required
@require_POST
def product_move(request, product_id):
    """제품을 다른 폴더로 이동 (AJAX)"""
    # Product는 MyLabel의 alias. PK는 my_label_id, 소유자 FK는 user_id
    label = get_object_or_404(MyLabel, my_label_id=product_id, user_id=request.user)
    # folder는 ProductMetadata에 있음
    metadata = get_object_or_404(ProductMetadata, label=label)
    folder_id = request.POST.get('folder_id')

    if folder_id:
        folder = get_object_or_404(ProductFolder, folder_id=folder_id, owner=request.user)
        metadata.folder = folder
    else:
        metadata.folder = None

    metadata.save(update_fields=['folder'])
    return JsonResponse({'success': True})


# ==================== 즐겨찾기 토글 ====================

@login_required
@require_POST
def product_favorite_toggle(request, product_id):
    """즐겨찾기 토글 (AJAX) - ProductMetadata 기반"""
    try:
        label = MyLabel.objects.get(my_label_id=product_id, user_id=request.user)
    except MyLabel.DoesNotExist:
        return JsonResponse({'success': False, 'error': '제품을 찾을 수 없습니다'}, status=404)
    
    # ProductMetadata 가져오기 또는 생성
    try:
        metadata = ProductMetadata.objects.get(label=label)
    except ProductMetadata.DoesNotExist:
        # 메타데이터가 없으면 생성
        user_product_count = ProductMetadata.objects.filter(
            label__user_id=request.user
        ).count()
        
        product_code = None
        for i in range(100):
            candidate_code = f"PRD-{request.user.id}-{user_product_count + i + 1:04d}"
            if not ProductMetadata.objects.filter(product_code=candidate_code).exists():
                product_code = candidate_code
                break
        
        if not product_code:
            import time
            product_code = f"PRD-{request.user.id}-{int(time.time())}"
        
        metadata = ProductMetadata.objects.create(
            label=label,
            product_code=product_code,
            starred_yn=False
        )
    
    # 즐겨찾기 토글
    metadata.starred_yn = not metadata.starred_yn
    if metadata.starred_yn:
        metadata.starred_datetime = timezone.now()
    else:
        metadata.starred_datetime = None
    metadata.save()
    
    return JsonResponse({
        'success': True,
        'starred_yn': metadata.starred_yn
    })


# ==================== 휴지통 ====================

@login_required
def product_trash(request):
    """휴지통 목록"""
    messages.info(request, '휴지통 화면은 준비 중입니다.')
    return redirect('products:product_explorer')


@login_required
@require_POST
def product_restore(request, product_id):
    """제품 복원 - 휴지통 기능 구현 전 차단"""
    messages.info(request, '복원 기능은 준비 중입니다.')
    return redirect('products:product_explorer')


@login_required
@require_POST
def product_permanent_delete(request, product_id):
    """제품 영구 삭제 - 휴지통 기능 구현 전 차단"""
    messages.info(request, '영구 삭제 기능은 준비 중입니다.')
    return redirect('products:product_explorer')


# ==================== 검색 ====================

@login_required
def product_search(request):
    """전체 제품 검색"""
    query = request.GET.get('q', '').strip()
    if query:
        return redirect(f"{reverse('products:product_explorer')}?q={quote(query)}")
    return redirect('products:product_explorer')


# ==================== 공유/협업 ====================

@login_required
def sharing_inbox(request):
    """공동작업 - 나에게 권한이 부여된 제품 목록"""
    status_filter = request.GET.get('status', '')
    role_filter   = request.GET.get('role', '')

    # ProductShare 기반으로 직접 조회 (recipient_user 또는 이메일 매칭 모두 포함)
    my_shares = ProductShare.objects.filter(
        Q(recipient_user=request.user) | Q(recipient_email__iexact=request.user.email),
        active_yn=True,
    ).filter(
        Q(share_end_date__isnull=True) | Q(share_end_date__gt=timezone.now())
    ).select_related(
        'label', 'created_by', 'permission',
    ).order_by('-share_start_date')

    # label_id → ProductMetadata 매핑 (status 조회용)
    label_ids = list({s.label_id for s in my_shares})
    metadata_map = {
        m.label_id: m
        for m in ProductMetadata.objects.filter(label_id__in=label_ids)
    }

    # label_id → 활성 문서 수 매핑 (미리보기용)
    from django.db.models import Count
    doc_count_map = {
        row['label_id']: row['cnt']
        for row in ProductDocument.objects.filter(
            label_id__in=label_ids, active_yn=True
        ).values('label_id').annotate(cnt=Count('document_id'))
    }

    # 역할별 할 일 매핑
    def _action_needed(role_code, status):
        mapping = {
            ('UPLOADER',  'REQUESTING'): ('업로드', 'bi bi-upload',        'text-primary'),
            ('REVIEWER',  'SUBMITTED'):  ('검토 필요', 'bi bi-eye',         'text-warning'),
            ('REVIEWER',  'REVIEW'):     ('검토 중',   'bi bi-eye-fill',    'text-warning'),
            ('APPROVER',  'PENDING'):    ('승인 대기', 'bi bi-check-circle', 'text-danger'),
            ('EDITOR',    'DRAFT'):      ('편집 가능', 'bi bi-pencil',       'text-success'),
            ('EDITOR',    'REQUESTING'): ('편집 가능', 'bi bi-pencil',       'text-success'),
        }
        hit = mapping.get((role_code, status))
        return hit if hit else (None, None, None)

    # 각 share에 metadata + action 속성 추가
    result_shares = []
    for share in my_shares:
        share.metadata = metadata_map.get(share.label_id)
        share.doc_count = doc_count_map.get(share.label_id, 0)
        perm = getattr(share, 'permission', None)
        role_code = perm.role_code if perm else 'VIEWER'
        status = share.metadata.status if share.metadata else 'DRAFT'
        share.role_code = role_code
        share.action_label, share.action_icon, share.action_style = _action_needed(role_code, status)
        result_shares.append(share)

    # 전체 상태별 카운트 (필터 적용 전 — 역할 필터 먼저 적용 후 집계)
    # 역할 필터 먼저 적용
    if role_filter:
        role_filtered = [s for s in result_shares if s.role_code == role_filter]
    else:
        role_filtered = result_shares

    # 상태별 카운트: 역할 필터 이후 기준으로 집계해야 카드 수치와 리스트가 일치
    status_counts = {}
    for share in role_filtered:
        s = share.metadata.status if share.metadata else 'DRAFT'
        status_counts[s] = status_counts.get(s, 0) + 1

    total_count = len(role_filtered)

    # 상태 필터 적용
    filtered = role_filtered
    if status_filter:
        filtered = [
            s for s in filtered
            if (s.metadata and s.metadata.status == status_filter)
            or (not s.metadata and status_filter == 'DRAFT')
        ]

    # SharedProductReceipt 자동 생성 및 각 share에 my_receipt 첨부
    existing_receipts = {
        r.share_id: r
        for r in SharedProductReceipt.objects.filter(
            receiver=request.user,
            share_id__in=[s.share_id for s in result_shares],
        )
    }
    shares_without_receipt = [s for s in result_shares if s.share_id not in existing_receipts]
    if shares_without_receipt:
        SharedProductReceipt.objects.bulk_create(
            [SharedProductReceipt(share=s, receiver=request.user) for s in shares_without_receipt],
            ignore_conflicts=True,
        )
        for r in SharedProductReceipt.objects.filter(
            receiver=request.user,
            share_id__in=[s.share_id for s in shares_without_receipt],
        ):
            existing_receipts[r.share_id] = r
    for share in result_shares:
        share.my_receipt = existing_receipts.get(share.share_id)

    context = {
        'received_shares': filtered,
        'title': '공동작업',
        'status_filter': status_filter,
        'role_filter': role_filter,
        'status_counts': status_counts,
        'total_count': total_count,
        'filtered_count': len(filtered),
        'Status': ProductMetadata.Status,
        'role_choices': SharePermission.ROLE_CHOICES,
    }
    return render(request, 'products/sharing/inbox.html', context)


@login_required
def product_detail_new(request, product_id):
    """제품 상세 (새 스타일)"""
    # product_detail과 동일하게 처리
    return product_detail(request, product_id)


@login_required
@xframe_options_sameorigin
def nutrition_workspace(request, label_id):
    """영양성분 입력 워크스페이스 (오너 + 공유 사용자 접근 가능)"""
    from v1.label.models import MyLabel
    from datetime import datetime

    # 오너 우선 조회
    is_owner = False
    can_edit = False
    try:
        label = MyLabel.objects.get(my_label_id=label_id, user_id=request.user, delete_YN='N')
        is_owner = True
        can_edit = True
    except MyLabel.DoesNotExist:
        # 공유 사용자 접근 허용
        shared_share = ProductShare.objects.filter(
            label__my_label_id=label_id,
            active_yn=True
        ).filter(
            Q(recipient_user=request.user) | Q(recipient_email__iexact=request.user.email)
        ).filter(
            Q(share_end_date__isnull=True) | Q(share_end_date__gt=timezone.now())
        ).select_related('label', 'permission').first()

        if not shared_share:
            from django.http import Http404
            raise Http404('접근 권한이 없습니다.')
        label = shared_share.label
        perm = getattr(shared_share, 'permission', None)
        can_edit = bool(perm and perm.can_edit_label)

    context = {
        'label': label,
        'LABEL_ID': label.my_label_id,
        'STATIC_BUILD_DATE': datetime.now().strftime('%Y%m%d%H%M%S'),
        'can_edit': can_edit,
    }
    log_activity(request, 'product', 'nutrition_view', label_id)
    return render(request, 'products/nutrition_editor.html', context)


@login_required
def nutrition_data_api(request, label_id):
    """영양성분 데이터 조회 API (오너 + 공유 사용자)"""
    from v1.label.models import MyLabel

    try:
        label = MyLabel.objects.get(my_label_id=label_id, user_id=request.user, delete_YN='N')
    except MyLabel.DoesNotExist:
        shared_share = ProductShare.objects.filter(
            label__my_label_id=label_id,
            active_yn=True
        ).filter(
            Q(recipient_user=request.user) | Q(recipient_email__iexact=request.user.email)
        ).filter(
            Q(share_end_date__isnull=True) | Q(share_end_date__gt=timezone.now())
        ).select_related('label').first()
        if not shared_share:
            from django.http import Http404
            raise Http404('접근 권한이 없습니다.')
        label = shared_share.label
    
    data = {
        # 기본 설정
        'serving_size': label.serving_size or '',
        'serving_size_unit': label.serving_size_unit or 'g',
        'units_per_package': label.units_per_package or '',
        'nutrition_display_unit': label.nutrition_display_unit or 'basic',
        'basic_display_type': label.basic_display_type or '',
        'parallel_display_type': label.parallel_display_type or '',
        
        # 필수 영양성분 9가지
        'calories': label.calories or '',
        'natriums': label.natriums or '',
        'carbohydrates': label.carbohydrates or '',
        'sugars': label.sugars or '',
        'fats': label.fats or '',
        'trans_fats': label.trans_fats or '',
        'saturated_fats': label.saturated_fats or '',
        'cholesterols': label.cholesterols or '',
        'proteins': label.proteins or '',
        
        # 추가 영양성분
        'dietary_fiber': label.dietary_fiber or '',
        'calcium': label.calcium or '',
        'iron': label.iron or '',
        'magnesium': label.magnesium or '',
        'phosphorus': label.phosphorus or '',
        'potassium': label.potassium or '',
        'zinc': label.zinc or '',
        'vitamin_a': label.vitamin_a or '',
        'vitamin_d': label.vitamin_d or '',
        'vitamin_c': label.vitamin_c or '',
        'thiamine': label.thiamine or '',
        'riboflavin': label.riboflavin or '',
        'niacin': label.niacin or '',
        'vitamin_b6': label.vitamin_b6 or '',
        'folic_acid': label.folic_acid or '',
        'vitamin_b12': label.vitamin_b12 or '',
    }
    
    return JsonResponse(data)


@login_required
@require_POST
def nutrition_save_api(request, label_id):
    """영양성분 데이터 저장 API (오너 + 쓰기 권한 공유 사용자)"""
    from v1.label.models import MyLabel
    import json

    try:
        label = MyLabel.objects.get(my_label_id=label_id, user_id=request.user, delete_YN='N')
    except MyLabel.DoesNotExist:
        shared_share = ProductShare.objects.filter(
            label__my_label_id=label_id,
            active_yn=True
        ).filter(
            Q(recipient_user=request.user) | Q(recipient_email__iexact=request.user.email)
        ).filter(
            Q(share_end_date__isnull=True) | Q(share_end_date__gt=timezone.now())
        ).select_related('label', 'permission').first()
        if not shared_share:
            return JsonResponse({'status': 'error', 'message': '접근 권한이 없습니다.'}, status=403)
        # can_edit_label 권한 확인 (오타 수정: can_edit → can_edit_label)
        perm = getattr(shared_share, 'permission', None)
        if not perm or not perm.can_edit_label:
            return JsonResponse({'status': 'error', 'message': '영양성분 편집 권한이 없습니다.'}, status=403)
        label = shared_share.label
    
    try:
        data = json.loads(request.body)
        
        # 기본 설정
        label.serving_size = data.get('serving_size', '') or '100'
        label.serving_size_unit = data.get('serving_size_unit', 'g') or 'g'
        label.units_per_package = data.get('units_per_package', '') or '1'
        label.nutrition_display_unit = data.get('nutrition_display_unit', 'basic') or 'basic'
        label.basic_display_type = data.get('basic_display_type', 'total') or 'total'
        label.parallel_display_type = data.get('parallel_display_type', 'unit_total') or 'unit_total'
        
        # 필수 영양성분 9가지
        label.calories = data.get('calories', '')
        label.natriums = data.get('natriums', '')
        label.carbohydrates = data.get('carbohydrates', '')
        label.sugars = data.get('sugars', '')
        label.fats = data.get('fats', '')
        label.trans_fats = data.get('trans_fats', '')
        label.saturated_fats = data.get('saturated_fats', '')
        label.cholesterols = data.get('cholesterols', '')
        label.proteins = data.get('proteins', '')
        
        # 추가 영양성분
        label.dietary_fiber = data.get('dietary_fiber', '')
        label.calcium = data.get('calcium', '')
        label.iron = data.get('iron', '')
        label.magnesium = data.get('magnesium', '')
        label.phosphorus = data.get('phosphorus', '')
        label.potassium = data.get('potassium', '')
        label.zinc = data.get('zinc', '')
        label.vitamin_a = data.get('vitamin_a', '')
        label.vitamin_d = data.get('vitamin_d', '')
        label.vitamin_c = data.get('vitamin_c', '')
        label.thiamine = data.get('thiamine', '')
        label.riboflavin = data.get('riboflavin', '')
        label.niacin = data.get('niacin', '')
        label.vitamin_b6 = data.get('vitamin_b6', '')
        label.folic_acid = data.get('folic_acid', '')
        label.vitamin_b12 = data.get('vitamin_b12', '')
        
        label.save(update_fields=[
            'serving_size', 'serving_size_unit', 'units_per_package',
            'nutrition_display_unit', 'basic_display_type', 'parallel_display_type',
            'calories', 'natriums', 'carbohydrates', 'sugars', 'fats',
            'trans_fats', 'saturated_fats', 'cholesterols', 'proteins',
            'dietary_fiber', 'calcium', 'iron', 'magnesium', 'phosphorus',
            'potassium', 'zinc', 'vitamin_a', 'vitamin_d', 'vitamin_c',
            'thiamine', 'riboflavin', 'niacin', 'vitamin_b6', 'folic_acid', 'vitamin_b12',
        ])
        
        return JsonResponse({'status': 'success', 'message': '저장되었습니다.'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


# ==================== 공유 상세 기능 (Stub) ====================

@login_required
def received_share_detail(request, receipt_id):
    """받은 공유 상세"""
    get_object_or_404(SharedProductReceipt, receipt_id=receipt_id, receiver=request.user)
    messages.info(request, '공유 상세 화면은 준비 중입니다.')
    return redirect('products:inbox')


@login_required
@require_POST
def received_share_accept(request, receipt_id):
    """공유 수락"""
    receipt = get_object_or_404(SharedProductReceipt, receipt_id=receipt_id, receiver=request.user)
    receipt.accepted_yn = True
    receipt.accepted_datetime = timezone.now()
    receipt.save()
    log_activity(request, 'sharing', 'share_accept', receipt_id)
    messages.success(request, '공유를 수락했습니다.')
    return redirect('products:inbox')


@login_required
@require_POST
def use_as_ingredient(request, receipt_id):
    """원료로 사용"""
    receipt = get_object_or_404(
        SharedProductReceipt.objects.select_related('share__label'),
        receipt_id=receipt_id,
        receiver=request.user,
    )
    receipt.used_as_ingredient_yn = True
    receipt.save()

    # 공유받은 제품을 내 원료 목록(MyIngredient)에 자동 등록
    from v1.label.models import MyIngredient as _MyIngredient
    label = receipt.share.label
    product_name = label.prdlst_nm or label.my_label_name
    existing = _MyIngredient.objects.filter(
        user_id=request.user,
        prdlst_nm=product_name,
        delete_YN='N',
    ).first()
    if not existing:
        _MyIngredient.objects.create(
            user_id=request.user,
            prdlst_nm=product_name,
            bssh_nm=label.bssh_nm or '',
            prdlst_dcnm=label.prdlst_dcnm or '',
            pog_daycnt=label.pog_daycnt or '',
            rawmtrl_nm=label.rawmtrl_nm or '',
            ingredient_display_name=label.rawmtrl_nm or product_name,
            food_category='processed',
            delete_YN='N',
        )

    log_activity(request, 'sharing', 'share_use_ingredient', receipt_id)
    messages.success(request, '원료로 등록했습니다.')
    return redirect('products:inbox')


def _get_editor_share_for_label(request, label):
    """요청 사용자가 해당 label에 EDITOR 역할로 공유받은 share를 반환. 없으면 None."""
    return ProductShare.objects.filter(
        label=label,
        active_yn=True,
        permission__role_code='EDITOR',
    ).filter(
        Q(recipient_user=request.user) | Q(recipient_email__iexact=request.user.email)
    ).filter(
        Q(share_end_date__isnull=True) | Q(share_end_date__gt=timezone.now())
    ).select_related('permission').first()


@login_required
@require_http_methods(["GET", "POST"])
def share_create(request, label_id):
    """공유 생성 – 소유자 또는 EDITOR 역할 공유자가 접근 가능"""
    # 소유자 확인
    label = MyLabel.objects.filter(
        my_label_id=label_id,
        user_id=request.user,
        delete_YN='N'
    ).first()

    if not label:
        # EDITOR 공유 접근 확인
        label = get_object_or_404(MyLabel, my_label_id=label_id, delete_YN='N')
        if not _get_editor_share_for_label(request, label):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied

    if request.method == 'GET':
        target_url = f"{reverse('products:product_detail', kwargs={'product_id': label_id})}#tab-share"
        return redirect(target_url)

    email = request.POST.get('email', '').strip()
    role_code = request.POST.get('role', 'VIEWER')
    expiration_date = request.POST.get('expiration_date')
    recipient_name = request.POST.get('name', '').strip()
    recipient_company = request.POST.get('company', '').strip()

    if not email:
        return JsonResponse({'success': False, 'error': '이메일은 필수입니다.'}, status=400)

    from django.core.validators import validate_email as _validate_email
    from django.core.exceptions import ValidationError as _DjValidationError
    try:
        _validate_email(email)
    except _DjValidationError:
        return JsonResponse({'success': False, 'error': '올바른 이메일 형식이 아닙니다.'}, status=400)

    if role_code not in dict(SharePermission.ROLE_CHOICES):
        return JsonResponse({'success': False, 'error': '잘못된 역할입니다.'}, status=400)

    # 이 API 는 PRIVATE(권한 관리) 공유만 만든다. PUBLIC 은 열람 화면이 미구현이다.
    if request.POST.get('share_mode', 'PRIVATE') != 'PRIVATE':
        return JsonResponse(
            {'success': False, 'error': '공개 링크 공유는 현재 지원하지 않습니다.'}, status=400)

    share_end_date = None
    if expiration_date:
        try:
            end_date = datetime.strptime(expiration_date, '%Y-%m-%d').date()
            share_end_date = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))
        except ValueError:
            return JsonResponse({'success': False, 'error': '만료일 형식이 올바르지 않습니다.'}, status=400)

    existing_share = ProductShare.objects.filter(
        label=label,
        recipient_email__iexact=email
    ).order_by('-created_datetime').first()

    recipient_user = User.objects.filter(email__iexact=email).first()

    if existing_share and existing_share.active_yn:
        # 이미 활성 공유 중 → 역할/만료일만 업데이트 (upsert)
        share = existing_share
        if recipient_name:
            share.recipient_name = recipient_name
        if recipient_company:
            share.recipient_company = recipient_company
        share.share_end_date = share_end_date
        share.save()
        permission, _ = SharePermission.objects.get_or_create(share=share)
        permission.apply_role_defaults(role_code=role_code, save=True)
        return JsonResponse({'success': True, 'updated': True, 'share_id': share.share_id})

    if existing_share and not existing_share.active_yn:
        share = existing_share
        share.active_yn = True
        share.share_mode = 'PRIVATE'  # 재활성화 시 모드를 PRIVATE으로 보장
        share.recipient_email = email
        share.recipient_user = recipient_user
        share.recipient_name = recipient_name
        share.recipient_company = recipient_company
        share.share_end_date = share_end_date
        share.created_by = request.user  # 재활성화 시 생성자를 현재 사용자로 업데이트
        share.save()
    else:
        share = ProductShare.objects.create(
            label=label,
            share_mode='PRIVATE',
            recipient_email=email,
            recipient_user=recipient_user,
            recipient_name=recipient_name,
            recipient_company=recipient_company,
            share_end_date=share_end_date,
            created_by=request.user
        )

    permission, _ = SharePermission.objects.get_or_create(share=share)
    permission.apply_role_defaults(role_code=role_code, save=True)

    if recipient_user:
        SharedProductReceipt.objects.get_or_create(share=share, receiver=recipient_user)

    # 연락처로도 등록해 둔다.
    # 연락처 목록은 '활성 공유 ∪ UserContact' 인데 여기서 UserContact 를 만들지 않아,
    # 공유를 해제하면 협력업체 연락처가 목록에서 사라졌다.
    UserContact.objects.update_or_create(
        owner=request.user,
        email=email,
        defaults={
            'name': recipient_name or None,
            'company': recipient_company or None,
        },
    )

    # 활동 로그 생성
    from .models import ProductActivityLog
    ProductActivityLog.objects.create(
        label=label,
        user=request.user,
        action='SHARE_CREATED',
        details={
            'recipient_email': email,
            'recipient_name': recipient_name,
            'role': role_code,
            'role_label': permission.role_label
        }
    )

    # ── 초대 이메일 발송 ──
    inviter_name, inviter_company = _get_sender_info(request.user)
    product_name = label.my_label_name or label.prdlst_nm or '제품'
    role_label   = permission.role_label
    from django.conf import settings as _ds
    inbox_url = f'{getattr(_ds, "SITE_URL", "https://labeldata.pythonanywhere.com")}/products/inbox/'
    _role_actions = {
        'UPLOADER': '공유받은 제품의 원료 규격서, 성분 데이터 등 요청 자료를 시스템에 업로드해 주세요.',
        'REVIEWER': '공유받은 제품의 라벨 내용을 검토하고 의견을 남겨주세요.',
        'EDITOR':   '공유받은 제품 정보를 확인하고 필요 시 수정·보완해 주세요.',
        'APPROVER': '검토된 내용을 확인하고 최종 승인 또는 반려해 주세요.',
    }
    _invite_subject = f'[EzLabeling] {inviter_name}님이 제품을 공유했습니다 — {product_name}'
    _txt, _html = _render_email('emails/product_invite.html', {
        'subject': _invite_subject,
        'sender_name': inviter_name, 'sender_company': inviter_company, 'sender_email': request.user.email,
        'product_name': product_name, 'role_label': role_label,
        'task_description': _role_actions.get(role_code, '시스템에 접속하여 내용을 확인해 주세요.'),
        'inbox_url': inbox_url,
    })
    _send_email_safe(subject=_invite_subject, body=_txt, to_email=email, html_body=_html)

    # ── 인앱 알림 (시스템 계정이 있는 경우) ──
    _create_notification(
        label=label,
        recipient_user=recipient_user,
        message=f'[{product_name}] {inviter_name}님이 {role_label} 역할로 초대했습니다.',
    )

    log_activity(request, 'sharing', 'share_create', label.my_label_id)
    return JsonResponse({'success': True, 'message': f'{email}님을 초대했습니다.'})


@login_required
def share_detail(request, share_id):
    """공유 상세"""
    messages.info(request, '공유 상세 기능은 준비 중입니다.')
    return redirect('products:inbox')


@login_required
@require_POST
def share_revoke(request, share_id):
    """공유 취소 – 소유자 또는 EDITOR(본인 제외) 접근 가능"""
    share = ProductShare.objects.filter(share_id=share_id, label__user_id=request.user).first()
    if not share:
        share = get_object_or_404(ProductShare, share_id=share_id)
        editor_share = _get_editor_share_for_label(request, share.label)
        if not editor_share:
            return JsonResponse({'success': False, 'error': '권한이 없습니다.'}, status=403)
        if share.share_id == editor_share.share_id:
            return JsonResponse({'success': False, 'error': '자신의 공유는 취소할 수 없습니다.'}, status=403)
    
    # 활동 로그 생성 (삭제 전에 정보 저장)
    from .models import ProductActivityLog
    ProductActivityLog.objects.create(
        label=share.label,
        user=request.user,
        action='SHARE_DELETED',
        details={
            'recipient_email': share.recipient_email,
            'recipient_name': share.recipient_name
        }
    )
    
    share.active_yn = False
    share.save(update_fields=['active_yn', 'updated_datetime'])
    return JsonResponse({'success': True})


@login_required
@require_POST
def share_update_permission(request, share_id):
    """공유 권한 수정 – 소유자 또는 EDITOR(본인 제외) 접근 가능"""
    share = ProductShare.objects.filter(share_id=share_id, label__user_id=request.user).first()
    if not share:
        share = get_object_or_404(ProductShare, share_id=share_id)
        editor_share = _get_editor_share_for_label(request, share.label)
        if not editor_share:
            return JsonResponse({'success': False, 'error': '권한이 없습니다.'}, status=403)
        if share.share_id == editor_share.share_id:
            return JsonResponse({'success': False, 'error': '자신의 권한은 변경할 수 없습니다.'}, status=403)
    role_code = request.POST.get('role', '').strip()

    if role_code not in dict(SharePermission.ROLE_CHOICES):
        return JsonResponse({'success': False, 'error': '잘못된 역할입니다.'}, status=400)

    permission, _ = SharePermission.objects.get_or_create(share=share)
    permission.apply_role_defaults(role_code=role_code, save=True)

    # 활동 로그 생성
    from .models import ProductActivityLog
    ProductActivityLog.objects.create(
        label=share.label,
        user=request.user,
        action='SHARE_UPDATED',
        details={
            'recipient_email': share.recipient_email,
            'recipient_name': share.recipient_name,
            'role': role_code,
            'role_label': permission.role_label,
            'change_type': 'permission'
        }
    )

    # ── 역할 변경 알림 (이메일 + 인앱) ──
    changer_name, changer_company = _get_sender_info(request.user)
    product_name = share.label.my_label_name or share.label.prdlst_nm or '제품'
    role_label   = permission.role_label
    notify_msg   = f'[{product_name}] 역할이 {role_label}(으)로 변경되었습니다.'
    from django.conf import settings as _ds
    _inbox_url = f'{getattr(_ds, "SITE_URL", "https://labeldata.pythonanywhere.com")}/products/inbox/'

    _role_subj = f'[EzLabeling] {product_name} 제품 권한이 변경되었습니다'
    _txt, _html = _render_email('emails/role_change.html', {
        'subject': _role_subj,
        'sender_name': changer_name, 'sender_company': changer_company, 'sender_email': request.user.email,
        'product_name': product_name, 'role_label': role_label,
        'inbox_url': _inbox_url,
    })
    _send_email_safe(subject=_role_subj, body=_txt, to_email=share.recipient_email, html_body=_html)
    _create_notification(
        label=share.label,
        recipient_user=share.recipient_user,
        message=notify_msg,
    )

    return JsonResponse({'success': True, 'role_label': permission.role_label})


@login_required
@require_http_methods(["POST"])
def share_update_info(request, share_id):
    """공유 멤버 정보 수정 – 소유자 또는 EDITOR(본인 제외) 접근 가능"""
    # 첫 번째 쿼리가 성공이면 라벨 오너, 실패하면 에디터 경로
    owner_share = ProductShare.objects.filter(share_id=share_id, label__user_id=request.user).first()
    is_label_owner = owner_share is not None

    if is_label_owner:
        share = owner_share
    else:
        share = get_object_or_404(ProductShare, share_id=share_id)
        editor_share = _get_editor_share_for_label(request, share.label)
        if not editor_share:
            return JsonResponse({'success': False, 'error': '권한이 없습니다.'}, status=403)
        if share.share_id == editor_share.share_id:
            return JsonResponse({'success': False, 'error': '자신의 정보는 변경할 수 없습니다.'}, status=403)

    name = request.POST.get('name', '').strip()
    company = request.POST.get('company', '').strip()
    license_no = request.POST.get('license_no', '').strip()
    role_code = request.POST.get('role', '').strip()

    # 이름·회사·인허가번호: 라벨 오너만 수정 가능
    if is_label_owner:
        share.recipient_name = name or None
        share.recipient_company = company or None
        share.recipient_license_no = license_no or None
    share.save()

    # 동일 이메일의 다른 공유 레코드에도 이름·회사·인허가번호 전파 (오너만)
    if is_label_owner:
        ProductShare.objects.filter(
            Q(label__user_id=share.label.user_id) | Q(created_by=share.created_by)
        ).filter(
            recipient_email__iexact=share.recipient_email,
            share_mode='PRIVATE',
            active_yn=True
        ).exclude(share_id=share.share_id).update(
            recipient_name=name or None,
            recipient_company=company or None,
            recipient_license_no=license_no or None,
        )
    
    # 역할 업데이트
    if role_code and role_code in dict(SharePermission.ROLE_CHOICES):
        permission, _ = SharePermission.objects.get_or_create(share=share)
        permission.apply_role_defaults(role_code=role_code, save=True)

    # 문서함 전체 보기 — 역할 기본값을 덮어쓴다.
    #
    # 자료 제출(협력업체)은 기본이 꺼짐이다. 다른 회사 서류가 같은 문서함에
    # 쌓이기 때문이다. 그래도 켜야 할 때가 있어(협력사가 검토도 맡는 경우)
    # 소유자가 여기서 켠다. **역할을 다시 고르면 기본값으로 돌아간다.**
    see_all = request.POST.get('see_all_documents', '')
    if see_all in ('true', 'false'):
        permission, _ = SharePermission.objects.get_or_create(share=share)
        permission.can_view_all_documents = (see_all == 'true')
        permission.save(update_fields=['can_view_all_documents'])
    
    # 활동 로그 생성
    from .models import ProductActivityLog
    detail_info = {
        'recipient_email': share.recipient_email,
        'name': name,
        'company': company,
        'change_type': 'info'
    }
    if role_code:
        detail_info['role'] = role_code
        detail_info['role_label'] = permission.role_label
    
    ProductActivityLog.objects.create(
        label=share.label,
        user=request.user,
        action='SHARE_UPDATED',
        details=detail_info
    )
    
    return JsonResponse({
        'success': True, 
        'message': '변경사항이 저장되었습니다.',
        'name': share.recipient_name,
        'company': share.recipient_company
    })


def public_share_view(request, share_token):
    """
    공개 링크 공유 — 미구현.

    ProductShare.share_mode 에 'PUBLIC' 선택지와 public_token 필드가 있지만 열람 화면이
    없다. 예전에는 "준비 중" 안내 후 제품 목록으로 보냈는데, 인증 없이 닿는 엔드포인트를
    열어두는 셈이라 404 로 닫는다.
    구현하려면 토큰 검증·만료·열람 범위 설계가 함께 필요하다.
    """
    raise Http404("공개 공유 링크는 제공하지 않습니다.")


# ==================== 문서 관리 ====================

# 문서 타입 자동 감지
def detect_document_type(filename):
    """파일명을 분석하여 적절한 문서 타입을 자동 감지"""
    all_types = DocumentType.objects.filter(active_yn=True).exclude(detection_keywords='')
    
    for dtype in all_types:
        if dtype.matches_filename(filename):
            return dtype
    
    # 분류 실패 시 '기타' 타입 반환 (없으면 자동 생성)
    other_type, created = DocumentType.objects.get_or_create(
        type_code='OTHER',
        defaults={
            'type_name': '기타',
            'icon': 'bi-file-earmark',
            'color': '#6c757d',
            'required_yn': False,
            'default_validity_days': 365,
            'detection_keywords': '',
            'active_yn': True
        }
    )
    return other_type


@login_required
def document_type_list(request):
    """문서 타입 목록 (관리자 전용 - Admin 사용 권장)"""
    if not request.user.is_staff:
        messages.warning(request, '문서 타입은 관리자만 조회할 수 있습니다. Django Admin을 이용해주세요.')
        return redirect('/admin/documents/documenttype/')
    
    types = DocumentType.objects.filter(active_yn=True).order_by('display_order', 'type_name')
    
    context = {
        'types': types,
    }
    
    return render(request, 'products/documents/type_list.html', context)


@login_required
def document_type_create(request):
    """문서 타입 생성 (Django Admin 사용 권장)"""
    if not request.user.is_staff:
        messages.error(request, '권한이 없습니다.')
        return redirect('products:document_type_list')
    
    messages.info(request, 'Django Admin에서 문서 타입을 관리해주세요.')
    return redirect('/admin/documents/documenttype/add/')


@login_required
def document_type_update(request, type_id):
    """문서 타입 수정 (Django Admin 사용 권장)"""
    if not request.user.is_staff:
        messages.error(request, '권한이 없습니다.')
        return redirect('products:document_type_list')
    
    messages.info(request, 'Django Admin에서 문서 타입을 수정해주세요.')
    return redirect(f'/admin/documents/documenttype/{type_id}/change/')


@login_required
def document_types_api(request):
    """문서 타입 목록 API (드롭존 생성용)"""
    types = DocumentType.objects.filter(active_yn=True).order_by('display_order')
    
    data = [{
        'type_id': t.type_id,
        'type_code': t.type_code,
        'type_name': t.type_name,
        'icon': t.icon,
        'color': t.color,
        'required_yn': t.required_yn,
        'default_validity_days': t.default_validity_days,
        'detection_keywords': t.detection_keywords,
    } for t in types]
    
    return JsonResponse({'types': data})


@login_required
@require_POST
def document_upload_api(request, label_id):
    """
    스마트 문서 업로드 API
    - document_type_id가 있으면 해당 타입으로 저장
    - slot_id가 있으면 해당 슬롯에 연결
    - 없으면 파일명 분석하여 자동 분류
    - 유효기간 자동 계산
    """
    import traceback
    from django.db.models import Max

    try:
        # 오너 우선 조회, 실패 시 공유 사용자 확인
        try:
            label = MyLabel.objects.get(my_label_id=label_id, user_id=request.user, delete_YN='N')
        except MyLabel.DoesNotExist:
            shared_share = ProductShare.objects.filter(
                label__my_label_id=label_id,
                active_yn=True,
            ).filter(
                Q(recipient_user=request.user) | Q(recipient_email__iexact=request.user.email)
            ).filter(
                Q(share_end_date__isnull=True) | Q(share_end_date__gt=timezone.now())
            ).select_related('label', 'permission').first()

            if not shared_share:
                return JsonResponse({'success': False, 'error': '접근 권한이 없습니다.'}, status=403)

            perm = getattr(shared_share, 'permission', None)
            if not perm or not perm.can_upload_documents:
                role_label = perm.role_code if perm else 'VIEWER'
                return JsonResponse({
                    'success': False,
                    'error': f'문서 업로드 권한이 없습니다. (현재 역할: {role_label})\n문서 업로드는 오너, 편집자, 자료 제출자만 가능합니다.'
                }, status=403)

            label = shared_share.label
        
        uploaded_file = request.FILES.get('file')
        document_type_id = request.POST.get('document_type') or request.POST.get('document_type_id')
        slot_id = request.POST.get('slot_id')  # 슬롯 ID 추가
        expiry_date_str = request.POST.get('expiry_date')
        expiry_unlimited = request.POST.get('expiry_unlimited', 'false') == 'true'
        description = request.POST.get('description', '')
        notification_enabled = request.POST.get('notification_enabled', 'false') == 'true'

        expiry_date = None
        if expiry_date_str:
            from datetime import datetime
            try:
                expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'error': '만료일 형식이 올바르지 않습니다.'
                }, status=400)
        
        # 파일 유효성 검사
        if not uploaded_file:
            return JsonResponse({
                'success': False,
                'error': '파일을 선택해주세요.'
            }, status=400)
        
        # 파일 크기 제한 (50MB)
        if uploaded_file.size > 50 * 1024 * 1024:
            return JsonResponse({
                'success': False,
                'error': '파일 크기는 50MB를 초과할 수 없습니다.'
            }, status=400)
        
        # 슬롯 정보 조회 (있는 경우)
        slot = None
        if slot_id:
            from .models import DocumentSlot
            slot = get_object_or_404(DocumentSlot, slot_id=slot_id, label=label)
            # 슬롯이 있으면 document_type은 슬롯의 타입 사용
            document_type = slot.document_type
        else:
            # 문서 타입 결정 (지정 vs 자동 분류)
            auto_detected = False
            if document_type_id:
                document_type = get_object_or_404(DocumentType, type_id=document_type_id, active_yn=True)
            else:
                # 파일명 기반 자동 분류
                document_type = detect_document_type(uploaded_file.name)
                auto_detected = True
                
                if not document_type:
                    return JsonResponse({
                        'success': False,
                        'error': '문서 타입을 자동으로 분류할 수 없습니다. 직접 선택해주세요.'
                    }, status=400)
        
        # 기존 문서가 슬롯에 있는 경우 버전 관리
        parent_document = None
        version_number = 1
        if slot and slot.current_document:
            parent_document = slot.current_document
            # 같은 부모의 최신 버전 번호 찾기
            latest_version = ProductDocument.objects.filter(
                Q(document_id=parent_document.document_id) | Q(parent_document=parent_document)
            ).aggregate(Max('version'))['version__max'] or 1
            version_number = latest_version + 1
        elif not slot:
            # 슬롯 없이 업로드 시 같은 label+document_type 기존 문서 중 최신 버전 처리
            existing_root = ProductDocument.objects.filter(
                label=label,
                document_type=document_type,
                parent_document__isnull=True,
                active_yn=True,
            ).order_by('-version', '-uploaded_datetime').first()
            if existing_root:
                parent_document = existing_root
                latest_version = ProductDocument.objects.filter(
                    Q(document_id=existing_root.document_id) | Q(parent_document=existing_root)
                ).aggregate(Max('version'))['version__max'] or 1
                version_number = latest_version + 1
        
        # 문서 생성
        metadata = {'expiry_unlimited': True} if expiry_unlimited else {}

        document = ProductDocument.objects.create(
            label=label,
            document_type=document_type,
            slot=slot,  # 슬롯 연결
            file=uploaded_file,
            original_filename=uploaded_file.name,
            file_size=uploaded_file.size,
            expiry_date=expiry_date if expiry_date else None,
            description=description,
            uploaded_by=request.user,
            parent_document=parent_document,
            version=version_number,
            expiry_notification_enabled=notification_enabled,
            metadata=metadata
        )
        
        # 슬롯 업데이트 (current_document 설정 및 상태 업데이트)
        if slot:
            slot.current_document = document
            slot.save()  # save()에서 update_status() 호출
        
        # 활동 로그 생성
        from .models import ProductActivityLog
        log_details = {
            'file_name': uploaded_file.name,
            'document_type': document_type.type_name,
            'file_size': uploaded_file.size
        }
        if slot:
            log_details['slot_id'] = slot.slot_id
            log_details['action_type'] = 'slot_upload'
        if version_number > 1:
            log_details['version'] = version_number
            log_details['previous_version'] = version_number - 1
        
        ProductActivityLog.objects.create(
            label=label,
            user=request.user,
            action='DOCUMENT_UPLOADED',
            details=log_details
        )
        
        # 응답 메시지 구성
        if slot:
            message = f"'{document_type.type_name}' 슬롯에 문서가 업로드되었습니다."
            if version_number > 1:
                message += f" (버전 {version_number})"
        elif 'auto_detected' in locals() and auto_detected:
            message = f"'{document_type.type_name}'(으)로 자동 분류되었습니다."
        else:
            message = '문서가 업로드되었습니다.'
        
        response_data = {
            'success': True,
            'document_id': document.document_id,
            'filename': document.original_filename,
            'file_size': document.file_size,
            'document_type': document.document_type.type_name,
            'document_type_id': document.document_type.type_id,
            'expiry_date': document.expiry_date.isoformat() if document.expiry_date else None,
            'uploaded_at': document.uploaded_datetime.strftime('%Y-%m-%d %H:%M'),
            'message': message,
            'version': version_number
        }
        
        # 슬롯 정보 추가
        if slot:
            response_data['slot'] = {
                'slot_id': slot.slot_id,
                'status': slot.status,
                'document_type': slot.document_type.type_name
            }
        elif 'auto_detected' in locals():
            response_data['auto_detected'] = auto_detected
        
        return JsonResponse(response_data)
        
    except Exception as e:
        # 상세한 에러 로깅
        error_trace = traceback.format_exc()
        print(f"[Document Upload Error] {error_trace}")
        
        return JsonResponse({
            'success': False,
            'error': f'업로드 중 오류가 발생했습니다: {str(e)}'
        }, status=500)


@login_required
@require_POST
def company_document_import_api(request, label_id):
    """
    고정 서류 관리에서 제품 문서로 불러오기 API
    - company_document_id: CompanyDocument PK
    - document_type_id: 대상 DocumentType (없으면 파일명으로 자동 감지)
    - slot_id: 연결할 슬롯 (선택)
    """
    import shutil
    import os
    from django.core.files import File
    from v1.user_management.models import CompanyDocument

    try:
        label = MyLabel.objects.get(my_label_id=label_id, user_id=request.user, delete_YN='N')
    except MyLabel.DoesNotExist:
        return JsonResponse({'success': False, 'error': '제품을 찾을 수 없거나 권한이 없습니다.'}, status=403)

    company_doc_id = request.POST.get('company_document_id')
    if not company_doc_id:
        return JsonResponse({'success': False, 'error': 'company_document_id가 필요합니다.'}, status=400)

    try:
        company_doc = CompanyDocument.objects.get(pk=company_doc_id, user=request.user)
    except CompanyDocument.DoesNotExist:
        return JsonResponse({'success': False, 'error': '고정 서류를 찾을 수 없습니다.'}, status=404)

    document_type_id = request.POST.get('document_type_id')
    slot_id = request.POST.get('slot_id')

    # 문서 타입 결정
    slot = None
    if slot_id:
        try:
            slot = DocumentSlot.objects.get(slot_id=slot_id, label=label)
        except DocumentSlot.DoesNotExist:
            return JsonResponse({'success': False, 'error': '슬롯을 찾을 수 없습니다.'}, status=400)
        document_type = slot.document_type
    elif document_type_id:
        document_type = DocumentType.objects.filter(type_id=document_type_id, active_yn=True).first()
        if not document_type:
            # 지정된 타입이 없으면 자동 감지로 fallback
            document_type = detect_document_type(company_doc.doc_file.name)
    else:
        document_type = detect_document_type(company_doc.doc_file.name)

    if not document_type:
        return JsonResponse({
            'success': False,
            'error': '문서 타입을 자동으로 분류할 수 없습니다. 직접 선택해주세요.',
        }, status=400)

    # 파일 복사하여 ProductDocument 생성
    original_name = os.path.basename(company_doc.doc_file.name)
    try:
        company_doc.doc_file.open('rb')
        file_content = company_doc.doc_file.read()
        company_doc.doc_file.close()
    except Exception:
        return JsonResponse({'success': False, 'error': '파일을 읽을 수 없습니다. 파일이 존재하는지 확인해주세요.'}, status=400)

    from io import BytesIO
    with BytesIO(file_content) as buf:
        document = ProductDocument.objects.create(
            label=label,
            document_type=document_type,
            slot=slot,
            file=File(buf, name=original_name),
            original_filename=original_name,
            file_size=company_doc.doc_file.size,
            document_title=company_doc.doc_name,
            uploaded_by=request.user,
            source_company_document=company_doc,
        )

    if slot:
        slot.current_document = document
        slot.save()
    else:
        # slot이 지정되지 않은 경우, document_type에 맞는 기존 슬롯 자동 연결
        auto_slot = DocumentSlot.objects.filter(
            label=label,
            document_type=document_type,
            hidden_yn=False,
        ).first()
        if auto_slot:
            auto_slot.current_document = document
            auto_slot.status = DocumentSlot.SlotStatus.ACTIVE
            auto_slot.save(update_fields=['current_document', 'status'])

    from .models import ProductActivityLog
    ProductActivityLog.objects.create(
        label=label,
        user=request.user,
        action='DOCUMENT_UPLOADED',
        details={
            'file_name': original_name,
            'document_type': document_type.type_name,
            'source': 'company_document',
            'company_document_id': company_doc.pk,
        },
    )

    return JsonResponse({
        'success': True,
        'document_id': document.document_id,
        'filename': document.original_filename,
        'document_type': document.document_type.type_name,
        'uploaded_at': document.uploaded_datetime.strftime('%Y-%m-%d %H:%M'),
        'message': f'고정 서류 "{company_doc.doc_name}"를 제품 문서로 불러왔습니다.',
    })


@login_required
@require_POST
def document_delete_api(request, document_id):
    """문서 삭제 AJAX API - JSON 응답 반환"""
    try:
        document = get_object_or_404(
            ProductDocument,
            document_id=document_id,
            label__user_id=request.user,
            active_yn=True
        )
        
        # 삭제 전 정보 저장
        file_name = document.original_filename
        document_type_name = document.document_type.type_name if document.document_type else '미분류'
        label = document.label
        
        # Soft Delete
        document.active_yn = False
        document.save()
        
        # 활동 로그 생성
        from .models import ProductActivityLog
        ProductActivityLog.objects.create(
            label=label,
            user=request.user,
            action='DOCUMENT_DELETED',
            details={
                'file_name': file_name,
                'document_type': document_type_name
            }
        )
        
        return JsonResponse({
            'success': True,
            'message': '문서가 삭제되었습니다.'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'삭제 중 오류가 발생했습니다: {str(e)}'
        }, status=500)


@login_required
@require_POST
def bulk_download(request):
    """
    선택한 문서들을 ZIP으로 일괄 다운로드
    
    POST 데이터:
    - document_ids: 다운로드할 문서 ID 목록 (comma-separated 또는 JSON 배열)
    - organize_by: 폴더 구성 방식 ('product', 'type', 'flat')
    """
    import json
    
    # 문서 ID 파싱
    document_ids_raw = request.POST.get('document_ids', '')
    organize_by = request.POST.get('organize_by', 'product')
    
    if not document_ids_raw:
        return JsonResponse({
            'success': False,
            'error': '다운로드할 문서를 선택해주세요.'
        }, status=400)
    
    # JSON 배열 또는 콤마 구분 처리
    try:
        if document_ids_raw.startswith('['):
            document_ids = json.loads(document_ids_raw)
        else:
            document_ids = [int(id.strip()) for id in document_ids_raw.split(',') if id.strip()]
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({
            'success': False,
            'error': '잘못된 문서 ID 형식입니다.'
        }, status=400)
    
    if not document_ids:
        return JsonResponse({
            'success': False,
            'error': '다운로드할 문서를 선택해주세요.'
        }, status=400)
    
    # 문서 조회 (사용자 소유 확인)
    documents = ProductDocument.objects.filter(
        Q(label__user_id=request.user) | Q(label__my_label_id__in=downloadable_label_ids(request.user)),
        document_id__in=document_ids,
        active_yn=True
    ).select_related('label', 'document_type')
    
    if not documents.exists():
        return JsonResponse({
            'success': False,
            'error': '다운로드 가능한 문서가 없습니다.'
        }, status=404)
    
    # ZIP 파일 생성
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        used_filenames = {}  # 중복 파일명 처리용
        
        for doc in documents:
            try:
                # 파일 경로 구성
                if organize_by == 'product':
                    product_name = _sanitize_filename(doc.label.my_label_name or doc.label.prdlst_nm or '제품')
                    folder = product_name
                elif organize_by == 'type':
                    type_name = _sanitize_filename(doc.document_type.type_name)
                    folder = type_name
                else:
                    folder = ''
                
                # 파일명 생성 (중복 처리)
                original_name = doc.original_filename
                
                if organize_by == 'type':
                    display_name = f"{_sanitize_filename(doc.label.my_label_name or doc.label.prdlst_nm or '제품')}_{original_name}"
                elif organize_by == 'flat':
                    display_name = f"{_sanitize_filename(doc.label.my_label_name or doc.label.prdlst_nm or '제품')}_{_sanitize_filename(doc.document_type.type_name)}_{original_name}"
                else:
                    display_name = original_name
                
                # 전체 경로
                if folder:
                    arcname = f"{folder}/{display_name}"
                else:
                    arcname = display_name
                
                # 중복 파일명 처리
                if arcname in used_filenames:
                    used_filenames[arcname] += 1
                    base, ext = os.path.splitext(arcname)
                    arcname = f"{base}_{used_filenames[arcname]}{ext}"
                else:
                    used_filenames[arcname] = 0
                
                # 파일 추가
                zip_file.writestr(arcname, doc.file.read())
                
            except Exception as e:
                continue
    
    # ZIP 응답
    zip_buffer.seek(0)
    
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    zip_filename = f"documents_{timestamp}.zip"
    
    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
    
    return response


@login_required
@require_POST
def bulk_download_version(request, label_id):
    """특정 표시사항의 모든 문서를 ZIP으로 다운로드 (소유자 + 다운로드 권한 공유자)"""
    label = get_object_or_404(MyLabel, my_label_id=label_id)
    if not user_can_download_label_files(request.user, label):
        raise Http404("표시사항을 찾을 수 없습니다.")
    
    # 한 건씩 묻는 자리가 아니다. 받을 수 있는 것만 담는다 —
    # 못 받는 것이 섞이면 ZIP 이 그대로 새는 통로가 된다.
    documents = visible_documents(request.user, label, ProductDocument.objects.filter(
        label=label,
        active_yn=True
    ).select_related('document_type'))
    
    if not documents.exists():
        return JsonResponse({
            'success': False,
            'error': '다운로드 가능한 문서가 없습니다.'
        }, status=404)
    
    # ZIP 파일 생성
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        used_filenames = {}
        
        for doc in documents:
            try:
                type_name = _sanitize_filename(doc.document_type.type_name)
                original_name = doc.original_filename
                
                arcname = f"{type_name}/{original_name}"
                
                # 중복 처리
                if arcname in used_filenames:
                    used_filenames[arcname] += 1
                    base, ext = os.path.splitext(arcname)
                    arcname = f"{base}_{used_filenames[arcname]}{ext}"
                else:
                    used_filenames[arcname] = 0
                
                zip_file.writestr(arcname, doc.file.read())
                
            except Exception:
                continue
    
    zip_buffer.seek(0)
    
    product_name = _sanitize_filename(label.my_label_name or label.prdlst_nm or '제품')
    timestamp = timezone.now().strftime('%Y%m%d')
    zip_filename = f"{product_name}_{timestamp}.zip"
    
    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
    
    return response


@login_required
def document_detail(request, document_id):
    """문서 상세"""
    document = get_object_or_404(
        ProductDocument.objects.select_related('label', 'document_type', 'uploaded_by'),
        document_id=document_id,
    )
    if not user_can_download_label_files(request.user, document.label, document):
        raise Http404("문서를 찾을 수 없습니다.")
    
    context = {
        'document': document,
    }
    
    return render(request, 'products/documents/document_detail.html', context)


@login_required
def document_download(request, document_id):
    """
    문서 다운로드 — 소유자 또는 다운로드 권한이 있는 공유자.

    이전에는 label__user_id=request.user 로 소유자만 허용해, can_download_documents=True
    인 검토자·협력사도 404 를 받았다. 제출 자료를 못 보면 검토 단계가 성립하지 않는다.
    """
    document = get_object_or_404(
        ProductDocument,
        document_id=document_id,
        active_yn=True
    )
    if not user_can_download_label_files(request.user, document.label, document):
        raise Http404("문서를 찾을 수 없습니다.")
    
    try:
        return FileResponse(
            document.file.open('rb'),
            as_attachment=True,
            filename=document.original_filename
        )
    except FileNotFoundError:
        raise Http404("파일을 찾을 수 없습니다.")


@login_required
def expiring_documents(request):
    """만료 예정 문서 목록"""
    today = timezone.now().date()
    alert_date = today + timedelta(days=30)
    
    documents = ProductDocument.objects.filter(
        label__user_id=request.user,
        active_yn=True,
        expiry_date__isnull=False,
        expiry_date__gte=today,
        expiry_date__lte=alert_date
    ).select_related('label', 'document_type').order_by('expiry_date')
    
    context = {
        'documents': documents,
        'today': today,
    }
    
    return render(request, 'products/documents/expiring_documents.html', context)


@login_required
def expired_documents(request):
    """만료된 문서 목록"""
    today = timezone.now().date()
    
    documents = ProductDocument.objects.filter(
        label__user_id=request.user,
        active_yn=True,
        expiry_date__isnull=False,
        expiry_date__lt=today
    ).select_related('label', 'document_type').order_by('-expiry_date')
    
    context = {
        'documents': documents,
        'today': today,
    }
    
    return render(request, 'products/documents/expired_documents.html', context)


def _sanitize_filename(name):
    """파일/폴더명에 사용할 수 없는 문자 제거"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name.strip()[:50]  # 50자 제한


# ==================== 협업 기능 (Stub) ====================

def _get_label_access(request, label_id):
    """댓글용 접근 권한 확인 (오너 또는 공유 사용자)"""
    try:
        label = MyLabel.objects.get(my_label_id=label_id, user_id=request.user)
        return label, 'OWNER'
    except MyLabel.DoesNotExist:
        shared_share = ProductShare.objects.filter(
            label__my_label_id=label_id,
            active_yn=True
        ).filter(
            Q(recipient_user=request.user) | Q(recipient_email__iexact=request.user.email)
        ).filter(
            Q(share_end_date__isnull=True) | Q(share_end_date__gt=timezone.now())
        ).select_related('label').first()

        if not shared_share:
            return None, None

        share_permission = SharePermission.objects.filter(share=shared_share).first()
        role = share_permission.role_code if share_permission else 'VIEWER'
        return shared_share.label, role

@login_required
def comment_list(request, label_id):
    """댓글 목록 API"""
    label, role = _get_label_access(request, label_id)
    if not label:
        return JsonResponse({'success': False, 'error': '접근 권한이 없습니다.'}, status=403)

    comments = ProductComment.objects.filter(label=label, resolved_yn=False).select_related('author').order_by('-created_at')
    payload = []
    for comment in comments:
        payload.append({
            'comment_id': comment.comment_id,
            'author': comment.author.username,
            'content': comment.content,
            'created_at': comment.created_at.isoformat()
        })

    return JsonResponse({'success': True, 'comments': payload})


@login_required
@require_POST
def comment_create(request, label_id):
    """댓글 생성 API"""
    label, role = _get_label_access(request, label_id)
    if not label:
        return JsonResponse({'success': False, 'error': '접근 권한이 없습니다.'}, status=403)

    if role not in ['OWNER', 'UPLOADER', 'EDITOR', 'REVIEWER', 'APPROVER']:
        return JsonResponse({'success': False, 'error': '댓글 작성 권한이 없습니다.'}, status=403)

    content = request.POST.get('content', '').strip()
    field_name = request.POST.get('field_name', '').strip() or None
    parent_id = request.POST.get('parent_id')
    if not content:
        return JsonResponse({'success': False, 'error': '댓글 내용을 입력해주세요.'}, status=400)

    parent_comment = None
    if parent_id:
        parent_comment = get_object_or_404(ProductComment, comment_id=parent_id, label=label)

    comment = ProductComment.objects.create(
        label=label,
        author=request.user,
        content=content,
        field_name=field_name,
        parent=parent_comment
    )
    
    # 활동 로그 기록
    from .models import ProductActivityLog
    ProductActivityLog.objects.create(
        label=label,
        user=request.user,
        action='COMMENT_ADDED',
        details={
            'comment_id': comment.comment_id,
            'field_name': field_name,
            'content': content[:100],  # 처음 100자만 저장
            'is_reply': parent_id is not None,
        }
    )

    display_name = comment.author.get_full_name().strip() or comment.author.email

    return JsonResponse({
        'success': True,
        'comment': {
            'comment_id': comment.comment_id,
            'author': comment.author.username,
            'display_name': display_name,
            'email': comment.author.email,
            'content': comment.content,
            'created_at': comment.created_at.isoformat(),
            'field_name': comment.field_name,
            'parent_id': comment.parent_id
        }
    })


@login_required
@require_POST
def comment_resolve(request, comment_id):
    """댓글 해결 API"""
    comment = get_object_or_404(ProductComment, comment_id=comment_id)
    if comment.author != request.user and comment.label.user_id != request.user:
        return JsonResponse({'success': False, 'error': '권한이 없습니다.'}, status=403)

    comment.resolve(request.user)
    return JsonResponse({'success': True})


@login_required
@require_POST
def comment_delete(request, comment_id):
    """댓글 삭제 API"""
    comment = get_object_or_404(ProductComment, comment_id=comment_id)
    if comment.author != request.user and comment.label.user_id != request.user:
        return JsonResponse({'success': False, 'error': '권한이 없습니다.'}, status=403)

    comment.delete()
    return JsonResponse({'success': True})


@login_required
@require_POST
def bulk_delete_documents(request):
    """문서 일괄 삭제"""
    import json
    from .models import ProductActivityLog
    
    try:
        data = json.loads(request.body)
        document_ids = data.get('document_ids', [])
        
        if not document_ids:
            return JsonResponse({
                'success': False,
                'error': '삭제할 문서를 선택해주세요.'
            }, status=400)
        
        # 권한 확인 및 삭제
        documents = ProductDocument.objects.filter(
            document_id__in=document_ids,
            label__user_id=request.user,
            active_yn=True
        )
        
        deleted_count = 0
        for document in documents:
            file_name = document.original_filename
            document_type_name = document.document_type.type_name if document.document_type else '미분류'
            label = document.label
            
            # Soft Delete
            document.active_yn = False
            document.save()
            
            # 활동 로그 생성
            ProductActivityLog.objects.create(
                label=label,
                user=request.user,
                action='DOCUMENT_DELETED',
                details={
                    'file_name': file_name,
                    'document_type': document_type_name
                }
            )
            
            deleted_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'{deleted_count}개의 문서가 삭제되었습니다.'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'삭제 중 오류가 발생했습니다: {str(e)}'
        }, status=500)


@login_required
@require_POST
def toggle_document_notification(request, document_id):
    """문서 만료일 알림 토글"""
    import json
    
    try:
        document = get_object_or_404(
            ProductDocument,
            document_id=document_id,
            label__user_id=request.user,
            active_yn=True
        )
        
        # 알림 토글
        document.expiry_notification_enabled = not document.expiry_notification_enabled
        document.save()
        
        return JsonResponse({
            'success': True,
            'enabled': document.expiry_notification_enabled
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def document_versions(request, document_id):
    """문서 버전 이력 조회"""
    try:
        document = get_object_or_404(
            ProductDocument,
            document_id=document_id,
            label__user_id=request.user
        )
        
        # 현재 문서의 모든 버전 조회 (parent_document가 같거나 자기 자신이 parent인 경우)
        if document.parent_document:
            parent_id = document.parent_document.document_id
        else:
            parent_id = document.document_id
        
        versions = ProductDocument.objects.filter(
            Q(document_id=parent_id) |
            Q(parent_document_id=parent_id)
        ).order_by('-version')
        
        version_list = []
        for v in versions:
            version_list.append({
                'version': v.version,
                'filename': v.original_filename,
                'uploaded_by': v.uploaded_by.username if v.uploaded_by else '알 수 없음',
                'uploaded_date': v.uploaded_datetime.strftime('%Y-%m-%d %H:%M'),
                'file_url': f'/v2/products/documents/{v.document_id}/',
                'is_current': v.document_id == document.document_id
            })
        
        return JsonResponse({
            'success': True,
            'versions': version_list
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_POST
def document_update(request, document_id):
    """문서 메타데이터 업데이트 (구분, 발행일, 만료일, 설명)"""
    import json
    
    try:
        document = get_object_or_404(
            ProductDocument,
            document_id=document_id,
            label__user_id=request.user
        )
        
        data = json.loads(request.body)
        
        # 문서 타입 변경 (슬롯 재연결 포함)
        if 'document_type_id' in data:
            old_document_type = document.document_type
            new_document_type = get_object_or_404(DocumentType, type_id=data['document_type_id'])
            
            if old_document_type != new_document_type:
                from .models import DocumentSlot
                
                # 기존 슬롯 저장
                old_slot = document.slot
                
                # 새 타입으로 변경
                document.document_type = new_document_type
                
                # 슬롯 재연결
                if new_document_type.required_yn:
                    # 필수 문서: 해당 타입의 슬롯 찾기/생성
                    new_slot, created = DocumentSlot.objects.get_or_create(
                        label=document.label,
                        document_type=new_document_type
                    )
                    document.slot = new_slot
                else:
                    # 일반 문서: 슬롯 연결 해제
                    document.slot = None
                
                # 문서 저장
                document.save()
                
                # 기존 슬롯 상태 업데이트
                if old_slot:
                    # current_document가 이 문서였다면 교체
                    if old_slot.current_document == document:
                        # 같은 슬롯의 다른 활성 문서 찾기
                        replacement = ProductDocument.objects.filter(
                            slot=old_slot,
                            active_yn=True
                        ).exclude(document_id=document.document_id).order_by('-uploaded_datetime').first()
                        old_slot.current_document = replacement
                    old_slot.update_status()
                    old_slot.save()
                
                # 새 슬롯 상태 업데이트
                if document.slot:
                    # 이 문서가 가장 최신인지 확인
                    latest_doc = ProductDocument.objects.filter(
                        slot=document.slot,
                        active_yn=True
                    ).order_by('-uploaded_datetime').first()
                    
                    if latest_doc == document:
                        document.slot.current_document = document
                    document.slot.update_status()
                    document.slot.save()
            else:
                document.document_type = new_document_type
        
        # 발행일 변경
        if 'issue_date' in data:
            if data['issue_date']:
                from datetime import datetime
                document.issue_date = datetime.strptime(data['issue_date'], '%Y-%m-%d').date()
            else:
                document.issue_date = None
        
        # 만료일 변경
        if 'expiry_date' in data:
            if data['expiry_date']:
                from datetime import datetime
                document.expiry_date = datetime.strptime(data['expiry_date'], '%Y-%m-%d').date()
            else:
                document.expiry_date = None
        
        # 설명 변경
        if 'description' in data:
            document.description = data['description']
        
        # 만료 알림 설정 변경
        if 'expiry_notification_enabled' in data:
            document.expiry_notification_enabled = data['expiry_notification_enabled']
        
        document.save()
        
        # 연결된 슬롯이 있으면 상태 업데이트
        if document.slot:
            document.slot.update_status()
            document.slot.save()
        
        return JsonResponse({
            'success': True,
            'message': '문서 정보가 업데이트되었습니다.',
            'document': {
                'document_id': document.document_id,
                'document_type': document.document_type.type_name,
                'issue_date': document.issue_date.strftime('%Y-%m-%d') if document.issue_date else None,
                'expiry_date': document.expiry_date.strftime('%Y-%m-%d') if document.expiry_date else None,
                'description': document.description,
                'expiry_notification_enabled': document.expiry_notification_enabled
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_POST
def toggle_slot_visibility(request, slot_id):
    """슬롯 숨김/표시 토글"""
    from .models import DocumentSlot
    
    try:
        slot = get_object_or_404(
            DocumentSlot,
            slot_id=slot_id,
            label__user_id=request.user
        )
        
        # 필수 문서는 숨길 수 없음
        if slot.document_type.required_yn:
            return JsonResponse({
                'success': False,
                'error': '필수 문서 슬롯은 숨길 수 없습니다.'
            }, status=400)
        
        # 숨김 상태 토글
        slot.hidden_yn = not slot.hidden_yn
        slot.save()
        
        # 활동 로그
        from .models import ProductActivityLog
        ProductActivityLog.objects.create(
            label=slot.label,
            user=request.user,
            action='SLOT_VISIBILITY_CHANGED',
            details={
                'slot_id': slot.slot_id,
                'document_type': slot.document_type.type_name,
                'hidden_yn': slot.hidden_yn
            }
        )
        
        return JsonResponse({
            'success': True,
            'hidden_yn': slot.hidden_yn
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_POST
def add_document_slot(request, label_id):
    """문서 슬롯 추가 (필수 문서 확장)"""
    import json

    try:
        label = get_object_or_404(
            MyLabel,
            my_label_id=label_id,
            user_id=request.user
        )

        data = json.loads(request.body)
        document_type_id = data.get('document_type_id')
        if not document_type_id:
            return JsonResponse({
                'success': False,
                'error': '문서 종류를 선택해주세요.'
            }, status=400)

        document_type = get_object_or_404(DocumentType, type_id=document_type_id, active_yn=True)

        slot = DocumentSlot.objects.filter(label=label, document_type=document_type).first()
        if slot:
            if slot.hidden_yn:
                slot.hidden_yn = False
                slot.save()
                message = '숨겨진 문서가 다시 표시됩니다.'
            else:
                message = '이미 추가된 문서입니다.'
        else:
            slot = DocumentSlot.objects.create(label=label, document_type=document_type)
            message = '문서가 추가되었습니다.'

        return JsonResponse({
            'success': True,
            'message': message,
            'slot_id': slot.slot_id,
            'document_type': document_type.type_name
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'문서 추가 중 오류가 발생했습니다: {str(e)}'
        }, status=500)


@login_required
@require_POST
def remove_document_slot(request, slot_id):
    """문서 슬롯 삭제 (숨김 처리 — 필수 문서 포함)"""
    try:
        slot = get_object_or_404(
            DocumentSlot,
            slot_id=slot_id,
            label__user_id=request.user
        )

        slot.hidden_yn = True
        slot.save()

        from .models import ProductActivityLog
        ProductActivityLog.objects.create(
            label=slot.label,
            user=request.user,
            action='SLOT_REMOVED',
            details={
                'slot_id': slot.slot_id,
                'document_type': slot.document_type.type_name,
                'was_required': slot.document_type.required_yn,
            }
        )

        return JsonResponse({
            'success': True,
            'message': f'"{slot.document_type.type_name}" 슬롯이 제거되었습니다.'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'슬롯 삭제 중 오류가 발생했습니다: {str(e)}'
        }, status=500)


# ==================== 알림 (Notifications) ====================

@login_required
def notification_list(request):
    """읽지 않은 알림 목록 + 최근 알림 반환 (JSON) - 부적합.처분 알림 통합"""
    qs = ProductNotification.objects.filter(
        recipient=request.user
    ).select_related('label').order_by('-created_at')[:30]

    unread_count = ProductNotification.objects.filter(
        recipient=request.user, read_yn=False
    ).count()

    items = []
    for n in qs:
        items.append({
            'id': n.id,
            'message': n.message,
            'read_yn': n.read_yn,
            'status_code': n.status_code,
            'created_at': n.created_at.strftime('%Y-%m-%d %H:%M'),
            'label_id': n.label_id,
            'label_name': (n.label.my_label_name or n.label.prdlst_nm or '') if n.label else '',
            'url': f'/products/{n.label.my_label_id}/' if n.label else '',
        })

    # ── 부적합.처분 알림 통합 ───────────────────────────────────────────
    try:
        from v1.regulatory.models import NewsProductMatch, NewsIngredientMatch

        # 읽지 않은 규제 알림 건수 (read_yn 기반 — "모두 읽음" 처리와 동일 기준)
        prod_unread = set(
            NewsProductMatch.objects.filter(
                product__user_id=request.user, false_positive_yn=False, read_yn=False,
            ).values_list('news_id', flat=True)
        )
        ing_unread = set(
            NewsIngredientMatch.objects.filter(
                user=request.user, dismissed_yn=False, read_yn=False,
            ).values_list('news_id', flat=True)
        )
        reg_unread = len(prod_unread | ing_unread)
        unread_count += reg_unread

        # 뉴스 단위로 중복 제거 (같은 뉴스에 여러 제품 매칭돼도 1건만 표시)
        seen_news_ids = set()
        reg_matches = (
            NewsProductMatch.objects
            .filter(product__user_id=request.user, false_positive_yn=False)
            .select_related('news', 'product')
            .order_by('-created_at')
        )
        for m in reg_matches:
            if m.news_id in seen_news_ids:
                continue
            seen_news_ids.add(m.news_id)
            items.append({
                'id':          f'reg_{m.id}',
                'message': (
                    f"⚠️ 부적합.처분 알림: '{m.news.product_name[:20]}' 부적합 — "
                    f"내 제품 '{m.product.my_label_name[:20]}' 연관 가능성"
                ),
                'read_yn':     m.read_yn,
                'status_code': 'REGULATORY',
                'created_at':  m.created_at.strftime('%Y-%m-%d %H:%M'),
                'label_id':    m.product.my_label_id,
                'label_name':  m.product.my_label_name or '',
                'url':         f'/regulatory/?id={m.news_id}',
            })
            if len(seen_news_ids) >= 10:
                break
        # ── 원료 매칭 알림 (NewsIngredientMatch) ──────────────────────────
        ing_matches = (
            NewsIngredientMatch.objects
            .filter(user=request.user, dismissed_yn=False)
            .select_related('news', 'ingredient')
            .order_by('-created_at')[:20]
        )
        seen_ing_news = set()
        for m in ing_matches:
            if not m.news_id or m.news_id in seen_ing_news or m.news_id in seen_news_ids:
                continue
            seen_ing_news.add(m.news_id)
            ingr_name = (getattr(m.ingredient, 'prdlst_nm', '') or '') if m.ingredient else ''
            pname = (m.news.product_name or '') if m.news else ''
            items.append({
                'id':          f'ing_{m.id}',
                'message':     f"원료 '{ingr_name[:15]}' 관련 부적합 알림: '{pname[:20]}'",
                'read_yn':     m.read_yn,
                'status_code': 'REGULATORY',
                'created_at':  m.created_at.strftime('%Y-%m-%d %H:%M'),
                'label_id':    None,
                'label_name':  ingr_name,
                'url':         f'/regulatory/?id={m.news_id}',
            })
    except Exception:
        pass  # 앱 미설치 환경 안전 처리

    # ── 키워드 알림 (PushNotificationLog, trigger_type='keyword') ────────
    try:
        from v1.mobile.models import AppDevice as _AppDev, PushNotificationLog as _PushLog
        _dev_ids = list(_AppDev.objects.filter(user=request.user).values_list('id', flat=True))
        if _dev_ids:
            _kw_logs = (
                _PushLog.objects
                .filter(device_id__in=_dev_ids, trigger_type='keyword')
                .select_related('news')
                .order_by('-created_at')[:200]
            )
            # 뉴스 단위로 그룹핑 (다른 탭과 동일 규칙) → 같은 뉴스가 여러 키워드에
            # 매칭돼도 1건만 표시. 매칭된 키워드는 메시지에 모아서 보여준다.
            _kw_groups = {}
            for _log in _kw_logs:
                _kw_groups.setdefault(_log.news_id, []).append(_log)
            for _news_id, _logs in _kw_groups.items():
                _first = _logs[0]
                _pname = (_first.news.product_name or '') if _first.news else ''
                _unread = sum(1 for l in _logs if not l.is_read)
                unread_count += _unread
                _kw_labels = list(dict.fromkeys(l.trigger_label for l in _logs if l.trigger_label))
                _kw_text = ', '.join(_kw_labels[:2])
                if len(_kw_labels) > 2:
                    _kw_text += f' 외 {len(_kw_labels) - 2}개'
                _msg = f"키워드 '{_kw_text}': '{_pname[:20]}' 이(가) 수집되었습니다."
                items.append({
                    'id':          f'kw_{_first.id}',
                    'message':     _msg,
                    'read_yn':     _unread == 0,
                    'status_code': 'KEYWORD',
                    'created_at':  _first.created_at.strftime('%Y-%m-%d %H:%M'),
                    'label_id':    None,
                    'label_name':  _kw_text,
                    'url':         f'/regulatory/?q={_kw_labels[0] if _kw_labels else ""}',
                })
    except Exception:
        pass

    # ── 수거검사 알림 (InspectionMatch) ─────────────────────────────────
    try:
        from v1.regulatory.models import InspectionMatch as _InspMatch
        _insp_list = (
            _InspMatch.objects
            .filter(user=request.user, notified_at__isnull=False)
            .select_related('inspection')
            .order_by('-notified_at')[:20]
        )
        for _m in _insp_list:
            if not _m.read_yn:
                unread_count += 1
            _pname = (_m.inspection.prdtnm or '') if _m.inspection else ''
            _phase = _m.get_alert_phase_display() if hasattr(_m, 'get_alert_phase_display') else ''
            items.append({
                'id':          f'insp_{_m.id}',
                'message':     f"수거검사 알림: '{_pname[:20]}' {_phase}",
                'read_yn':     _m.read_yn,
                'status_code': 'INSPECTION',
                'created_at':  _m.notified_at.strftime('%Y-%m-%d %H:%M') if _m.notified_at else '',
                'label_id':    None,
                'label_name':  '',
                'url':         '/regulatory/?tab=insp',
            })
    except Exception:
        pass

    # 미읽음 우선 정렬
    items.sort(key=lambda x: (0 if not x['read_yn'] else 1, x['created_at']), reverse=False)
    items = items[:30]

    return JsonResponse({'notifications': items, 'unread': unread_count})


@login_required
@require_POST
def notification_mark_read(request):
    """알림 읽음 처리 - 특정 id 또는 전체 (부적합.처분 알림 포함)"""
    import json
    from django.utils import timezone as tz
    try:
        body = json.loads(request.body)
        notification_id = body.get('id')
    except (ValueError, AttributeError):
        notification_id = None

    nid = str(notification_id) if notification_id else ''

    if nid.startswith('reg_'):
        # 내제품 매칭 알림
        try:
            from v1.regulatory.models import NewsProductMatch
            NewsProductMatch.objects.filter(
                id=int(nid[4:]), product__user_id=request.user
            ).update(read_yn=True, read_at=tz.now())
        except Exception:
            pass
    elif nid.startswith('ing_'):
        # 원료 매칭 알림
        try:
            from v1.regulatory.models import NewsIngredientMatch
            NewsIngredientMatch.objects.filter(
                id=int(nid[4:]), user=request.user
            ).update(read_yn=True)
        except Exception:
            pass
    elif nid.startswith('kw_'):
        # 키워드 알림 — 해당 키워드의 모든 로그를 읽음 처리
        try:
            from v1.mobile.models import AppDevice as _AppDev, PushNotificationLog as _PushLog
            _log = _PushLog.objects.filter(
                id=int(nid[3:]), device__user=request.user
            ).select_related('device').first()
            if _log:
                _dev_ids = list(_AppDev.objects.filter(user=request.user).values_list('id', flat=True))
                _PushLog.objects.filter(
                    device_id__in=_dev_ids,
                    trigger_label=_log.trigger_label,
                    trigger_type='keyword',
                    is_read=False,
                ).update(is_read=True)
        except Exception:
            pass
    elif nid.startswith('insp_'):
        # 수거검사 알림
        try:
            from v1.regulatory.models import InspectionMatch
            InspectionMatch.objects.filter(
                id=int(nid[5:]), user=request.user
            ).update(read_yn=True)
        except Exception:
            pass
    elif nid:
        ProductNotification.objects.filter(
            id=int(nid), recipient=request.user
        ).update(read_yn=True)
    else:
        # 전체 읽음
        ProductNotification.objects.filter(recipient=request.user, read_yn=False).update(read_yn=True)
        try:
            from v1.regulatory.models import NewsProductMatch, NewsIngredientMatch, InspectionMatch
            NewsProductMatch.objects.filter(
                product__user_id=request.user, read_yn=False
            ).update(read_yn=True, read_at=tz.now())
            NewsIngredientMatch.objects.filter(user=request.user, read_yn=False).update(read_yn=True)
            InspectionMatch.objects.filter(user=request.user, read_yn=False, notified_at__isnull=False).update(read_yn=True)
        except Exception:
            pass
        try:
            from v1.mobile.models import AppDevice as _AppDev, PushNotificationLog as _PushLog
            _dev_ids = list(_AppDev.objects.filter(user=request.user).values_list('id', flat=True))
            if _dev_ids:
                _PushLog.objects.filter(device_id__in=_dev_ids, is_read=False).update(is_read=True)
        except Exception:
            pass
        # 규제 알림 사이드바 캐시 무효화
        from django.core.cache import cache
        cache.delete(f'regulatory_alert_count_{request.user.id}')

    unread_count = ProductNotification.objects.filter(recipient=request.user, read_yn=False).count()
    try:
        from v1.regulatory.models import (
            NewsProductMatch, NewsIngredientMatch, InspectionMatch
        )
        _prod_unread = set(NewsProductMatch.objects.filter(product__user_id=request.user, false_positive_yn=False, read_yn=False).values_list('news_id', flat=True))
        _ing_unread  = set(NewsIngredientMatch.objects.filter(user=request.user, dismissed_yn=False, read_yn=False).values_list('news_id', flat=True))
        unread_count += len(_prod_unread | _ing_unread)
        unread_count += InspectionMatch.objects.filter(user=request.user, read_yn=False, notified_at__isnull=False).count()
    except Exception:
        pass
    try:
        from v1.mobile.models import AppDevice as _AppDev, PushNotificationLog as _PushLog
        _dev_ids = list(_AppDev.objects.filter(user=request.user).values_list('id', flat=True))
        if _dev_ids:
            unread_count += _PushLog.objects.filter(device_id__in=_dev_ids, trigger_type='keyword', is_read=False).count()
    except Exception:
        pass

    return JsonResponse({'success': True, 'unread': unread_count})

@login_required
def contacts(request):
    """연락처 관리 - 공유 이력 기반 연락처 목록 및 공유 현황"""
    from v1.label.models import MyLabel

    # 내가 공유한 이메일 목록 (고유값, 최신 이름/회사명 우선)
    # 라벨 소유자로서 공유하거나, EDITOR 권한으로 공유를 생성한 경우 모두 포함
    sent_shares = (
        ProductShare.objects
        .filter(
            Q(label__user_id=request.user) | Q(created_by=request.user)
        )
        .filter(share_mode='PRIVATE', active_yn=True)
        .exclude(recipient_email__isnull=True)
        .exclude(recipient_email='')
        .select_related('permission', 'recipient_user')
        .distinct()
        .order_by('recipient_email', '-created_datetime')
    )

    contact_map = {}
    for s in sent_shares:
        email = s.recipient_email.lower()
        if email not in contact_map:
            name = (s.recipient_name
                    or (s.recipient_user.get_full_name() or s.recipient_user.username
                        if s.recipient_user else '')
                    or '')
            contact_map[email] = {
                'email': email,
                'name': name,
                'company': s.recipient_company or '',
                'license_no': s.recipient_license_no or '',
                'sent': 1,
                'received': 0,
            }
        else:
            contact_map[email]['sent'] += 1

    # UserContact 기반 연락처 추가 (ProductShare에 없는 경우에만)
    for uc in UserContact.objects.filter(owner=request.user):
        email = uc.email.lower()
        if email not in contact_map:
            contact_map[email] = {
                'email': email,
                'name': uc.name or '',
                'company': uc.company or '',
                'license_no': uc.license_no or '',
                'sent': 0,
                'received': 0,
            }

    contacts_list = sorted(contact_map.values(), key=lambda x: x['email'])

    # 문서 요청 집계 (내가 보낸 요청 기준)
    from v1.products.models import DocumentRequest
    from django.db.models import Count, Q as Qdr
    dr_agg = (
        DocumentRequest.objects
        .filter(requester=request.user)
        .values('recipient_email')
        .annotate(
            total=Count('request_id'),
            pending=Count('request_id', filter=Qdr(status='PENDING')),
        )
    )
    dr_map = {row['recipient_email'].lower(): row for row in dr_agg}

    # 내가 받은 요청 집계 (요청수신 필터용)
    dr_recv_agg = (
        DocumentRequest.objects
        .filter(recipient_email__iexact=request.user.email)
        .values('requester__email')
        .annotate(total=Count('request_id'))
    )
    dr_recv_map = {row['requester__email'].lower(): row['total'] for row in dr_recv_agg}

    for c in contacts_list:
        dr = dr_map.get(c['email'], {})
        c['doc_sent']     = dr.get('total', 0)
        c['doc_pending']  = dr.get('pending', 0)
        c['doc_received'] = dr_recv_map.get(c['email'], 0)  # 이 연락처로부터 받은 요청

    context = {
        'contacts_list': contacts_list,
    }
    return render(request, 'products/contacts.html', context)


@login_required
def contacts_api_list(request):
    """연락처 목록 JSON API"""
    # ① ProductShare 기반 연락처
    sent_shares = (
        ProductShare.objects
        .filter(
            Q(label__user_id=request.user) | Q(created_by=request.user)
        )
        .filter(share_mode='PRIVATE', active_yn=True)
        .exclude(recipient_email__isnull=True)
        .exclude(recipient_email='')
        .select_related('recipient_user')
        .distinct()
    )

    contact_map = {}
    for s in sent_shares:
        email = (s.recipient_email or '').lower()
        if email and email not in contact_map:
            name = (s.recipient_name
                    or (s.recipient_user.get_full_name() or s.recipient_user.username
                        if s.recipient_user else '')
                    or '')
            contact_map[email] = {
                'email': email,
                'name': name,
                'company': s.recipient_company or '',
                'license_no': s.recipient_license_no or '',
            }

    # ② UserContact 기반 연락처 (ProductShare에 없는 경우에만 추가)
    for uc in UserContact.objects.filter(owner=request.user):
        email = uc.email.lower()
        if email not in contact_map:
            contact_map[email] = {
                'email': email,
                'name': uc.name or '',
                'company': uc.company or '',
                'license_no': uc.license_no or '',
            }

    data = sorted(contact_map.values(), key=lambda x: x['email'])
    return JsonResponse({'contacts': data})


@login_required
@require_POST
def contacts_api_update(request):
    """연락처 정보 업데이트 API - 이메일 수정 지원"""
    old_email = request.POST.get('old_email', '').strip().lower()  # 기존 이메일
    new_email = request.POST.get('email', '').strip().lower()      # 새 이메일
    name = request.POST.get('name', '').strip()
    company = request.POST.get('company', '').strip()
    license_no = request.POST.get('license_no', '').strip()

    if not old_email or not new_email:
        return JsonResponse({'success': False, 'error': '이메일이 필요합니다.'}, status=400)

    # 새 이메일 유효성 검증 (Django 표준)
    from django.core.validators import validate_email as _validate_email
    from django.core.exceptions import ValidationError as _DjValidationError
    try:
        _validate_email(new_email)
    except _DjValidationError:
        return JsonResponse({'success': False, 'error': '올바른 이메일 형식이 아닙니다.'}, status=400)

    # ① ProductShare 레코드 일괄 업데이트 (이메일 변경 포함)
    updated = (
        ProductShare.objects
        .filter(
            Q(label__user_id=request.user) | Q(created_by=request.user)
        )
        .filter(recipient_email__iexact=old_email, share_mode='PRIVATE', active_yn=True)
        .distinct()
        .update(
            recipient_email=new_email if old_email != new_email else old_email,  # 이메일 변경
            recipient_name=name or None,
            recipient_company=company or None,
            recipient_license_no=license_no or None,
        )
    )

    # ② UserContact 동기화 (있으면 업데이트)
    if old_email != new_email:
        # 이메일이 변경되면 기존 레코드 삭제 및 새로 생성
        UserContact.objects.filter(owner=request.user, email__iexact=old_email).delete()
        UserContact.objects.create(
            owner=request.user,
            email=new_email,
            name=name or None,
            company=company or None,
            license_no=license_no or None,
        )
    else:
        # 이메일은 같고 다른 정보만 수정
        UserContact.objects.filter(owner=request.user, email__iexact=old_email).update(
            name=name or None,
            company=company or None,
            license_no=license_no or None,
        )

    return JsonResponse({'success': True, 'updated': updated})


@login_required
@require_POST
def contacts_api_add(request):
    """새 연락처 추가 API – UserContact 테이블에 저장 (이미 있으면 업데이트)"""
    email = request.POST.get('email', '').strip().lower()
    name = request.POST.get('name', '').strip()
    company = request.POST.get('company', '').strip()
    license_no = request.POST.get('license_no', '').strip()

    if not email:
        return JsonResponse({'success': False, 'error': '이메일이 필요합니다.'}, status=400)

    # 이메일 유효성 검증 (Django 표준)
    from django.core.validators import validate_email as _validate_email
    from django.core.exceptions import ValidationError as _DjValidationError
    try:
        _validate_email(email)
    except _DjValidationError:
        return JsonResponse({'success': False, 'error': '올바른 이메일 형식이 아닙니다.'}, status=400)

    contact, created = UserContact.objects.update_or_create(
        owner=request.user,
        email=email,
        defaults={
            'name': name or None,
            'company': company or None,
            'license_no': license_no or None,
        },
    )
    return JsonResponse({'success': True, 'created': created})


@login_required
def contacts_api_shares(request):
    """특정 이메일의 공유 현황 JSON API"""
    email = request.GET.get('email', '').strip().lower()
    if not email:
        return JsonResponse({'sent': [], 'received': []})

    # 내가 그 이메일에 공유한 항목 (라벨 소유자 또는 공유 생성자 기준)
    sent = []
    for s in (
        ProductShare.objects
        .filter(
            Q(label__user_id=request.user) | Q(created_by=request.user)
        )
        .filter(recipient_email__iexact=email, share_mode='PRIVATE', active_yn=True)
        .select_related('label', 'permission')
        .distinct()
        .order_by('-created_datetime')
    ):
        try:
            role_code    = s.permission.role_code
            role_display = s.permission.get_role_code_display()
        except Exception:
            role_code, role_display = '', '미지정'
        sent.append({
            'share_id': s.share_id,
            'label_id': s.label.my_label_id,
            'label_name': s.label.my_label_name or s.label.prdlst_nm or '(이름 없음)',
            'role': role_code,
            'role_display': role_display,
            'share_end_date': s.share_end_date.strftime('%Y-%m-%d') if s.share_end_date else '',
            'created_at': s.created_datetime.strftime('%Y-%m-%d') if s.created_datetime else '',
        })

    # 그 이메일(사람)이 나에게 공유한 항목
    received = []
    from django.contrib.auth.models import User as DjangoUser
    sharer_users = DjangoUser.objects.filter(email__iexact=email)
    if sharer_users.exists():
        sharer = sharer_users.first()
        for s in (
            ProductShare.objects
            .filter(label__user_id=sharer, share_mode='PRIVATE', active_yn=True)
            .filter(Q(recipient_user=request.user) | Q(recipient_email__iexact=request.user.email))
            .select_related('label', 'permission')
            .order_by('-created_datetime')
        ):
            try:
                role_code    = s.permission.role_code
                role_display = s.permission.get_role_code_display()
            except Exception:
                role_code, role_display = '', '미지정'
            received.append({
                'share_id': s.share_id,
                'label_id': s.label.my_label_id,
                'label_name': s.label.my_label_name or s.label.prdlst_nm or '(이름 없음)',
                'role': role_code,
                'role_display': role_display,
                'share_end_date': s.share_end_date.strftime('%Y-%m-%d') if s.share_end_date else '',
                'created_at': s.created_datetime.strftime('%Y-%m-%d') if s.created_datetime else '',
            })

    return JsonResponse({'sent': sent, 'received': received})


@login_required
def contacts_api_doc_requests(request):
    """연락처별 자료 요청 이력 JSON API"""
    from v1.products.models import DocumentRequest, DocumentSubmission, ProductDocument
    email = request.GET.get('email', '').strip().lower()
    if not email:
        return JsonResponse({'requests': []})

    try:
        dr_list = (
            DocumentRequest.objects
            .filter(requester=request.user, recipient_email__iexact=email)
            .select_related('linked_label')
            .prefetch_related('submissions')
            .order_by('-created_datetime')
        )
        data = []
        for dr in dr_list:
            subs = []
            for s in dr.submissions.filter(active_yn=True):
                # ProductDocument 찾기 (vendor_submission_id로)
                try:
                    pd = ProductDocument.objects.filter(
                        metadata__vendor_submission_id=s.submission_id
                    ).first()
                except Exception as e:
                    logger.warning(f"ProductDocument lookup error for submission {s.submission_id}: {e}")
                    pd = None
                
                subs.append({
                    'submission_id':   str(s.submission_id),
                    'document_type':   s.document_type or '',
                    'filename':        s.original_filename or '',
                    'file_url':        request.build_absolute_uri(s.file.url) if s.file else '',
                    'file_size':       int(s.file_size) if s.file_size else 0,
                    'submitted':       s.submitted_datetime.strftime('%Y-%m-%d %H:%M') if s.submitted_datetime else '',
                    'submitted_by':    s.submitted_by_name or s.submitted_by_email or '',
                    'document_id':     str(pd.document_id) if pd and pd.document_id else None,
                    'ai_status':       pd.metadata.get('ai_status') if pd and pd.metadata else None,
                })
            data.append({
                'request_id':          str(dr.request_id),
                'status':              dr.status or '',
                'status_display':      dr.get_status_display() or '',
                'requested_documents': [{'type_name': d.get('type_name', '')} for d in (dr.requested_documents or [])],
                'message':             dr.message or '',
                'due_date':            dr.due_date.strftime('%Y-%m-%d') if dr.due_date else '',
                'created':             dr.created_datetime.strftime('%Y-%m-%d %H:%M') if dr.created_datetime else '',
                'submissions':         subs,
                'linked_label_id':     str(dr.linked_label_id) if dr.linked_label_id else None,
                'linked_label_name':   dr.linked_label.my_label_name if dr.linked_label else None,
            })
        return JsonResponse({'requests': data})
    except Exception as e:
        logger.exception(f"contacts_api_doc_requests error for email={email}, user={request.user}")
        return JsonResponse({'error': str(e), 'requests': []}, status=400)


@login_required
@require_POST
def doc_request_submit(request, req_id):
    """수신자가 문서 파일을 제출하고 요청을 수낙"""
    from v1.products.models import DocumentRequest, DocumentSubmission
    try:
        dr = DocumentRequest.objects.get(
            request_id=req_id,
            recipient_email__iexact=request.user.email,
        )
    except DocumentRequest.DoesNotExist:
        return JsonResponse({'error': '요청을 찾을 수 없습니다.'}, status=404)
    if dr.status == DocumentRequest.STATUS_CANCELLED:
        return JsonResponse({'error': '취소된 요청입니다.'}, status=400)

    files   = request.FILES  # key = doc_type_name (또는 type_id)
    notes   = request.POST.get('notes', '')
    saved   = []

    for key, f in files.items():
        sub = DocumentSubmission(
            request            = dr,
            document_type      = key,
            file               = f,
            original_filename  = f.name,
            file_size          = f.size,
            submitted_by_email = request.user.email,
            submitted_by_name  = request.user.get_full_name() or request.user.username,
            notes              = notes,
            active_yn          = True,
        )
        sub.save()
        saved.append({'document_type': key, 'filename': f.name})

    # 파일이 하나라도 제출되면 수낙 상태로 변경
    if saved:
        dr.status = DocumentRequest.STATUS_ACCEPTED
        dr.save(update_fields=['status', 'updated_datetime'])

        # 요청자의 제품(linked_label)에 제출된 파일을 ProductDocument로 자동 등록
        if dr.linked_label:
            from django.core.files.base import File as DjangoFile
            for sub in DocumentSubmission.objects.filter(request=dr, active_yn=True):
                # document_type 매핑: DocumentSubmission.document_type (문자열) → DocumentType
                dtype = DocumentType.objects.filter(type_name__iexact=sub.document_type, active_yn=True).first()
                if dtype and sub.file:
                    try:
                        sub.file.seek(0)
                        ProductDocument.objects.create(
                            label=dr.linked_label,
                            document_type=dtype,
                            file=DjangoFile(sub.file, name=sub.original_filename),
                            original_filename=sub.original_filename,
                            file_size=sub.file_size or 0,
                            uploaded_by=dr.requester,
                        )
                    except Exception:
                        pass  # 파일 복사 실패 시 제출 자체는 유지

    return JsonResponse({'success': True, 'submitted': saved})


@login_required
def contacts_api_received_doc_requests(request):
    """내가 받은 자료 요청 목록 JSON API"""
    from v1.products.models import DocumentRequest, DocumentSubmission
    dr_list = (
        DocumentRequest.objects
        .filter(recipient_email__iexact=request.user.email)
        .select_related('requester')
        .prefetch_related('submissions')
        .order_by('-created_datetime')
    )
    data = []
    for dr in dr_list:
        subs = [
            {
                'document_type': s.document_type,
                'filename':      s.original_filename,
                'file_url':      s.file.url if s.file else '',
                'submitted':     s.submitted_datetime.strftime('%Y-%m-%d %H:%M') if s.submitted_datetime else '',
            }
            for s in dr.submissions.filter(active_yn=True)
        ]
        data.append({
            'request_id':          dr.request_id,
            'status':              dr.status,
            'status_display':      dr.get_status_display(),
            'requester_name':      dr.requester.get_full_name() or dr.requester.username,
            'requester_email':     dr.requester.email,
            'requested_documents': dr.requested_documents or [],
            'submissions':         subs,
            'message':             dr.message or '',
            'due_date':            dr.due_date.strftime('%Y-%m-%d') if dr.due_date else '',
            'created':             dr.created_datetime.strftime('%Y-%m-%d %H:%M') if dr.created_datetime else '',
            'attachment_url':      dr.attachment.url if dr.attachment else '',
        })
    return JsonResponse({'requests': data})


@login_required
def doc_requests_dashboard(request):
    """연락처 관리로 통합 — 자료 요청 서브메뉴는 연락처 관리 페이지에서 확인합니다."""
    from django.shortcuts import redirect
    return redirect('products:contacts')


@login_required
@require_POST
def doc_request_cancel(request, req_id):
    """내가 보낸 문서 요청 취소"""
    from v1.products.models import DocumentRequest
    try:
        dr = DocumentRequest.objects.get(request_id=req_id, requester=request.user)
    except DocumentRequest.DoesNotExist:
        return JsonResponse({'error': '요청을 찾을 수 없습니다.'}, status=404)
    if dr.status != DocumentRequest.STATUS_PENDING:
        return JsonResponse({'error': '이미 처리된 요청입니다.'}, status=400)
    dr.status = DocumentRequest.STATUS_CANCELLED
    dr.save(update_fields=['status', 'updated_datetime'])
    return JsonResponse({'success': True})


@login_required
@require_POST
def doc_request_accept(request, req_id):
    """내가 받은 문서 요청 수락"""
    from v1.products.models import DocumentRequest
    try:
        dr = DocumentRequest.objects.get(
            request_id=req_id,
            recipient_email__iexact=request.user.email,
        )
    except DocumentRequest.DoesNotExist:
        return JsonResponse({'error': '요청을 찾을 수 없습니다.'}, status=404)
    if dr.status != DocumentRequest.STATUS_PENDING:
        return JsonResponse({'error': '이미 처리된 요청입니다.'}, status=400)
    dr.status = DocumentRequest.STATUS_ACCEPTED
    dr.save(update_fields=['status', 'updated_datetime'])
    log_activity(request, 'document', 'doc_request_accept', req_id)
    return JsonResponse({'success': True})


@login_required
def api_doc_types(request):
    """활성화된 문서 유형 목록 JSON"""
    from v1.products.models import DocumentType
    types = DocumentType.objects.filter(active_yn=True).values(
        'type_id', 'type_code', 'type_name', 'icon', 'color', 'required_yn', 'description'
    ).order_by('display_order', 'type_name')
    return JsonResponse({'doc_types': list(types)})


@login_required
def api_my_labels(request):
    """내 제품(MyLabel) 목록 JSON — 자료 요청 연결용"""
    from v1.label.models import MyLabel
    labels = (
        MyLabel.objects.filter(user_id=request.user, delete_YN='N')
        .order_by('-my_label_id')
        .values('my_label_id', 'my_label_name', 'prdlst_nm', 'prdlst_dcnm')
    )
    data = [
        {
            'id': l['my_label_id'],
            'name': l['my_label_name'] or l['prdlst_nm'] or f"제품 #{l['my_label_id']}",
            'prdlst_dcnm': l['prdlst_dcnm'] or '',
        }
        for l in labels
    ]
    return JsonResponse({'labels': data})


@login_required
def api_my_company_documents(request):
    """내 고정 서류 목록 JSON — 제품 문서 불러오기용"""
    from v1.user_management.models import CompanyDocument
    docs = CompanyDocument.objects.filter(user=request.user).select_related('linked_document_type').order_by('doc_type', '-uploaded_at')
    data = [
        {
            'id': d.pk,
            'doc_type': d.doc_type,
            'doc_type_display': (d.linked_document_type.type_name if d.linked_document_type else d.get_doc_type_display()),
            'doc_name': d.doc_name,
            'file_url': d.doc_file.url if d.doc_file else '',
            'uploaded_at': d.uploaded_at.strftime('%Y-%m-%d') if d.uploaded_at else '',
            'note': d.note,
            'linked_doc_type_id': d.linked_document_type_id,
            'linked_doc_type_name': d.linked_document_type.type_name if d.linked_document_type else None,
        }
        for d in docs
    ]
    return JsonResponse({'documents': data})


@login_required
@require_POST
def api_update_doc_request_label(request, req_id):
    """이미 보낸 자료 요청에 제품 연결"""
    from v1.products.models import DocumentRequest
    from v1.label.models import MyLabel as _MyLabel
    try:
        dr = DocumentRequest.objects.get(request_id=req_id, requester=request.user)
    except DocumentRequest.DoesNotExist:
        return JsonResponse({'error': '요청을 찾을 수 없습니다.'}, status=404)

    label_id = request.POST.get('linked_label_id') or None
    if label_id:
        label = _MyLabel.objects.filter(
            my_label_id=label_id, user_id=request.user, delete_YN='N'
        ).first()
        if not label:
            return JsonResponse({'error': '해당 제품을 찾을 수 없습니다.'}, status=404)
        dr.linked_label = label
    else:
        dr.linked_label = None
    dr.save(update_fields=['linked_label', 'updated_datetime'])

    # 새로 제품이 연결된 경우, 이미 제출된 문서를 ProductDocument로 소급 등록
    imported_count = 0
    if label:
        from .models import DocumentSubmission, DocumentType, ProductDocument
        from django.core.files.base import File as DjangoFile

        for sub in DocumentSubmission.objects.filter(request=dr, active_yn=True):
            if not sub.file:
                continue
            # 동일 파일이 이미 등록돼 있으면 skip (중복 방지)
            already = ProductDocument.objects.filter(
                label=label,
                original_filename=sub.original_filename,
                file_size=sub.file_size,
            ).exists()
            if already:
                continue

            # requested_documents JSON에서 type_id로 먼저 조회 (더 정확)
            dtype = None
            for rd in (dr.requested_documents or []):
                if rd.get('type_name') == sub.document_type and rd.get('type_id'):
                    dtype = DocumentType.objects.filter(
                        type_id=rd['type_id'], active_yn=True
                    ).first()
                    if dtype:
                        break

            # fallback: type_name 문자열로 조회
            if not dtype:
                dtype = DocumentType.objects.filter(
                    type_name__iexact=sub.document_type, active_yn=True
                ).first()

            if dtype:
                try:
                    sub.file.seek(0)
                    ProductDocument.objects.create(
                        label=label,
                        document_type=dtype,
                        file=DjangoFile(sub.file, name=sub.original_filename),
                        original_filename=sub.original_filename,
                        file_size=sub.file_size or 0,
                        uploaded_by=dr.requester,
                    )
                    imported_count += 1
                except Exception:
                    pass

    return JsonResponse({
        'success': True,
        'linked_label_id': dr.linked_label_id,
        'linked_label_name': dr.linked_label.my_label_name if dr.linked_label else None,
        'imported_count': imported_count,
    })


@login_required
@require_POST
def api_send_doc_request(request):
    """문서 요청 생성 + 이메일 발송 (multipart/form-data)"""
    import json
    from v1.products.models import DocumentRequest, DocumentType
    from django.core.mail import EmailMessage
    from django.conf import settings

    # FormData로 받음 (파일 첨부 지원)
    try:
        recipients   = json.loads(request.POST.get('recipients', '[]'))
        type_ids     = json.loads(request.POST.get('type_ids', '[]'))
    except (ValueError, TypeError):
        return JsonResponse({'error': '잘못된 요청 형식입니다.'}, status=400)

    message_text     = request.POST.get('message', '').strip()
    attachment       = request.FILES.get('attachment')
    linked_label_id  = request.POST.get('linked_label_id')
    linked_ingredient_id = request.POST.get('linked_ingredient_id')

    # 연결 제품/원료 검증 (소유자 확인)
    linked_label = None
    linked_ingredient = None
    if linked_label_id:
        linked_label = MyLabel.objects.filter(my_label_id=linked_label_id, user_id=request.user, delete_YN='N').first()
    if linked_ingredient_id:
        from v1.label.models import MyIngredient as _MyIngredient
        linked_ingredient = _MyIngredient.objects.filter(my_ingredient_id=linked_ingredient_id, user_id=request.user, delete_YN='N').first()

    if not recipients:
        return JsonResponse({'error': '수신자를 선택해주세요.'}, status=400)
    if not type_ids:
        return JsonResponse({'error': '요청할 문서 종류를 선택해주세요.'}, status=400)

    doc_types = list(DocumentType.objects.filter(type_id__in=type_ids, active_yn=True))
    if not doc_types:
        return JsonResponse({'error': '유효한 문서 종류가 없습니다.'}, status=400)
    doc_names = ', '.join(dt.type_name for dt in doc_types)

    requester_name, requester_company = _get_sender_info(request.user)

    created_count = 0
    email_errors  = []
    for r in recipients:
        email = (r.get('email') or '').strip().lower()
        if not email:
            continue
        doc_info = [{'type_id': dt.type_id, 'type_name': dt.type_name} for dt in doc_types]

        dr = DocumentRequest.objects.create(
            requester            = request.user,
            recipient_email      = email,
            recipient_name       = r.get('name', ''),
            recipient_company    = r.get('company', ''),
            requested_documents  = doc_info,
            message              = message_text,
            attachment           = attachment,
            linked_label         = linked_label,
            linked_ingredient    = linked_ingredient,
        )
        created_count += 1

        # ── 이메일 발송 ──
        from django.conf import settings as _ds
        # 동적 사이트 URL: 현재 요청 도메인 사용
        # request.build_absolute_uri('/')는 현재 스킴과 호스트를 기반으로 생성
        # 예: localhost:8000 또는 ezlabeling.com
        base_url = request.build_absolute_uri('/').rstrip('/')
        
        subject = f'[EzLabeling] {requester_name}님이 문서 제출을 요청했습니다'
        _txt, _html = _render_email('emails/doc_request.html', {
            'subject': subject,
            'sender_name': requester_name, 'sender_company': requester_company, 'sender_email': request.user.email,
            'recipient_name': r.get('name', ''), 'recipient_company': r.get('company', ''),
            'doc_list': [dt.type_name for dt in doc_types],
            'message_text': message_text,
            'doc_request_url': f'{base_url}/vendor/upload/{dr.upload_token}/',  # ← 동적 URL 사용
        })

        # 첨부파일 처리
        attach_tuple = None
        if attachment and dr.attachment:
            try:
                attachment.seek(0)
                attach_tuple = (attachment.name, attachment.read(), attachment.content_type)
            except Exception:
                attach_tuple = None

        sent_ok = _send_email_safe(subject=subject, body=_txt, to_email=email, html_body=_html, attachment=attach_tuple)
        if sent_ok:
            dr.email_sent = True
            dr.email_sent_datetime = timezone.now()
            dr.save(update_fields=['email_sent', 'email_sent_datetime'])
        else:
            email_errors.append(email)

    if created_count == 0:
        return JsonResponse({'error': '요청 대상 이메일이 없습니다.'}, status=400)

    log_activity(request, 'document', 'doc_request_send')
    return JsonResponse({'success': True, 'created': created_count, 'email_errors': email_errors})


 
# ==================== AI 문서 리뷰 (Human-in-the-Loop) ====================
 
@login_required
def document_ai_review(request, document_id):
    """AI 문서 리뷰 페이지"""
    from v1.products.services.vision_service import _pdf_to_base64_images, _image_to_base64
    
    doc = get_object_or_404(
        ProductDocument.objects.select_related('label', 'document_type', 'label__user_id'),
        pk=document_id,
        label__user_id=request.user,
    )
    
    meta = doc.metadata or {}
    ai_group = meta.get('ai_group', 'A')
    ai_status = meta.get('ai_status', 'PENDING')
    extracted = meta.get('extracted_data') or {}
    test_items = extracted.get('test_items') or []
    
    # PDF/이미지 변환
    document_images = []
    if doc.file:
        try:
            file_path = doc.file.path
            ext = os.path.splitext(file_path)[1].lower()
            if ext == '.pdf':
                document_images = _pdf_to_base64_images(file_path, max_pages=2)
            elif ext in ['.jpg', '.jpeg', '.png', '.gif']:
                img_b64 = _image_to_base64(file_path)
                if img_b64:
                    document_images = [img_b64]
        except Exception as e:
            logger.warning(f"Image convert error: {e}")
    
    # 필드 설정
    fields_config = {
        'product_name': {'label': '제품명', 'desc': '제품 명칭', 'ph': '예: PINE SOFT-T', 'req': True},
        'food_type': {'label': '제품의유형', 'desc': '식품 분류', 'ph': '예: 음료수', 'req': True},
        'manufacturer': {'label': '제조업소명', 'desc': '제조사', 'ph': '예: 알글리딘(주)', 'req': True},
        'raw_materials': {'label': '원재료명', 'desc': '원재료 목록', 'ph': '예: 생강, 민트', 'req': True},
        'blend_ratios': {'label': '배합비', 'desc': '배합 비율', 'ph': '예: 37%, 17%', 'req': False},
        'origins': {'label': '원산지', 'desc': '원산지 정보', 'ph': '예: 미국', 'req': False},
        'allergens': {'label': '알레르기 함유', 'desc': '알레르기 정보', 'ph': '예: 없음', 'req': False},
        'storage_method': {'label': '보관방법', 'desc': '보관 조건', 'ph': '예: 실온보관', 'req': False},
        'shelf_life': {'label': '유통기한', 'desc': '유통기한', 'ph': '예: 2년', 'req': False},
    }
    
    extracted_with_labels = [
        {'key': k, 'label': v['label'], 'value': extracted.get(k, '')}
        for k, v in fields_config.items()
    ]
    
    return render(request, 'products/document_ai_review_v2.html', {
        'doc': doc,
        'ai_group': ai_group,
        'ai_status': ai_status,
        'extracted': extracted,
        'extracted_with_labels': extracted_with_labels,
        'document_images': document_images,
        'compliance_status': meta.get('compliance_status', ''),
        'test_items': test_items,
    })
 
 
@login_required
@require_POST
def document_ai_create_from_submission(request):
    """ProductDocument 생성 API"""
    import json as _json
    from v1.products.models import DocumentSubmission, DocumentType
    
    try:
        payload = _json.loads(request.body)
        submission_id = payload.get('submission_id')
        if not submission_id:
            return JsonResponse({'success': False, 'error': 'submission_id required'}, status=400)
        
        sub = DocumentSubmission.objects.select_related('document_request__linked_label').get(
            submission_id=submission_id, active_yn=True
        )
        label = sub.document_request.linked_label if sub.document_request else None
        
        if not label or (hasattr(label.user_id, 'id') and label.user_id.id != request.user.id) or \
           (not hasattr(label.user_id, 'id') and label.user_id != request.user.id):
            return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
        
        existing = ProductDocument.objects.filter(metadata__vendor_submission_id=submission_id).first()
        if existing:
            return JsonResponse({
                'success': True,
                'document_id': existing.document_id,
                'redirect_url': f'/products/documents/{existing.document_id}/ai-review/',
            })
        
        doc = ProductDocument()
        doc.label = label
        doc_type = DocumentType.objects.filter(active_yn=True).first()
        if doc_type:
            doc.document_type = doc_type
        if sub.file:
            doc.file = sub.file
        
        doc.metadata = {
            'vendor_submission_id': str(submission_id),
            'ai_status': 'PENDING',
            'document_title': sub.original_filename or 'Document',
            'ai_group': 'A',
            'extracted_data': {},
            'compliance_status': 'UNKNOWN',
            'test_items': [],
        }
        doc.save()
        
        return JsonResponse({
            'success': True,
            'document_id': doc.document_id,
            'redirect_url': f'/products/documents/{doc.document_id}/ai-review/',
        })
    
    except DocumentSubmission.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Not found'}, status=404)
    except Exception as e:
        logger.exception(f"Error: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
@require_POST
def document_ai_extract_api(request, document_id):
    """
    AI 문서 분석 실행 API (비동기)
    사용자가 "AI 분석" 버튼 클릭 시 호출되어 실제 AI 추출을 수행합니다.
    
    Response:
        {
            "success": true,
            "status": "COMPLETED",
            "extracted_data": {...},
            "message": "분석 완료"
        }
    """
    import json as _json
    from v1.products.services.vision_service import VisionAIService
    
    doc = get_object_or_404(
        ProductDocument.objects.select_related('label', 'document_type'),
        pk=document_id,
        label__user_id=request.user,
    )
    
    try:
        # PDF/이미지 경로 확인
        if not doc.file or not os.path.exists(doc.file.path):
            return JsonResponse({
                'success': False,
                'error': '문서 파일을 찾을 수 없습니다.',
            }, status=400)
        
        file_path = str(doc.file.path)
        file_ext = os.path.splitext(file_path)[1].lower()
        
        # VisionAIService 초기화 및 실행
        service = VisionAIService()
        
        # 문서 타입에 따른 처리
        if file_ext == '.pdf':
            # PDF → 처음 2페이지 이미지로 변환 후 분석
            from v1.products.services.vision_service import _pdf_to_base64_images
            images_b64 = _pdf_to_base64_images(file_path, max_pages=2)
            if not images_b64:
                raise ValueError("PDF에서 이미지를 추출할 수 없습니다.")
        elif file_ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp']:
            # 이미지 파일 → Base64
            from v1.products.services.vision_service import _image_to_base64
            img_b64 = _image_to_base64(file_path)
            if not img_b64:
                raise ValueError("이미지를 처리할 수 없습니다.")
            images_b64 = [img_b64]
        else:
            return JsonResponse({
                'success': False,
                'error': f'지원하지 않는 파일 형식: {file_ext}',
            }, status=400)
        
        # ── Group A 실행 (데이터 추출) ──
        extracted_data = {}
        try:
            extracted_data = service.extract_group_a(images_b64)
            logger.info(f"AI Group A 추출 성공: document_id={document_id}, extracted={extracted_data}")
        except Exception as e:
            logger.error(f"AI Group A 추출 실패: document_id={document_id}, error={e}")
            extracted_data = {}
        
        # ── Group B 실행 (규정 검증) ──
        compliance_data = {}
        if extracted_data:  # 추출이 성공한 경우만 규정 검증
            try:
                compliance_data = service.extract_group_b(images_b64, extracted_data)
                logger.info(f"AI Group B 검증 성공: document_id={document_id}, compliance={compliance_data}")
            except Exception as e:
                logger.error(f"AI Group B 검증 실패: document_id={document_id}, error={e}")
                compliance_data = {}
        
        # ── 결과 저장 ──
        metadata = doc.metadata or {}
        metadata['ai_status'] = 'COMPLETED'
        metadata['extracted_data'] = extracted_data
        metadata['compliance_status'] = compliance_data.get('overall_status', 'UNKNOWN')
        metadata['test_items'] = compliance_data.get('test_items', [])
        
        doc.metadata = metadata
        doc.save(update_fields=['metadata'])
        
        # ── 응답 ──
        return JsonResponse({
            'success': True,
            'status': 'COMPLETED',
            'extracted_data': extracted_data,
            'test_items': compliance_data.get('test_items', []),
            'message': '분석 완료',
        })
    
    except Exception as e:
        logger.exception(f"document_ai_extract_api error: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e),
            'message': 'AI 분석 중 오류가 발생했습니다.',
        }, status=500)


@login_required
@require_POST
def document_ai_review_save(request, document_id):
    """AI 추출 데이터 수동 수정 저장."""
    import json as _json
    doc = get_object_or_404(
        ProductDocument.objects.select_related('label'),
        pk=document_id,
        label__user_id=request.user,
    )
 
    try:
        payload = _json.loads(request.body)
    except (_json.JSONDecodeError, Exception):
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)
 
    meta = dict(doc.metadata) if doc.metadata else {}
    meta['extracted_data'] = payload.get('extracted_data', meta.get('extracted_data', {}))
    doc.metadata = meta
    doc.save(update_fields=['metadata'])
    return JsonResponse({'success': True})
 
 
def register_ingredient_bom(user, label, fields):
    """
    원료 한 건을 BOM 에 넣는다. 사진에서 왔든 품목보고번호에서 왔든 규칙은 같다.

    같은 원료를 두 번 만들지 않는다 - 이름이 비슷하면(RapidFuzz) 기존 "내 원료"
    에 붙이고, 없을 때만 만든다. 같은 원료가 이 제품 BOM 에 이미 있으면 행을
    늘리지 않고 값을 갱신한다.

    배합비는 넣지 않는다. 그 원료가 완제품에서 몇 %인지는 원료 봉지에도
    품목보고 정보에도 없다. BOM 화면에서 사람이 넣는다.

    Returns: (bom, 만들었는지, 기존 원료에 붙었는지, 유사도, 후보 목록)
    """
    from v1.bom.models import ProductBOM
    from v1.label.services.ingredient_matching import (
        get_or_create_my_ingredient, load_pool, match_my_ingredient,
    )

    name = (fields.get('ingredient_name') or '').strip()
    pool = load_pool(user)
    ingredient, score, candidates = match_my_ingredient(user, name, pool=pool)
    matched = ingredient is not None

    if not matched:
        ingredient, _created = get_or_create_my_ingredient(
            user,
            prdlst_nm=name,
            prdlst_report_no=fields.get('report_no') or '',
            prdlst_dcnm=fields.get('food_type') or '',
            ingredient_display_name=name,
            allergens=fields.get('allergens') or '',
            bssh_nm=fields.get('manufacturer') or '',
            rawmtrl_nm=fields.get('sub_ingredients') or '',
            delete_YN='N',
        )

    # 원재료 표시명에는 **읽어낸 원재료명과 함량**을 넣는다. 이 원료가 완제품에
    # 쓰이면 표시 문구가 "표고버섯볶음(새송이버섯 57.64%, ...)" 로 나가야 한다.
    # 원료명을 그대로 복사하면 BOM 표의 앞 두 칸이 똑같아 "원재료명을 못 읽었다"
    # 로 보이고, 정작 읽은 값은 표에 컬럼이 없는 sub_ingredients 에만 남는다.
    printed = (fields.get('sub_ingredients') or '').strip()

    bom = ProductBOM.objects.filter(
        parent_label=label, source_ingredient=ingredient, active_yn=True).first()
    created = False
    if bom is None:
        bom = ProductBOM.objects.create(
            parent_label=label,
            created_by=user,
            ingredient_name=name,
            raw_material_name=printed or name,
            food_type=fields.get('food_type') or '',
            sub_ingredients=printed,
            allergens=fields.get('allergens') or '',
            allergen=fields.get('allergens') or '',
            origin=fields.get('origin') or '',
            manufacturer=fields.get('manufacturer') or '',
            report_no=fields.get('report_no') or '',
            source_ingredient=ingredient,
            sort_order=ProductBOM.objects.filter(
                parent_label=label, active_yn=True).count(),
            active_yn=True,
        )
        created = True
    else:
        # 다시 읽은 경우. 행을 새로 만들지 않고 값을 갱신한다.
        bom.ingredient_name = name
        bom.raw_material_name = printed or name
        bom.food_type = fields.get('food_type') or bom.food_type
        bom.sub_ingredients = printed or bom.sub_ingredients
        bom.allergens = fields.get('allergens') or bom.allergens
        bom.allergen = fields.get('allergens') or bom.allergen
        bom.origin = fields.get('origin') or bom.origin
        bom.manufacturer = fields.get('manufacturer') or bom.manufacturer
        bom.report_no = fields.get('report_no') or bom.report_no
        bom.save()

    return bom, created, matched, score, candidates


@login_required
@require_POST
def ingredient_to_bom(request, label_id):
    """
    사진 없이 원료 한 건을 BOM 에 넣는다 (품목보고번호로 불러온 경우).

    첨부 파일이 없으므로 문서함에는 아무것도 남기지 않는다. 사진으로 올린
    경우와 달리 근거 자료가 파일이 아니라 품목보고번호 자체다 - 그 번호를
    BOM 행에 적어 둔다.
    """
    from django.db import transaction

    label = _resolve_editable_label(request, label_id)

    try:
        payload = json.loads(request.body or '{}')
    except (ValueError, TypeError):
        payload = {}

    fields = payload.get('fields') or {}
    if not (fields.get('ingredient_name') or '').strip():
        return JsonResponse({
            'success': False,
            'error': '원료명이 없습니다.',
        }, status=400)

    with transaction.atomic():
        bom, created, matched, score, candidates = register_ingredient_bom(
            request.user, label, fields)

    log_activity(request, 'product', 'ingredient_to_bom', label.my_label_id)
    return JsonResponse({
        'success': True,
        'created': created,
        'matched_existing': matched,
        'match_score': score,
        'candidates': candidates,
        'bom_id': bom.bom_id,
        'ingredient_name': bom.ingredient_name,
        'message': ('BOM 에 원료를 추가했습니다.' if created
                    else '이미 BOM 에 있는 원료라 정보를 갱신했습니다.'),
    })


@login_required
@require_POST
def label_photo_upload(request, label_id):
    """
    사진으로 불러오기에 쓴 **원본 사진**을 문서함의 '한글표시사항도안' 으로 남긴다.

    판독값은 사진에서 나온 것이고, 그 사진이 없으면 나중에 "이 값이 어디서
    왔는지" 를 되짚을 수가 없다. 표시사항은 법적 표시물이라 근거가 남아야 한다.

    **원본을 남긴다.** 영역을 골라 읽었더라도 문서함에는 자르기 전 사진을 넣는다 —
    조각은 판독을 위해 우리가 만든 것이지 사용자가 가진 자료가 아니다.

    미리보기 PDF 저장(upload_label_pdf)과 같은 문서 종류를 쓰고, 같은 방식으로
    판을 올린다. 다만 metadata 의 source 가 다르다 — 그 값으로 "도안을 만들었다"
    와 "판독에 쓴 사진" 을 가른다(제품 목록의 표시사항 완료 판정, 확정 통보
    메일의 PDF 첨부가 그 구분을 본다).
    """
    from django.db.models import Max, Q

    label = _resolve_editable_label(request, label_id)

    uploaded = request.FILES.get('image')
    if not uploaded:
        return JsonResponse({'success': False, 'error': '사진이 없습니다.'}, status=400)
    if uploaded.size > 10 * 1024 * 1024:
        return JsonResponse({
            'success': False,
            'error': f'파일 크기는 10MB 이하여야 합니다 (현재 {uploaded.size / 1024 / 1024:.1f}MB).',
        }, status=400)

    doc_type, _ = DocumentType.objects.get_or_create(
        type_code='LABEL_DESIGN',
        defaults={
            'type_name': '한글표시사항도안',
            'description': '한글표시사항 도안 PDF 및 판독에 사용한 표시사항 사진',
            'required_yn': False, 'active_yn': True, 'display_order': 0,
            'icon': 'bi-file-pdf', 'color': '#e8710a',
            'detection_keywords': '한글표시사항,표시사항도안',
            'expiry_alert_days': 30, 'requires_expiry': False,
        },
    )

    existing = (ProductDocument.objects
                .filter(label=label, document_type=doc_type, active_yn=True)
                .order_by('-version', '-uploaded_datetime').first())
    version_number = 1
    parent_document = None
    if existing:
        parent_document = existing
        latest = ProductDocument.objects.filter(
            Q(document_id=existing.document_id) | Q(parent_document=existing)
        ).aggregate(Max('version'))['version__max'] or 1
        version_number = latest + 1

    _, ext = os.path.splitext(uploaded.name or '')
    ext = (ext or '.jpg').lower()
    product_name = label.prdlst_nm or label.my_label_name or 'label'
    filename = f'표시사항사진_{product_name}_{timezone.now().strftime("%Y%m%d")}{ext}'

    document = ProductDocument.objects.create(
        label=label,
        document_type=doc_type,
        file=uploaded,
        original_filename=filename,
        file_size=uploaded.size,
        file_extension=ext,
        document_title='표시사항 사진 (판독 원본)',
        uploaded_by=request.user,
        parent_document=parent_document,
        version=version_number,
        metadata={'expiry_unlimited': True, 'source': 'ocr_import'},
    )

    # 준수율 카드가 보는 슬롯. 사진도 이 칸을 채운 것은 맞다 - 도안 PDF 를
    # 나중에 저장하면 그때 다시 최신 문서로 바뀐다.
    slot, _created = DocumentSlot.objects.get_or_create(
        label=label, document_type=doc_type,
        defaults={'status': DocumentSlot.SlotStatus.VALID})
    slot.current_document = document
    slot.update_status()
    slot.save()

    from v1.products.models import ProductActivityLog

    ProductActivityLog.objects.create(
        label=label, user=request.user, action='DOCUMENT_UPLOADED',
        details={'file_name': filename, 'document_type': doc_type.type_name,
                 'file_size': uploaded.size, 'source': 'ocr_import',
                 'version': version_number},
    )
    log_activity(request, 'product', 'ocr_photo_saved', label.my_label_id)

    return JsonResponse({
        'success': True,
        'document_id': document.document_id,
        'filename': filename,
        'version': version_number,
        'message': '판독에 사용한 사진을 문서함에 남겼습니다.',
    })


@login_required
@require_POST
def ingredient_photo_upload(request, label_id):
    """
    원료 표시사항 사진을 문서함에 넣고 바로 읽는다. BOM 에는 아직 쓰지 않는다.

    "원료로 등록" 은 두 가지를 한 번에 한다 - 사진을 문서함에 남기고, 그 내용을
    BOM 원료로 만든다. 사진이 문서함에 남아야 나중에 "이 원료 정보가 어디서
    왔는지" 를 되짚을 수 있다(원료 표시사항은 근거 자료다).

    여기서는 저장과 읽기까지만 한다. BOM 쓰기는 사용자가 확인한 뒤
    document_ingredient_photo_to_bom 이 맡는다 - OCR 은 틀리고, 틀린 원료가
    BOM 에 들어가면 배합비·알레르기·표시 문구가 전부 그 위에 쌓인다.
    """
    from v1.products.services.ingredient_photo import (
        parse_ingredient_photo, read_document_image,
    )
    from v1.label.services.ingredient_matching import match_my_ingredient

    label = _resolve_editable_label(request, label_id)

    uploaded = request.FILES.get('image')
    if not uploaded:
        return JsonResponse({'success': False, 'error': '사진이 없습니다.'}, status=400)
    if uploaded.size > 10 * 1024 * 1024:
        return JsonResponse({
            'success': False,
            'error': f'파일 크기는 10MB 이하여야 합니다 (현재 {uploaded.size / 1024 / 1024:.1f}MB).',
        }, status=400)

    doc_type = DocumentType.objects.filter(type_code='INGREDIENT_LABEL').first()
    if doc_type is None:
        return JsonResponse({
            'success': False,
            'error': '"원료 표시사항" 문서 타입이 없습니다. migrate 를 먼저 실행하세요.',
        }, status=500)

    _, ext = os.path.splitext(uploaded.name)
    document = ProductDocument.objects.create(
        label=label,
        document_type=doc_type,
        file=uploaded,
        original_filename=uploaded.name,
        file_size=uploaded.size,
        file_extension=ext.lower(),
        document_title=doc_type.type_name,
        uploaded_by=request.user,
        metadata={'source': 'ingredient_photo_upload'},
    )

    ocr_data, error = read_document_image(document)
    if error:
        # 문서는 남긴다. 사진 자체는 근거 자료로 쓸모가 있고, 사용자가 문서함에서
        # 다시 읽어 볼 수 있다.
        return JsonResponse({
            'success': False,
            'error': error,
            'document_id': document.document_id,
        }, status=400)

    fields = parse_ingredient_photo(ocr_data)
    ingredient, score, candidates = match_my_ingredient(
        request.user, fields.get('ingredient_name') or '')

    return JsonResponse({
        'success': True,
        'document_id': document.document_id,
        'filename': document.original_filename,
        'fields': fields,
        'matched_existing': ingredient is not None,
        'matched_name': ingredient.prdlst_nm if ingredient else '',
        'match_score': score,
        'candidates': candidates,
    })


@login_required
@require_POST
def report_no_lookup(request, label_id):
    """
    품목보고번호로 식약처 등록 정보를 불러온다.

    사진이 없어도 되는 입구다. 번호만 알면 제품명·식품유형·원재료명·제조사가
    나온다 - 사진을 읽는 것보다 정확하다(OCR 을 거치지 않는다).

    제품으로 쓸지 원료로 쓸지는 화면이 정한다. 여기서는 찾아서 돌려주기만 한다.
    """
    _resolve_editable_label(request, label_id)

    try:
        payload = json.loads(request.body or '{}')
    except (ValueError, TypeError):
        payload = {}

    report_no = ''.join(str(payload.get('report_no') or '').split())
    if not report_no:
        return JsonResponse({'success': False, 'error': '품목보고번호를 입력하세요.'},
                            status=400)

    item = FoodItem.objects.filter(prdlst_report_no=report_no).first()
    if item is None:
        return JsonResponse({
            'success': False,
            'error': f'등록된 품목을 찾지 못했습니다 ({report_no}). 번호를 확인해 주세요.',
        }, status=404)

    return JsonResponse({
        'success': True,
        'fields': {
            'prdlst_report_no': item.prdlst_report_no or '',
            'prdlst_nm': item.prdlst_nm or '',
            'prdlst_dcnm': item.prdlst_dcnm or '',
            'rawmtrl_nm': item.rawmtrl_nm or '',
            'bssh_nm': item.bssh_nm or '',
        },
    })


@login_required
@require_POST
def ocr_record_corrections(request, label_id):
    """
    판독 결과를 사용자가 어떻게 고쳤는지 남긴다.

    이 기록이 없으면 "무엇을 얼마나 틀리는지" 를 셀 수 없고, 프롬프트를 고쳐도
    나아졌는지 알 수 없다. 튜닝을 하려 해도 원본과 정답의 쌍이 안 쌓인다.

    고치지 않고 그대로 쓴 것도 남긴다 - 정답률을 재려면 맞은 것도 세야 한다.

    기록에 실패해도 200 을 돌려준다. 값은 이미 화면에 채워진 뒤이고, 이력이
    안 남았다고 사용자에게 오류를 보일 이유가 없다.
    """
    from django.conf import settings

    from v1.label.services.ocr_learning import invalidate, record

    _resolve_editable_label(request, label_id)

    try:
        payload = json.loads(request.body or '{}')
    except (ValueError, TypeError):
        payload = {}

    rows = payload.get('rows') or []
    model = getattr(settings, 'OCR_MODEL', '')
    # 영역을 골라 읽었는지(crop) 사진 전체를 읽었는지(whole).
    # 나눠 재지 않으면 "영역을 고르는 게 나은가" 를 답할 수 없다.
    variant = (payload.get('variant') or '')[:20]
    saved = 0
    corrected = 0
    for row in rows[:60]:      # 한 번에 들어올 수 있는 항목 수의 상한
        field = (row.get('field') or '').strip()
        if not field:
            continue
        entry = record(
            request.user, field,
            row.get('ocr_value'), row.get('final_value'),
            confidence=row.get('confidence'), model=model, variant=variant,
            # 사진만 봤는지, 품목보고 등록 정보와 대조했는지. 나눠 재지 않으면
            # "대조가 정확도를 올렸는가" 를 나중에 숫자로 답할 수 없다.
            source=row.get('source'),
        )
        if entry is not None:
            saved += 1
            corrected += int(entry.corrected)

    if corrected:
        # 다음 판독부터 새 교정이 반영되게 한다
        invalidate()

    return JsonResponse({'success': True, 'saved': saved, 'corrected': corrected})


@login_required
@require_POST
def ocr_apply_extras(request, label_id):
    """
    사진에서 읽은 값 중 기본 정보 탭 밖으로 가는 것을 반영한다.

    영양성분은 별도 탭(iframe)이고 분리배출은 미리보기 설정이라, 기본 정보 탭의
    폼으로는 채울 수가 없다. 화면이 확인받은 것만 여기로 보낸다.

    **고른 것만 쓴다.** 영양성분 저장 API(nutrition_save_api)는 넘어오지 않은
    항목을 빈 값으로 덮으므로 여기서 쓸 수 없다 - 사진에 없던 성분이 지워진다.
    """
    from v1.label.services.ocr_apply import (
        apply_nutrition, apply_recycling_mark, basis_is_total,
        parse_nutrition_basis, to_per_100,
    )

    label = _resolve_editable_label(request, label_id)

    try:
        payload = json.loads(request.body or '{}')
    except (ValueError, TypeError):
        payload = {}

    # **기준을 먼저 읽는다.** 사진의 표는 그 표가 밝힌 기준(총 내용량 87 g 등)으로
    # 인쇄돼 있고, 저장 칸은 언제나 100 g 당이다. 기준을 알아야 환산할 수 있다.
    #
    # 예전에는 값을 먼저 넣고 기준으로 serving_size 만 맞췄다. 분모(87)는
    # 바뀌는데 분자는 인쇄값 그대로라, 화면이 318 kcal 를 275 로 바꿔 보여 주고
    # 검증은 "열량이 맞지 않습니다" 라고 했다. 사진에도 라벨에도 318 인데도.
    basis_text = payload.get('nutrition_basis')
    basis_value, basis_unit = parse_nutrition_basis(basis_text)

    nutrition = payload.get('nutrition') or []
    applied = apply_nutrition(label, to_per_100(nutrition, basis_value))

    # 표의 기준(총 내용량 / 100g당)이 읽혔으면 함께 맞춘다. 기준이 어긋나면
    # 수치는 맞는데 표시가 틀린다.
    if basis_value:
        label.serving_size = basis_value
        label.serving_size_unit = basis_unit or label.serving_size_unit or 'g'
        fields = ['serving_size', 'serving_size_unit']
        # 표의 기준이 총 내용량이면 단위량이 곧 총량이다. 포장개수가 예전 값으로
        # 남아 있으면(2 등) 표의 총량이 그 배수가 되고, 인쇄된 내용량과 영양정보
        # 표의 머리가 서로 다른 총량을 말한다.
        if basis_is_total(basis_text):
            label.units_per_package = '1'
            fields.append('units_per_package')
        label.save(update_fields=fields)
        applied += fields

    # 분리배출은 읽은 문구를 저장용 종류로 바꿔 준다. 종류를 못 정하면 문구만
    # 남기고 켜지 않는다 - 틀린 종류를 넣으면 포장재질 대조 검증이 엉뚱하게 운다.
    from v1.label.services.ocr_apply import map_recycling_mark

    mark_text = (payload.get('recycling_mark_text') or '').strip()
    mark_type = (payload.get('recycling_mark_type') or '').strip()
    if mark_text and not mark_type:
        mark_type, mark_text = map_recycling_mark(mark_text)
    marked = apply_recycling_mark(label, mark_type, mark_text) if mark_text or mark_type else []

    log_activity(request, 'product', 'ocr_apply_extras', label.my_label_id)
    return JsonResponse({
        'success': True,
        'nutrition_applied': len([f for f in applied if not f.endswith('_unit')]),
        'recycling_applied': bool(marked),
        'recycling_type': mark_type,
    })


@login_required
@require_POST
def design_compare_record(request, label_id):
    """
    디자인 시안 대조 결과를 문서함에 남긴다.

    대조만 하고 아무것도 안 남기면 "확인했다" 는 말만 남는다. 누가 언제 어느
    시안과 맞춰 봤고 무엇이 달랐는지가 있어야 절차가 된다 — 인쇄가 나온 뒤에
    "그때 뭘 봤더라" 를 다시 세지 않아도 된다.

    시안 파일은 우리가 만든 도안(LABEL_DESIGN)과 **다른 종류**로 둔다. 하나는
    우리가 낸 것이고 하나는 받은 것이라, 같은 칸에 쌓으면 어느 것이 정본인지
    알 수 없다.
    """
    from v1.products.models import (
        DocumentType, ProductActivityLog, ProductDocument,
    )

    label = _resolve_editable_label(request, label_id)
    design_file = request.FILES.get('design_file')

    try:
        result = json.loads(request.POST.get('result') or '{}')
    except (ValueError, TypeError):
        result = {}

    diff = result.get('diff') or []
    same = int(result.get('same') or 0)

    doc_type, _ = DocumentType.objects.get_or_create(
        type_code='DESIGN_PROOF',
        defaults={
            'type_name': '포장지 시안',
            'description': '디자인 담당자가 만든 포장지 시안. 표시사항과 대조한 기록이 함께 남는다',
            'required_yn': False,
            'active_yn': True,
            'display_order': 1,
            'icon': 'bi-image',
            'color': '#1a73e8',
            'detection_keywords': '시안,도안,포장지',
            'expiry_alert_days': 0,
            'requires_expiry': False,
        },
    )

    document = None
    if design_file:
        if design_file.size > 20 * 1024 * 1024:
            return JsonResponse({'success': False, 'error': '파일 크기는 20MB 를 초과할 수 없습니다.'},
                                status=400)
        latest = ProductDocument.objects.filter(
            label=label, document_type=doc_type, active_yn=True,
        ).order_by('-version', '-uploaded_datetime').first()

        document = ProductDocument.objects.create(
            label=label,
            document_type=doc_type,
            file=design_file,
            original_filename=design_file.name,
            file_size=design_file.size,
            uploaded_by=request.user,
            parent_document=latest,
            version=(latest.version + 1) if latest else 1,
            metadata={
                'expiry_unlimited': True,
                'source': 'design_compare',
                # 대조 기록은 그 파일에 붙어 있어야 한다 — 파일과 결과가
                # 따로 놀면 "이 시안을 본 결과인가" 를 알 수 없다.
                'compare': {
                    'checked_at': timezone.now().isoformat(timespec='seconds'),
                    'checked_by': request.user.get_username(),
                    'diff_count': len(diff),
                    'same_count': same,
                    'diff': diff[:40],      # 화면이 보여 줄 만큼만
                },
            },
        )

    ProductActivityLog.objects.create(
        label=label,
        user=request.user,
        action='DESIGN_COMPARED',
        details={
            'diff_count': len(diff),
            'same_count': same,
            'file_name': design_file.name if design_file else '',
            'fields': [d.get('label') for d in diff][:20],
        },
    )

    log_activity(request, 'product', 'design_compare', label.my_label_id)
    return JsonResponse({
        'success': True,
        'document_id': document.document_id if document else None,
        'version': document.version if document else None,
        'diff_count': len(diff),
    })


def _resolve_editable_label(request, label_id):
    """내 라벨이거나 편집 권한이 있는 공유 라벨을 돌려준다."""
    return get_object_or_404(MyLabel, my_label_id=label_id, user_id=request.user)


@login_required
@require_POST
def rawmtrl_to_bom_preview(request, label_id):
    """
    표시사항의 원재료명 한 줄을 원료 목록으로 쪼개 보여 준다. 저장하지 않는다.

    사진에서 읽은 원재료명은 "새송이버섯(국산)57.64%,과·채가공품/표고버섯채
    (중국산)21.63%(표고버섯,정제수,정제소금,구연산),..." 같은 한 줄이다.
    이걸 원료마다 한 행으로 만들어야 배합비 순서 검사·알레르기 수집·표시 문구가
    올라갈 자리가 생긴다.

    쪼갠 결과를 바로 넣지 않는다. 화면이 목록을 보여 주고 사용자가 고친 뒤
    apply 로 저장한다.
    """
    from v1.label.services.ingredient_matching import load_pool, match_my_ingredient
    from v1.label.services.ingredient_text import parse_ingredient_list

    label = _resolve_editable_label(request, label_id)

    try:
        payload = json.loads(request.body or '{}')
    except (ValueError, TypeError):
        payload = {}

    text = (payload.get('text') or '').strip()
    if not text:
        text = (label.rawmtrl_nm_display or label.rawmtrl_nm or '').strip()
    if not text:
        return JsonResponse({
            'success': False,
            'error': '원재료명이 비어 있습니다.',
        }, status=400)

    parsed = parse_ingredient_list(text)
    if not parsed['items']:
        return JsonResponse({
            'success': False,
            'error': '원재료명에서 원료를 찾지 못했습니다.',
        }, status=400)

    pool = load_pool(request.user)
    rows = []
    for item in parsed['items']:
        ingredient, score, candidates = match_my_ingredient(
            request.user, item['name'], pool=pool)
        rows.append({
            **item,
            'matched': ingredient is not None,
            'matched_name': ingredient.prdlst_nm if ingredient else '',
            'score': score,
            'candidates': candidates,
        })

    return JsonResponse({
        'success': True,
        'rows': rows,
        'allergen_note': parsed['allergen_note'],
        'existing_bom': _active_bom_count(label),
    })


def _active_bom_count(label):
    from v1.bom.models import ProductBOM
    return ProductBOM.objects.filter(parent_label=label, active_yn=True).count()


@login_required
@require_POST
def rawmtrl_to_bom_apply(request, label_id):
    """
    쪼갠 원료들을 BOM 행으로 만든다.

    replace=True 면 기존 BOM 을 비우고 새로 채운다. 사진으로 다시 읽을 때
    같은 원료가 두 벌로 쌓이지 않게 하려는 것이라, 기본은 False 다.

    배합비는 사진에 적힌 값을 그대로 쓴다. 없는 원료는 비워 둔다 — 라벨에
    함량이 적히지 않은 원료가 흔하고, 없는 값을 0 으로 채우면 순서 검사가
    "함량 0" 을 사실로 받아들인다.
    """
    from django.db import transaction

    from v1.bom.models import ProductBOM
    from v1.bom.services import sync_relations_from_bom
    from v1.label.services.ingredient_matching import (
        get_or_create_my_ingredient, load_pool, match_my_ingredient, normalize_name,
    )

    label = _resolve_editable_label(request, label_id)

    try:
        payload = json.loads(request.body or '{}')
    except (ValueError, TypeError):
        payload = {}

    rows = payload.get('rows') or []
    if not rows:
        return JsonResponse({'success': False, 'error': '등록할 원료가 없습니다.'},
                            status=400)

    replace = bool(payload.get('replace'))
    pool = load_pool(request.user)
    created = matched = 0

    with transaction.atomic():
        if replace:
            ProductBOM.objects.filter(parent_label=label).update(active_yn=False)
            order = 0
        else:
            order = _active_bom_count(label)

        for row in rows:
            name = (row.get('name') or '').strip()
            if not name:
                continue

            ratio = row.get('ratio')
            try:
                ratio = float(ratio) if ratio not in (None, '') else None
            except (TypeError, ValueError):
                ratio = None

            subs = (row.get('sub_ingredients') or '').strip()
            origin = (row.get('origin') or '').strip()

            ingredient, _score, _cands = match_my_ingredient(
                request.user, name, pool=pool)
            if ingredient:
                matched += 1
            else:
                ingredient, _new = get_or_create_my_ingredient(
                    request.user,
                    prdlst_nm=name,
                    prdlst_report_no='',
                    prdlst_dcnm='',
                    ingredient_display_name=name,
                    rawmtrl_nm=subs,
                    delete_YN='N',
                )
                pool.setdefault(normalize_name(name), []).append(ingredient)

            bom = ProductBOM.objects.filter(
                parent_label=label, source_ingredient=ingredient).first()
            if bom:
                bom.ingredient_name = name
                bom.raw_material_name = name
                bom.usage_ratio = ratio
                bom.origin = origin
                bom.sub_ingredients = subs
                bom.active_yn = True
                bom.sort_order = order
                bom.save()
            else:
                ProductBOM.objects.create(
                    parent_label=label,
                    created_by=request.user,
                    ingredient_name=name,
                    raw_material_name=name,
                    usage_ratio=ratio,
                    origin=origin,
                    sub_ingredients=subs,
                    source_ingredient=ingredient,
                    sort_order=order,
                    active_yn=True,
                )
                created += 1
            order += 1

        linked, skipped = sync_relations_from_bom(label)

    log_activity(request, 'product', 'rawmtrl_to_bom', label.my_label_id)
    return JsonResponse({
        'success': True,
        'created': created,
        'matched_existing': matched,
        'linked_to_label': linked,
        'skipped_no_ingredient': skipped,
        'total': len(rows),
    })


@login_required
@require_POST
def document_ingredient_photo_to_bom(request, document_id):
    """
    문서함의 원료 표시사항 사진을 읽어 BOM 에 원료 한 건으로 등록한다.

    원료 봉지의 표시사항도 표시사항이라, 완제품 사진에 쓰던 OCR 을 그대로 쓴다.
    다만 읽어낸 값의 뜻이 다르다 - 제품명은 원료명이고, 원재료명은 그 원료의
    하위 원료(복합원재료)다. 그 옮김은 services/ingredient_photo.py 가 한다.

    배합비는 사진에 없다. 봉지에는 그 원료가 완제품에서 몇 %인지 적히지 않는다.
    함량은 비워 두고 BOM 화면에서 사람이 넣는다.

    같은 원료를 두 번 만들지 않는다 - 이름이 비슷하면(RapidFuzz) "내 원료" 의
    기존 레코드에 붙이고, 없을 때만 만든다.

    읽은 결과를 확인 없이 넣는 것이 아니다. 화면이 먼저 preview 로 무엇이 읽혔는지
    보여 주고, 사용자가 확인하면 이 API 가 실제로 등록한다.
    """
    from django.db import transaction

    from v1.bom.models import ProductBOM
    from v1.label.services.ingredient_matching import (
        get_or_create_my_ingredient, load_pool, match_my_ingredient,
    )
    from v1.products.services.ingredient_photo import (
        parse_ingredient_photo, read_document_image,
    )

    doc = get_object_or_404(
        ProductDocument.objects.select_related('label', 'document_type'),
        pk=document_id,
        label__user_id=request.user,
    )

    payload = {}
    if request.body:
        try:
            payload = json.loads(request.body)
        except (ValueError, TypeError):
            payload = {}

    # 화면에서 고친 값이 오면 그것을 쓴다. 없으면 사진을 다시 읽는다.
    fields = payload.get('fields')
    if not fields:
        ocr_data, error = read_document_image(doc)
        if error:
            return JsonResponse({'success': False, 'error': error}, status=400)
        fields = parse_ingredient_photo(ocr_data)

    name = (fields.get('ingredient_name') or '').strip()
    if not name:
        return JsonResponse({
            'success': False,
            'error': '사진에서 원료명을 찾지 못했습니다. 원료명을 직접 입력해 주세요.',
        }, status=400)

    label = doc.label

    with transaction.atomic():
        # 등록 규칙은 품목보고번호 경로와 같다 - register_ingredient_bom 한 곳에 있다.
        bom, created, matched, score, candidates = register_ingredient_bom(
            request.user, label, fields)

        # 문서에 "무엇으로 등록했는지" 를 남긴다. 같은 사진을 두 번 읽지 않게 하고,
        # 나중에 이 BOM 행이 어디서 왔는지 되짚을 수 있다.
        meta = dict(doc.metadata or {})
        meta['ingredient_bom_id'] = bom.bom_id
        meta['ingredient_fields'] = fields
        doc.metadata = meta
        doc.save(update_fields=['metadata'])

    log_activity(request, 'document', 'ingredient_photo_to_bom', document_id)
    return JsonResponse({
        'success': True,
        'created': created,
        'matched_existing': matched,
        'match_score': score,
        'candidates': candidates,
        'bom_id': bom.bom_id,
        'ingredient_name': name,
        'message': ('BOM 에 원료를 추가했습니다.' if created
                    else '이미 BOM 에 있는 원료입니다.'),
    })


@login_required
@require_GET
def document_ingredient_photo_preview(request, document_id):
    """
    원료 사진에서 무엇이 읽히는지 미리 보여 준다. 아무것도 저장하지 않는다.

    등록 전에 사람이 확인할 수 있어야 한다 - OCR 은 틀리고, 틀린 원료가 BOM 에
    들어가면 배합비·알레르기·표시 문구가 전부 그 위에 쌓인다.
    """
    from v1.label.services.ingredient_matching import match_my_ingredient
    from v1.products.services.ingredient_photo import (
        parse_ingredient_photo, read_document_image,
    )

    doc = get_object_or_404(
        ProductDocument.objects.select_related('label', 'document_type'),
        pk=document_id,
        label__user_id=request.user,
    )

    ocr_data, error = read_document_image(doc)
    if error:
        return JsonResponse({'success': False, 'error': error}, status=400)

    fields = parse_ingredient_photo(ocr_data)
    ingredient, score, candidates = match_my_ingredient(
        request.user, fields.get('ingredient_name') or '')

    return JsonResponse({
        'success': True,
        'fields': fields,
        'matched_existing': ingredient is not None,
        'matched_name': ingredient.prdlst_nm if ingredient else '',
        'match_score': score,
        'candidates': candidates,
    })


@login_required
@require_POST
def document_ai_apply_to_bom(request, document_id):
    """
    Group A 전용: AI 추출 원재료 데이터를 BOM에 병합.
    extracted_data의 raw_materials / blend_ratios / origins / allergens 활용.
    """
    import json as _json
    from v1.bom.models import ProductBOM
 
    doc = get_object_or_404(
        ProductDocument.objects.select_related('label'),
        pk=document_id,
        label__user_id=request.user,
    )
 
    meta = doc.metadata or {}
    extracted = meta.get('extracted_data') or {}
    raw_materials = extracted.get('raw_materials') or []
    blend_ratios = extracted.get('blend_ratios') or {}
    origins = extracted.get('origins') or {}
    allergens = extracted.get('allergens') or []
    allergen_str = ', '.join(allergens) if allergens else ''
 
    if not raw_materials:
        return JsonResponse({'error': '추출된 원재료 데이터가 없습니다.'}, status=400)
 
    created_count = 0
    for material in raw_materials:
        if not material:
            continue
        ratio_raw = blend_ratios.get(material)
        try:
            ratio = float(str(ratio_raw).replace('%', '').strip()) if ratio_raw else None
        except (ValueError, TypeError):
            ratio = None
 
        origin = origins.get(material, '')
 
        _, created = ProductBOM.objects.get_or_create(
            parent_label=doc.label,
            ingredient_name=material,
            defaults={
                'usage_ratio': ratio,
                'origin': origin,
                'allergens': allergen_str,
            },
        )
        if created:
            created_count += 1
 
    log_activity(request, 'document', 'ai_apply_to_bom', document_id)
    return JsonResponse({'success': True, 'created': created_count, 'total': len(raw_materials)})


# ============================================================
# 품목제조보고 정보 외부 export API (GAS 제품설명서 자동화 시트 연동용)
# ============================================================
# 식약처 OpenAPI(C002/I1250/I1310/C006)를 GAS가 직접 호출하면 인증키 미승인·타임아웃 등으로
# 조회가 불안정해서, 우리 서버가 ApiEndpoint(v1/common)로 매일 수집해 이미 저장해 둔
# FoodItem(v1/label/models.py, db_table=food_item) 마스터 데이터로 대신 응답한다.
# FoodItem은 prdlst_report_no가 PK이고 사용자/라벨 등록 여부와 무관하게 전체 품목이
# 들어 있으므로, 라벨을 만들지 않은 품목도 조회된다.
#
# 인증: X-Api-Key 헤더 또는 ?key= 쿼리파라미터 (settings.INSPECTION_EXPORT_API_KEY와 일치해야 함,
# 수거검사 export API와 키를 공유한다)
# 쿼리파라미터:
#   - report_no: 품목제조보고번호 (필수)
# 응답: {"data": {...}} 또는 매칭 없을 시 {"data": null}

def product_export_api(request):
    """FoodItem 데이터를 품목제조보고 정보로 JSON export (GAS 등 외부 연동용, 읽기 전용)."""
    api_key = getattr(settings, 'INSPECTION_EXPORT_API_KEY', '')
    if not api_key:
        return JsonResponse({'error': 'API 미설정'}, status=503)

    req_key = request.headers.get('X-Api-Key') or request.GET.get('key', '')
    if not req_key or not hmac.compare_digest(req_key, api_key):
        return JsonResponse({'error': '인증 실패'}, status=401)

    report_no = request.GET.get('report_no', '').strip()
    if not report_no:
        return JsonResponse({'error': 'report_no 파라미터가 필요합니다.'}, status=400)

    row = FoodItem.objects.filter(prdlst_report_no=report_no).first()
    if not row:
        return JsonResponse({'data': None})

    data = {
        'report_no':      row.prdlst_report_no,
        'prdt_nm':        row.prdlst_nm,
        'food_type':      row.prdlst_dcnm,
        'bssh_name':      row.bssh_nm,
        'appearance':     row.dispos,               # 성상
        'usage':          row.prpos,                # 제품용도
        'report_date':    row.prms_dt,               # 품목제조보고일 (YYYYMMDD)
        'change_date':    row.last_updt_dtm,          # 품목제조변경일 (YYYYMMDD)
        'sobigihan':      row.pog_daycnt,
        'packaging':      row.frmlc_mtrqlt,
        'rawmtrl_nm':     row.rawmtrl_nm_sorted or row.rawmtrl_nm,
        'updated_datetime': row.update_datetime.strftime('%Y-%m-%d %H:%M:%S') if row.update_datetime else '',
    }
    return JsonResponse({'data': data})