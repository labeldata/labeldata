import json
import re  # 정규식 처리를 위해 추가
from datetime import datetime, timedelta  # datetime과 timedelta를 import 추가
from decimal import Decimal, InvalidOperation
from urllib.parse import quote  # 파일명 인코딩을 위해 추가

# --- [수정] Third-Party Libraries ---
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation 

# --- [수정] Django Imports ---
from django.conf import settings  # Django settings import 추가
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.core.paginator import Paginator
from django.db import transaction  # 엑셀 업로드 무결성 보증 추가
from django.db.models import F, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse # [추가] URL 생성을 위해 import
from django.utils import timezone  # 추가
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET

# --- [수정] Local Application Imports ---
from .constants import CATEGORY_CHOICES
from .forms import LabelCreationForm, MyIngredientsForm
from .models import (AgriculturalProduct, CountryList, FoodAdditive, FoodItem, 
                     FoodType, ImportedFood, LabelIngredientRelation, MyIngredient, 
                     MyLabel, MyPhrase)

# --- [Import] utils에서 유틸리티 함수 및 상수 import ---
from .utils import ALLERGEN_LIST, GMO_LIST, get_expiry_recommendations, get_search_conditions


# ============================================
# 사용자 활동 로깅 헬퍼 함수
# ============================================

def log_user_activity(request, category, action, target_id=None):
    """
    사용자 활동을 기록하는 헬퍼 함수
    - request: HttpRequest 객체
    - category: 카테고리 (label, ingredient, preview, validation, search, calculator, board)
    - action: 액션 (label_create, ingredient_view 등)
    - target_id: 대상 ID (선택사항)
    """
    try:
        from v1.common.models import UserActivityLog
        if request.user.is_authenticated:
            UserActivityLog.objects.create(
                user=request.user,
                category=category,
                action=action,
                target_id=str(target_id) if target_id else None
            )
    except Exception:
        pass  # 로깅 실패해도 기능에는 영향 없도록


def format_cautions_text(text):
    """
    주의사항/기타표시 텍스트를 미리보기 형식으로 변환
    줄바꿈과 마침표를 " | "로 구분
    
    Args:
        text (str): 원본 텍스트
        
    Returns:
        str: 변환된 텍스트
    """
    if not text:
        return ''
    
    # 1. 줄바꿈을 먼저 |로 변경
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    result = ' | '.join(lines)
    
    # 2. 마침표 뒤에 |가 없으면 추가
    result = re.sub(r'\.\s+(?!\|)', '. | ', result)
    result = re.sub(r'\.(?=[^\s|])', '. | ', result)
    
    return result


def paginate_queryset(queryset, page_number, items_per_page):
    """
    페이징 처리를 수행합니다.
    """
    paginator = Paginator(queryset, items_per_page)
    page_obj = paginator.get_page(page_number)
    page_range = range(max(1, page_obj.number - 5), min(paginator.num_pages + 1, page_obj.number + 5))
    return paginator, page_obj, page_range

def process_sorting(request, default_sort):
    """
    정렬 필드와 정렬 순서를 처리합니다.
    """
    sort_field = request.GET.get("sort", default_sort)
    sort_order = request.GET.get("order", "asc")
    if sort_order == "desc":
        sort_field = f"-{sort_field}"
    return sort_field, sort_order

def get_querystring_without(request, keys):
    """
    요청의 GET 파라미터에서 지정한 키들을 제거한 쿼리 문자열을 반환합니다.
    """
    q = request.GET.copy()
    for key in keys:
        q.pop(key, None)
    return q.urlencode()

def to_bool(val):
    """imported_mode 값을 명확히 True/False로 변환"""
    if isinstance(val, bool):
        return val
    if isinstance(val, int):
        return val != 0
    if isinstance(val, str):
        return val.strip().lower() in ("true", "1", "on", "yes", "y")
    return False

# ------------------------------------------
# View 함수들
# ------------------------------------------

@login_required
def food_item_list(request):
    # 국내제품/수입제품 구분
    food_category = request.GET.get("food_category", "domestic")
    
    # 조회 활동 로깅
    if food_category == "domestic":
        log_user_activity(request, 'search', 'search_domestic')
    else:
        log_user_activity(request, 'search', 'search_import')
    
    # 검색 필드
    search_fields = {
        "prdlst_nm": "prdlst_nm",
        "bssh_nm": "bssh_nm",
        "prdlst_dcnm": "prdlst_dcnm",
        "pog_daycnt": "pog_daycnt",
        "prdlst_report_no": "prdlst_report_no",
        "frmlc_mtrqlt": "frmlc_mtrqlt",
        "rawmtrl_nm": "rawmtrl_nm",
    }
    imported_search_fields = {
        "prduct_korean_nm": "prdlst_nm",
        "itm_nm": "prdlst_dcnm",
        "xport_ntncd_nm": "xport_ntncd_nm",
        "bsn_ofc_name": "bsn_ofc_name",
        "ovsmnfst_nm": "bssh_nm",
        "expirde_dtm": "pog_daycnt",
        "rawmtrl_nm": "rawmtrl_nm",
    }
    # 정렬조건 변경
    if food_category == "imported":
        # 수입제품: 단일 컬럼 정렬만 허용 (기본: -expirde_dtm)
        sort_field, sort_order = process_sorting(request, "-expirde_dtm")
        items_per_page = int(request.GET.get("items_per_page", 10))
        page_number = request.GET.get("page", 1)
        imported_mode = True
        imported_conditions = Q()
        imported_search_values = {}
        has_search_params = False # 검색 파라미터 유무 플래그 추가
        for model_field, query_param in imported_search_fields.items():
            value = request.GET.get(query_param, "").strip()
            if value:
                imported_conditions &= Q(**{f"{model_field}__icontains": value})
                imported_search_values[query_param] = value
                has_search_params = True # 검색 파라미터가 하나라도 있으면 True
        
        if has_search_params:
            # 단일 컬럼 정렬 적용
            imported_items_qs = ImportedFood.objects.filter(imported_conditions).order_by(sort_field)
        else:
            imported_items_qs = ImportedFood.objects.none() # 검색 조건 없으면 빈 쿼리셋

        total_count = imported_items_qs.count()
        paginator, page_obj, page_range = paginate_queryset(imported_items_qs, page_number, items_per_page)
        search_values = imported_search_values
    else:
        # 국내제품: 단일 컬럼 정렬만 허용 (기본: -prms_dt)
        sort_field, sort_order = process_sorting(request, "-prms_dt")
        items_per_page = int(request.GET.get("items_per_page", 10))
        page_number = request.GET.get("page", 1)
        search_conditions, search_values = get_search_conditions(request, search_fields)
        has_search_params = any(search_values.values()) # 검색 파라미터 유무 확인

        if has_search_params:
            # 단일 컬럼 정렬 적용
            food_items_qs = FoodItem.objects.filter(search_conditions).order_by(sort_field)
        else:
            food_items_qs = FoodItem.objects.none() # 검색 조건 없으면 빈 쿼리셋
            
        total_count = food_items_qs.count()
        paginator, page_obj, page_range = paginate_queryset(food_items_qs, page_number, items_per_page)
        # imported_mode 변수를 domestic 케이스에서도 정의
        imported_mode = False 

    querystring_without_page = get_querystring_without(request, ["page"])
    querystring_without_sort = get_querystring_without(request, ["sort", "order"])

    # 검색 조건이 있는지 확인 (위에서 이미 계산된 has_search_params 사용)
    # search_result_count는 total_count를 사용하되, 검색 조건이 있었을 때만 의미가 있음
    search_result_count = total_count if has_search_params else None

    context = {
        "page_obj": page_obj,
        "paginator": paginator,
        "page_range": page_range,
        "search_values": search_values,
        "food_category": food_category,
        "items_per_page": items_per_page,
        "sort_field": sort_field.lstrip('-') if isinstance(sort_field, str) else sort_field, # 정렬 필드에서 '-' 제거
        "sort_order": sort_order,
        "querystring_without_page": querystring_without_page,
        "querystring_without_sort": querystring_without_sort,
        "imported_mode": imported_mode, # food_category 값에 따라 설정된 imported_mode 사용
        "imported_items": page_obj if imported_mode else [], # imported_mode가 True일 때만 imported_items 전달
        "search_result_count": search_result_count,
    }

    return render(request, "label/food_item_list.html", context)

@login_required
def my_label_list(request):
    # 표시사항 목록 조회 로깅
    log_user_activity(request, 'label', 'label_view')
    
    search_fields = {
        "my_label_name": "my_label_name",
        "prdlst_report_no": "prdlst_report_no",
        "prdlst_nm": "prdlst_nm",
        "prdlst_dcnm": "prdlst_dcnm",
        "bssh_nm": "bssh_nm",
        "storage_method": "storage_method",
        "frmlc_mtrqlt": "frmlc_mtrqlt",
        "pog_daycnt": "pog_daycnt",  # 소비기한 검색 추가
    }
    search_conditions, search_values = get_search_conditions(request, search_fields)
    # 표시사항 관리: 작성일 내림차순, 라벨명 오름차순(가나다순)
    sort_field, sort_order = process_sorting(request, "-update_datetime")
    # ↓ 추가: report_no_verify_yn → report_no_verify_YN로 변환
    if sort_field.lstrip('-') == 'report_no_verify_yn':
        sort_field = sort_field.replace('report_no_verify_yn', 'report_no_verify_YN')
    items_per_page = int(request.GET.get("items_per_page", 10))
    page_number = request.GET.get("page", 1)

    ingredient_id = request.GET.get("ingredient_id")
    ingredient_name = None
    
    # 특정 label_id로 필터링 (메인 페이지에서 저장 후 리다이렉트 시)
    label_id = request.GET.get("label_id")

    # 품보 신고 상태 필터 추가
    prdlst_report_status = request.GET.get("prdlst_report_status", "").strip()
    if prdlst_report_status == "completed":
        search_conditions &= Q(report_no_verify_YN="Y")
    elif prdlst_report_status == "pending":
        search_conditions &= Q(report_no_verify_YN="N")

    if ingredient_id:
        linked_label_ids = LabelIngredientRelation.objects.filter(ingredient_id=ingredient_id).values_list("label_id", flat=True)
        labels = MyLabel.objects.filter(user_id=request.user, my_label_id__in=linked_label_ids).filter(search_conditions).order_by(sort_field)
        try:
            ingredient_obj = MyIngredient.objects.get(my_ingredient_id=ingredient_id)
            ingredient_name = ingredient_obj.prdlst_nm or ingredient_obj.ingredient_display_name or ingredient_id
        except MyIngredient.DoesNotExist:
            ingredient_name = ingredient_id
    elif label_id:
        # 특정 label_id만 조회
        labels = MyLabel.objects.filter(user_id=request.user, my_label_id=label_id).filter(search_conditions).order_by(sort_field)
    else:
        labels = MyLabel.objects.filter(user_id=request.user).filter(search_conditions).order_by(sort_field)

    total_count = labels.count()
    paginator, page_obj, page_range = paginate_queryset(labels, page_number, items_per_page)
    querystring_without_sort = get_querystring_without(request, ["sort", "order"])

    # 항상 총 건수 표시 (검색 조건 유무와 관계없이)
    search_result_count = total_count

    context = {
        "page_obj": page_obj,
        "paginator": paginator,
        "page_range": page_range,
        "search_fields": [
            {"name": "prdlst_report_no", "placeholder": "품목제조번호", "value": search_values.get("prdlst_report_no", "")},
            {"name": "prdlst_nm", "placeholder": "제품명", "value": search_values.get("prdlst_nm", "")},
            {"name": "my_label_name", "placeholder": "라벨명", "value": search_values.get("my_label_name", "")},
            {"name": "prdlst_dcnm", "placeholder": "식품유형", "value": search_values.get("prdlst_dcnm", "")},
            {"name": "bssh_nm", "placeholder": "제조사명", "value": search_values.get("bssh_nm", "")},
            {"name": "storage_method", "placeholder": "보관조건", "value": search_values.get("storage_method", "")},
            {"name": "frmlc_mtrqlt", "placeholder": "포장재질", "value": search_values.get("frmlc_mtrqlt", "")},
            {"name": "pog_daycnt", "placeholder": "소비기한", "value": search_values.get("pog_daycnt", "")},
        ],
        "items_per_page": items_per_page,
        "sort_field": sort_field,
        "sort_order": sort_order,
        "querystring_without_sort": querystring_without_sort,
        "ingredient_id": ingredient_id,
        "ingredient_name": ingredient_name,
        "search_result_count": search_result_count,  # 검색 결과 건수 추가
        "prdlst_report_status": prdlst_report_status,  # 품보 신고 상태 추가
    }

    return render(request, "label/my_label_list.html", context)

def create_new_label(request):
    """
    [신규] '신규 작성' 요청을 받아 새 표시사항을 생성하고 편집 페이지로 리디렉션합니다.
    라벨명은 "임시 - 제품명 - N" 형식으로 자동 생성됩니다.
    GET 요청 시 HTML 응답, POST 요청 시 JSON 응답을 반환합니다.
    """
    # 로그인 확인 (AJAX 요청에 대해 JSON 응답)
    if not request.user.is_authenticated:
        if request.method == 'POST' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            login_url = reverse('user_management:login')
            # mode를 포함한 현재 URL을 next로 전달
            mode = request.GET.get('mode', 'detailed')
            next_url = f"{reverse('label:create_new_label')}?mode={mode}"
            return JsonResponse({
                'success': False,
                'error': 'login_required',
                'login_url': login_url,
                'next_url': next_url,
                'message': '로그인이 필요합니다.'
            }, status=401)
        else:
            # GET 요청 시 일반 로그인 리다이렉트
            login_url = reverse('user_management:login')
            return redirect(f'{login_url}?next={request.get_full_path()}')
    
    # 표시사항 생성 로깅
    log_user_activity(request, 'label', 'label_create')
    try:
        base_name = "임시 - 제품명"
        
        # 현재 사용자의 "임시 - 제품명 - X" 라벨들을 조회
        existing_labels = MyLabel.objects.filter(
            user_id=request.user,
            my_label_name__startswith=base_name
        ).values_list('my_label_name', flat=True)

        max_num = 0
        # 정규식을 사용하여 이름에서 숫자 부분을 추출하고 최대값을 찾음
        pattern = re.compile(rf'^{re.escape(base_name)}\s*-\s*(\d+)$')
        for name in existing_labels:
            match = pattern.match(name)
            if match:
                num = int(match.group(1))
                if num > max_num:
                    max_num = num
        
        # 새 라벨 이름 생성 (최대값 + 1)
        new_label_name = f"{base_name} - {max_num + 1}"

        # 새 라벨 객체 생성
        new_label = MyLabel.objects.create(
            user_id=request.user,
            my_label_name=new_label_name
        )

        # 생성된 라벨의 편집 페이지 URL 생성
        mode = request.GET.get('mode', '')
        if request.method == 'POST':
            try:
                body = json.loads(request.body)
                mode = body.get('mode', mode)
            except (json.JSONDecodeError, AttributeError):
                pass
        
        # 간편 모드면 홈 페이지로, 상세 모드면 상세 편집 페이지로 리다이렉트
        if mode == 'simple':
            redirect_url = f'/?label_id={new_label.my_label_id}'
        else:
            redirect_url = reverse('label:label_creation', kwargs={'label_id': new_label.my_label_id})
        
        if request.method == 'POST':
            return JsonResponse({'success': True, 'redirect_url': redirect_url})
        else:
            # GET 요청 시 바로 리디렉션
            return redirect(redirect_url)

    except Exception as e:
        # 예외 발생 시 에러 메시지 반환
        if request.method == 'POST':
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
        else:
            messages.error(request, f'표시사항 생성 실패: {str(e)}')
            return redirect('label:my_label_list')

@login_required
def food_item_detail(request, prdlst_report_no):
    # 국내/수입 제품 상세 조회 로깅
    log_user_activity(request, 'search', 'search_domestic', prdlst_report_no)
    """
    제품 상세 팝업: 가공식품/식품첨가물/수입식품 모두 지원.
    - 수입식품: ImportedFood에서 제품명(한글) 또는 pk로 조회
    - 일반식품: FoodItem에서 품목보고번호로 조회
    """
    # 수입식품: ImportedFood의 prduct_korean_nm 또는 pk(일반적으로 id)로 조회
    imported_item = None
    imported_mode = False

    # 우선 ImportedFood에서 prdlst_report_no와 일치하는 prduct_korean_nm이 있는지 확인
    try:
        imported_item = ImportedFood.objects.get(prduct_korean_nm=prdlst_report_no)
        imported_mode = True
    except ImportedFood.DoesNotExist:
        try:
            imported_item = ImportedFood.objects.get(pk=prdlst_report_no)
            imported_mode = True
        except (ImportedFood.DoesNotExist, ValueError):
            imported_item = None
            imported_mode = False

    if imported_mode and imported_item:
        context = {
            "item": imported_item,
            "imported_mode": True,
        }
        return render(request, "label/food_item_detail.html", context)

    # 일반식품
    food_item = get_object_or_404(FoodItem, prdlst_report_no=prdlst_report_no)
    context = {
        "item": food_item,
        "imported_mode": False,
        # "actions": ... # 필요시 추가
    }
    return render(request, "label/food_item_detail.html", context)


FOODITEM_MYLABEL_MAPPING = {
    'prdlst_report_no': 'prdlst_report_no',
    'prdlst_nm': 'prdlst_nm',
    'prdlst_dcnm': 'prdlst_dcnm',
    'bssh_nm': 'bssh_nm',
    'rawmtrl_nm': 'rawmtrl_nm',
}


@login_required
def save_to_my_label(request, prdlst_report_no):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid request method"}, status=400)

    # 검색에서 표시사항 복사 로깅
    log_user_activity(request, 'search', 'search_label_copy')

    try:
        imported_item = None
        food_item = None

        data = json.loads(request.body) if request.body else {}
        imported_mode = to_bool(data.get("imported_mode", False))
        confirm_flag = data.get("confirm", False)

        if imported_mode:
            imported_item = ImportedFood.objects.filter(pk=prdlst_report_no).first()
            if not imported_item:
                return JsonResponse({"success": False, "error": "해당 수입식품 ID에 대한 데이터를 찾을 수 없습니다."}, status=404)
        else:
            food_item = FoodItem.objects.filter(prdlst_report_no=prdlst_report_no).first()
            if not food_item:
                return JsonResponse({"success": False, "error": "해당 품목제조번호에 대한 데이터를 찾을 수 없습니다."}, status=404)

        if imported_mode and imported_item:
            # 수입식품 라벨 중복 체크(제품명+수입업체명+사용자)
            existing_label = MyLabel.objects.filter(prdlst_nm=imported_item.prduct_korean_nm, bssh_nm=imported_item.bsn_ofc_name, user_id=request.user).first()
            if existing_label and not confirm_flag:
                return JsonResponse({
                    "success": False,
                    "confirm_required": True,
                    "message": "이미 내 라벨에 저장된 수입식품입니다. 저장하시겠습니까?"
                }, status=200)
            # 저장
            new_label = MyLabel.objects.create(
                user_id=request.user,
                my_label_name=f"임시 - {imported_item.prduct_korean_nm}",
                prdlst_nm=imported_item.prduct_korean_nm or "미정",
                prdlst_dcnm=imported_item.itm_nm or "미정",
                bssh_nm=imported_item.bsn_ofc_name or "미정",
                pog_daycnt=imported_item.expirde_dtm or imported_item.expirde_end_dtm or "",
                rawmtrl_nm=imported_item.irdnt_nm or "미정",
                importer_address=imported_item.bsn_ofc_name or "",
                country_of_origin=imported_item.xport_ntncd_nm or imported_item.mnf_ntncn_nm or "",
                # 기타 필드는 필요시 추가
            )
            return JsonResponse({
                "success": True, 
                "message": "내 표시사항으로 저장되었습니다.",
                "label_name": new_label.my_label_name
            })

        # 일반식품 로직
        existing_label = MyLabel.objects.filter(prdlst_report_no=prdlst_report_no, user_id=request.user).first()
        if existing_label and not confirm_flag:
            return JsonResponse(
                {"success": False, "confirm_required": True, "message": "이미 내 라벨에 저장된 항목입니다. 저장하시겠습니까?"}, status=200
            )
        data_mapping = {field: getattr(food_item, field, "") for field in FOODITEM_MYLABEL_MAPPING.keys()}
        label_name = f"임시 - {food_item.prdlst_nm}"
        
        if existing_label and confirm_flag:
            new_label = MyLabel.objects.create(user_id=request.user, my_label_name=label_name, **data_mapping)
            return JsonResponse({
                "success": True, 
                "message": "내 표시사항으로 저장되었습니다.",
                "label_name": new_label.my_label_name
            })
        
        new_label = MyLabel.objects.create(user_id=request.user, my_label_name=label_name, **data_mapping)
        return JsonResponse({
            "success": True, 
            "message": "내 표시사항으로 저장되었습니다.",
            "label_name": new_label.my_label_name
        })
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)



@login_required
def label_creation(request, label_id=None):
    if request.method == 'POST':
        # 표시사항 작성 로깅
        log_user_activity(request, 'label', 'label_write', label_id)
        
        if label_id:
            label = get_object_or_404(MyLabel, my_label_id=label_id, user_id=request.user)
            form = LabelCreationForm(request.POST, instance=label)
        else:
            form = LabelCreationForm(request.POST)

        # 라벨명을 POST 데이터에서 직접 추출하여 폼에 설정
        my_label_name = request.POST.get('my_label_name', '').strip()
        if my_label_name:
            # POST 데이터를 복사하여 수정
            post_data = request.POST.copy()
            post_data['my_label_name'] = my_label_name
            
            # 폼을 새로 생성하여 라벨명이 포함되도록 함
            if label_id:
                form = LabelCreationForm(post_data, instance=label)
            else:
                form = LabelCreationForm(post_data)

        # 체크박스 필드들을 별도로 처리하기 위해 폼 검증 전에 미리 추출
        preservation_types = request.POST.getlist('preservation_type')
        processing_methods = request.POST.getlist('processing_method')
        processing_condition = request.POST.get('processing_condition', '')
        
        if form.is_valid():
            # 폼을 통해 기본 필드들은 처리하되, 문제가 되는 필드들은 별도 처리
            if label_id:
                # 기존 라벨을 업데이트하는 경우
                label = get_object_or_404(MyLabel, my_label_id=label_id, user_id=request.user)
                # 폼의 cleaned_data에서 안전한 값들만 가져와서 설정
                for field_name in form.fields:
                    if field_name in form.cleaned_data and field_name not in ['preservation_type', 'processing_method']:
                        setattr(label, field_name, form.cleaned_data[field_name])
            else:
                # 새 라벨을 생성하는 경우
                label = form.save(commit=False)
                label.user_id = request.user
            
            # 라벨명 다시 한번 확인하여 설정
            if my_label_name:
                label.my_label_name = my_label_name
            
            # hidden 필드에서 식품유형 정보 가져오기
            label.food_group = request.POST.get('food_group')
            label.food_type = request.POST.get('food_type')
            
            # 장기보존식품과 제조방법 정보 저장 (다중 선택 지원)
            label.preservation_type = ', '.join(preservation_types) if preservation_types else ''
            label.processing_method = ', '.join(processing_methods) if processing_methods else ''
            label.processing_condition = processing_condition
            
            # 체크박스 상태 처리
            checkbox_fields = [field for field in request.POST.keys() if field.startswith('chk_')]
            for field_name in checkbox_fields:
                model_field = 'chckd_' + field_name[4:]  # chk_ -> chckd_
                if hasattr(label, model_field):
                    setattr(label, model_field, 'Y')
            
            # 체크되지 않은 체크박스는 N으로 설정
            all_checkbox_fields = [
                'chckd_prdlst_dcnm', 'chckd_prdlst_nm', 'chckd_ingredient_info',
                'chckd_content_weight', 'chckd_weight_calorie', 'chckd_prdlst_report_no',
                'chckd_country_of_origin', 'chckd_storage_method', 'chckd_frmlc_mtrqlt',
                'chckd_bssh_nm', 'chckd_distributor_address', 'chckd_repacker_address',
                'chckd_importer_address', 'chckd_pog_daycnt', 'chckd_rawmtrl_nm_display',
                'chckd_cautions', 'chckd_additional_info', 'chckd_nutrition_text'
            ]
            
            for field_name in all_checkbox_fields:
                if hasattr(label, field_name):
                    corresponding_checkbox = 'chk_' + field_name[6:]  # chckd_ -> chk_
                    if corresponding_checkbox not in checkbox_fields:
                        setattr(label, field_name, 'N')

            # 품목보고번호 변경 감지 및 검증 상태 관리
            # POST 데이터에서 전달된 검증 상태 확인 (프론트엔드에서 검증 후 전달)
            post_verify_status = request.POST.get('report_no_verify_YN', '').strip()
            
            if label_id:
                orig_label = MyLabel.objects.get(my_label_id=label_id)
                
                # 1. POST로 검증 상태가 명시적으로 전달된 경우 우선 사용
                if post_verify_status and post_verify_status in ['Y', 'N', 'F', 'R']:
                    label.report_no_verify_YN = post_verify_status
                # 2. 품목보고번호가 변경된 경우 'N'으로 초기화
                elif orig_label.prdlst_report_no != label.prdlst_report_no:
                    label.report_no_verify_YN = 'N'
                # 3. 번호 변경도 없고 POST 데이터도 없으면 기존 상태 유지
                else:
                    label.report_no_verify_YN = orig_label.report_no_verify_YN
            else:
                # 신규 생성 시: POST 데이터가 있으면 사용, 없으면 'N'
                if post_verify_status and post_verify_status in ['Y', 'N', 'F', 'R']:
                    label.report_no_verify_YN = post_verify_status
                else:
                    label.report_no_verify_YN = 'N'
            
            # 알레르기 정보 저장 (원재료 사용 알레르기만 DB 저장)
            allergens_input = request.POST.get('allergens', '')
            label.allergens = allergens_input
            
            # 알레르기 자동 감지 활동 로깅
            if allergens_input:
                from v1.common.models import UserActivityLog
                UserActivityLog.objects.create(
                    user=request.user,
                    category='validation',
                    action='allergen_auto_detect',
                    target_id=str(label.my_label_id) if label.my_label_id else None
                )

            # 맞춤항목 저장
            custom_fields_json = request.POST.get('custom_fields_json', '')
            if custom_fields_json:
                try:
                    import json
                    custom_fields = json.loads(custom_fields_json)
                    # 최대 10개로 제한
                    if len(custom_fields) > 10:
                        custom_fields = custom_fields[:10]
                    label.custom_fields = custom_fields
                except json.JSONDecodeError as e:
                    label.custom_fields = []
            else:
                label.custom_fields = []

            label.save()
            
            # 표시사항 수정 로깅
            log_user_activity(request, 'label', 'label_update', label.my_label_id)
            
            # AJAX 요청인 경우 JSON 응답 반환
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': '저장되었습니다.',
                    'label_id': str(label.my_label_id)
                })
            
            return redirect('label:label_creation', label_id=label.my_label_id)
        else:
            messages.error(request, '입력 정보에 오류가 있습니다.')
    else:
        has_ingredient_relations = False
        
        if label_id:
            # 기존 라벨 편집
            label = get_object_or_404(MyLabel, my_label_id=label_id, user_id=request.user)
            has_ingredient_relations = label.ingredient_relations.exists()
            # 연결된 원료 개수 구하기
            count_ingredient_relations = LabelIngredientRelation.objects.filter(label_id=label.my_label_id).count()

            # 내 원료에 연결된 원재료명 가져오기
            if has_ingredient_relations:
                relations = LabelIngredientRelation.objects.filter(
                    label_id=label.my_label_id
                ).select_related('ingredient').order_by('relation_sequence')
                
                # 원재료명 정보를 생성 (순서대로)
                ingredients_info = []
                allergens_set = set()
                gmo_set = set()
                shellfish_collected = set()
                shellfish_pattern = re.compile(r'^조개류\(([^)]+)\)$')  # 조개류 패턴 정규식

                # 향료/동일 용도 카운터
                flavor_counter = {}
                purpose_counter = {}

                for relation in relations:
                    ingredient = relation.ingredient
                    food_category = getattr(ingredient, 'food_category', None) or getattr(ingredient, 'food_group', None) or ''
                    display_name = ingredient.ingredient_display_name or ingredient.prdlst_nm or ""
                    # display_name이 콤마로 여러 개일 때 최대 5개까지만 표시 (팝업과 동일)
                    if display_name and "," in display_name:
                        display_name = ", ".join([x.strip() for x in display_name.split(",")][:5])
                    food_type = ingredient.prdlst_dcnm or ""
                    ratio = None
                    try:
                        ratio = float(relation.ingredient_ratio) if relation.ingredient_ratio is not None else None
                    except Exception:
                        ratio = None

                    # 혼합제제(식품첨가물) 처리
                    if food_category == 'additive' and '혼합제제' in display_name:
                        ingredients_info.append(f"혼합제제[{display_name}]")
                        continue

                    # 향료/동일 용도(영양강화제 등) 번호 붙이기
                    # 향료: "향료" 또는 "향료(00향)" 형태
                    if food_category == 'additive' and (display_name.startswith('향료') or re.match(r'^향료(\(.+\))?$', display_name)):
                        flavor_counter[display_name] = flavor_counter.get(display_name, 0) + 1
                        count = flavor_counter[display_name]
                        suffix = f"{count}" if count > 1 else ""
                        ingredients_info.append(f"향료{suffix}{display_name[2:] if display_name.startswith('향료') else ''}")
                        continue
                    # 동일 용도(예: 영양강화제, 산화방지제 등) 번호 붙이기
                    m = re.match(r'^([가-힣]+제)(\(.+\))?$', display_name)
                    if food_category == 'additive' and m:
                        purpose = m.group(1)
                        purpose_counter[purpose] = purpose_counter.get(purpose, 0) + 1
                        count = purpose_counter[purpose]
                        suffix = f"{count}" if count > 1 else ""
                        ingredients_info.append(f"{purpose}{suffix}{m.group(2) or ''}")
                        continue

                    # 정제수는 식품유형만 표시(팝업과 동일)
                    if display_name == '정제수':
                        summary_item = food_type or display_name
                    # 팝업과 동일하게: 첨가물 또는 비율 5% 이상이면 식품유형[원재료명], 그 외는 식품유형만 표시
                    elif (food_category == 'additive') or (ratio is not None and ratio >= 5):
                        if food_type and display_name:
                            summary_item = f"{food_type}[{display_name}]"
                        elif food_type:
                            summary_item = food_type
                        else:
                            summary_item = display_name
                    else:
                        summary_item = food_type or display_name
                    if summary_item:
                        ingredients_info.append(summary_item)
                    # 알레르기/GMO 수집
                    if ingredient.allergens:
                        allergen_list = ingredient.allergens.split(',')
                        for allergen in allergen_list:
                            allergen = allergen.strip() if allergen else ""
                            if not allergen:
                                continue
                            match = shellfish_pattern.match(allergen)
                            if match:
                                # 조개류(홍합,전복) 형태 처리
                                items = [item.strip() for item in match.group(1).split(',') if item.strip()]
                                shellfish_collected.update(items)
                            elif '조개류' in allergen:
                                allergens_set.add(allergen)
                            else:
                                allergens_set.add(allergen)
                    if ingredient.gmo:
                        for g in ingredient.gmo.split(','):
                            g = g.strip() if g else ""
                            if g:
                                gmo_set.add(g)  # GMO를 별도 집합에 추가
                
                # 조개류 항목이 있으면 통합
                if shellfish_collected:
                    shellfish_str = f"조개류({', '.join(sorted(shellfish_collected))})"
                    allergens_set.add(shellfish_str)
                
                # 콤마로 연결하여 원재료명(참고) 필드에 설정
                rawmtrl_nm_str = ", ".join(ingredients_info)
                # 알레르기/GMO 요약 추가 (각각 별도 대괄호로 분리)
                summary_parts = []
                if allergens_set:
                    # 문자열로 안전하게 변환
                    allergens_list = [str(allergen) for allergen in sorted(allergens_set) if allergen]
                    if allergens_list:
                        summary_parts.append(f"[알레르기 성분: {', '.join(allergens_list)}]")
                if gmo_set:
                    # 문자열로 안전하게 변환
                    gmo_list = [str(gmo) for gmo in sorted(gmo_set) if gmo]
                    if gmo_list:
                        summary_parts.append(f"[GMO: {', '.join(gmo_list)}]")
                if summary_parts:
                    rawmtrl_nm_str += f"  {' '.join(summary_parts)}"
                label.rawmtrl_nm = rawmtrl_nm_str
            
            if request.method == 'POST':
                # POST 요청 처리
                form = LabelCreationForm(request.POST, instance=label)
                
                if form.is_valid():
                    label = form.save(commit=False)  # commit=False를 올바르게 사용
                    label.user_id = request.user
                    
                    # hidden 필드에서 식품유형 정보 가져오기
                    label.food_group = request.POST.get('food_group')
                    label.food_type = request.POST.get('food_type')
                    
                    # 변경 사항 저장
                    label.save()
                    # 메시지 제거
                    return redirect('label:label_creation', label_id=label.my_label_id)
                else:
                    messages.error(request, '입력 정보에 오류가 있습니다.')
            else:
                # GET 요청 처리
                form = LabelCreationForm(instance=label)
        else:
            # 새 라벨 생성
            label = None  # label 변수 초기화
            has_ingredient_relations = False
            count_ingredient_relations = 0
            
            if request.method == 'POST':
                # POST 요청 처리
                form = LabelCreationForm(request.POST)
                
                if form.is_valid():
                    label = form.save(commit=False)  # commit=False를 올바르게 사용
                    label.user_id = request.user
                    
                    # hidden 필드에서 식품유형 정보 가져오기
                    label.food_group = request.POST.get('food_group')
                    label.food_type = request.POST.get('food_type')
                    
                    # 변경 사항 저장
                    label.save()
                    # 메시지 제거
                    return redirect('label:label_creation', label_id=label.my_label_id)
                else:
                    messages.error(request, '입력 정보에 오류가 있습니다.')
            else:
                # GET 요청 처리
                form = LabelCreationForm()
        
        # 식품유형 대분류 목록 조회
        food_groups = FoodType.objects.values_list('food_group', flat=True).distinct().order_by('food_group')
        
        # 소분류 목록 필터링 (초기 로드 시)
        current_food_group = getattr(label, 'food_group', '') if label_id else ''
        
        if current_food_group:
            # 선택된 대분류가 있는 경우 해당 대분류에 속하는 소분류만 가져옴
            food_types = FoodType.objects.filter(food_group=current_food_group).values('food_type', 'food_group').order_by('food_type')
        else:
            # 대분류가 선택되지 않은 경우 모든 소분류 가져옴
            food_types = FoodType.objects.values('food_type', 'food_group').order_by('food_type')
            
        # 자주 사용하는 문구 관련 코드 삭제
        
        # 맞춤항목 JSON 직렬화
        custom_fields_json = ''
        if label and label.custom_fields:
            import json
            custom_fields_json = json.dumps(label.custom_fields, ensure_ascii=False)

        context = {
            'form': form,
            'label': label if label_id else None,
            'food_types': food_types,
            'food_groups': food_groups,
            'country_list': CountryList.objects.all(),
            'has_ingredient_relations': has_ingredient_relations,
            'count_ingredient_relations': count_ingredient_relations,
            'custom_fields_json': custom_fields_json,
            # 프론트엔드 상수들은 /static/js/constants.js 파일에서 직접 로드됨
        }
        return render(request, 'label/label_creation.html', context)


@login_required
@csrf_exempt
def save_to_my_ingredients(request, prdlst_report_no=None):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid request method"}, status=400)

    # 검색에서 내원료 복사 로깅
    log_user_activity(request, 'search', 'search_ingredient_copy')

    try:
        data = json.loads(request.body) if request.body else {}
        imported_mode = to_bool(data.get("imported_mode", False))
        confirm = data.get("confirm", False)

        food_item = None
        imported_item = None

        if imported_mode:
            imported_item = ImportedFood.objects.filter(pk=prdlst_report_no).first()
            if not imported_item:
                return JsonResponse({"success": False, "error": "해당 수입식품 ID에 대한 데이터를 찾을 수 없습니다."}, status=404)
            prdlst_dcnm = (imported_item.itm_nm or "").strip()
            prdlst_nm = (imported_item.prduct_korean_nm or "").strip()
            rawmtrl_nm = imported_item.irdnt_nm or "미정"
            # 수입식품도 식품유형에 따라 가공식품/첨가물로 구분
            if prdlst_dcnm and FoodType.objects.filter(food_type=prdlst_dcnm).exists():
                food_category = "processed"
            else:
                food_category = "additive"
            # 제조사명 필드: ovsmnfst_nm (수입식품 테이블에서 제조사명)
            bssh_nm_value = imported_item.ovsmnfst_nm or "미정"
        else:
            food_item = FoodItem.objects.filter(prdlst_report_no=prdlst_report_no).first()
            if not food_item:
                return JsonResponse({"success": False, "error": "해당 품목제조번호에 대한 데이터를 찾을 수 없습니다."}, status=404)
            prdlst_dcnm = (food_item.prdlst_dcnm or "").strip()
            prdlst_nm = (food_item.prdlst_nm or "").strip()
            rawmtrl_nm = food_item.rawmtrl_nm or "미정"
            if prdlst_dcnm and FoodType.objects.filter(food_type=prdlst_dcnm).exists():
                food_category = "processed"
            else:
                food_category = "additive"
            bssh_nm_value = food_item.bssh_nm or "미정"

        # 중복 확인 로직 추가
        if not confirm:
            existing_ingredient = MyIngredient.objects.filter(
                user_id=request.user,
                prdlst_nm=prdlst_nm,
                delete_YN='N'
            ).first()
            
            if existing_ingredient:
                return JsonResponse({
                    "success": False,
                    "confirm_required": True,
                    "message": f"동일한 이름의 내원료 '{prdlst_nm}'가 이미 존재합니다. 그래도 저장하시겠습니까?"
                })

        # ------------------- 추천 로직 통합 -------------------
        mixture_types = [
            "L-글루탐산나트륨제제", "면류첨가알칼리제", "발색제제", "보존료제제",
            "사카린나트륨제제", "타르색소제제", "팽창제제", "향료제제", "혼합제제",
        ]
        ingredient_display_name = prdlst_nm or "미정"
        prdlst_dcnm_for_save = prdlst_dcnm

        # 1. 혼합제제류(지정된 식품유형 포함)는 "식품유형[원재료명]"으로 표시
        if (
            food_category == "additive"
            and any(mix in prdlst_dcnm for mix in mixture_types)
        ):
            ingredient_display_name = f"{prdlst_dcnm}[{rawmtrl_nm}]" if rawmtrl_nm and rawmtrl_nm != "미정" else prdlst_dcnm
            prdlst_dcnm_for_save = prdlst_dcnm

        # 2. 기타 식품첨가물: 원재료명을 기본 표시명으로 저장
        elif food_category == "additive":
            ingredient_display_name = rawmtrl_nm or prdlst_dcnm or "미정"
            prdlst_dcnm_for_save = prdlst_dcnm
        
        # 3. 가공식품 또는 위에서 처리되지 않은 원료
        else:
            ingredient_display_name = rawmtrl_nm or "미정"
            prdlst_dcnm_for_save = prdlst_dcnm
        # ---------------------------------------------------

        new_ingredient = MyIngredient.objects.create(
            user_id=request.user,
            prdlst_report_no=None if imported_mode else food_item.prdlst_report_no,
            prdlst_nm=prdlst_nm or "미정",
            bssh_nm=bssh_nm_value,
            prms_dt="" if imported_mode else (food_item.prms_dt or "00000000"),
            food_category=food_category,
            prdlst_dcnm=prdlst_dcnm_for_save if 'prdlst_dcnm_for_save' in locals() else (prdlst_dcnm or "미정"),
            pog_daycnt=(imported_item.expirde_dtm if imported_mode else food_item.pog_daycnt) or "0",
            frmlc_mtrqlt="" if imported_mode else (food_item.frmlc_mtrqlt or "미정"),
            rawmtrl_nm=rawmtrl_nm or "미정",
            ingredient_display_name=ingredient_display_name,
            delete_YN="N"
        )
        # 메시지 제거 - JSON 응답만 반환 (저장된 원료명 포함)
        return JsonResponse({
            "success": True, 
            "message": "내 원료로 저장되었습니다.",
            "ingredient_name": new_ingredient.prdlst_nm
        })
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)



def ingredient_popup(request):
    label_id = request.GET.get('label_id')
    ingredients_data = []
    has_relations = False
    if label_id:
        # relation_sequence 순서로 명시적 정렬 (모델의 기본 ordering 무시)
        relations = LabelIngredientRelation.objects.filter(label_id=label_id).select_related('ingredient').order_by('relation_sequence')
        for relation in relations:
            ingredient = relation.ingredient
            food_category = getattr(ingredient, 'food_category', None) or getattr(ingredient, 'food_group', None) or ''
            ingredients_data.append({
                'my_ingredient_id': ingredient.my_ingredient_id,
                'ingredient_name': ingredient.prdlst_nm,
                'prdlst_report_no': ingredient.prdlst_report_no or '',
                'ratio': float(relation.ingredient_ratio) if relation.ingredient_ratio else '',
                'food_type': ingredient.prdlst_dcnm or '',
                'food_category': food_category,
                'display_name': ingredient.ingredient_display_name,
                'allergen': ingredient.allergens or '',
                'gmo': ingredient.gmo or '',
                'manufacturer': ingredient.bssh_nm or '',
                'summary_type_flag': ingredient.summary_type_flag or 'Y'  # Y: 식품유형, N: 원재료명
            })
        if relations.exists():
            has_relations = True  # 관계 데이터가 있는 경우 플래그 설정
    # 국가 목록 데이터 추가
    country_list = list(CountryList.objects.values('country_name_ko').order_by('country_name_ko'))
    country_names = [country['country_name_ko'] for country in country_list]
    
    # utils.py에서 알레르기 목록 가져오기
    from .utils import ALLERGEN_LIST
    
    context = {
        'saved_ingredients': ingredients_data,
        'has_relations': has_relations,
        'country_names': country_names,  # 국가 목록 추가
        'allergen_list': ALLERGEN_LIST,  # 알레르기 목록 추가
        'food_types': list(FoodType.objects.all().values('food_type')),
        'agricultural_products': list(AgriculturalProduct.objects.all().values(name_kr=F('rprsnt_rawmtrl_nm'))),
        'food_additives': list(FoodAdditive.objects.all().values('name_kr'))
    }
    return render(request, 'label/ingredient_popup.html', context)


def fetch_food_item(request, prdlst_report_no):
    try:
        food_item = FoodItem.objects.get(prdlst_report_no=prdlst_report_no)
        data = {
            'success': True,
            'prdlst_nm': food_item.prdlst_nm,
            'prdlst_dcnm': food_item.prdlst_dcnm,
            'rawmtrl_nm': food_item.rawmtrl_nm,
            'bssh_nm': food_item.bssh_nm,
        }
    except FoodItem.DoesNotExist:
        data = {'success': False}
    return JsonResponse(data)

@login_required
@csrf_exempt
def check_my_ingredient(request):
    if request.method != 'POST':
        return JsonResponse({'exists': False, 'error': 'Invalid request method'}, status=400)
    try:
        data = json.loads(request.body)
        prdlst_nm = data.get('prdlst_nm', '').strip()
        exists = False
        if prdlst_nm:
            exists = MyIngredient.objects.filter(user_id=request.user, prdlst_nm=prdlst_nm, delete_YN='N').exists()
        return JsonResponse({'exists': exists})
    except Exception as e:
        return JsonResponse({'exists': False, 'error': str(e)}, status=500)
    
@login_required
def my_ingredient_list(request):
    # 원료 목록 조회 로깅
    log_user_activity(request, 'ingredient', 'ingredient_view')
    
    search_fields = {
        'prdlst_nm': 'prdlst_nm',
        'prdlst_report_no': 'prdlst_report_no',
        'prdlst_dcnm': 'prdlst_dcnm',
        'bssh_nm': 'bssh_nm',
        'ingredient_display_name': 'ingredient_display_name',
    }
    search_conditions, search_values = get_search_conditions(request, search_fields)
    # 기존: search_conditions &= Q(delete_YN='N') & Q(user_id=request.user)
    search_conditions &= Q(delete_YN='N') & (Q(user_id=request.user) | Q(user_id__isnull=True))
    sort_field, sort_order = process_sorting(request, 'prdlst_nm')
    # ↓ 추가: report_no_verify_yn → report_no_verify_YN로 변환
    if sort_field.lstrip('-') == 'report_no_verify_yn':
        sort_field = sort_field.replace('report_no_verify_yn', 'report_no_verify_YN')
    items_per_page = int(request.GET.get('items_per_page', 10))
    page_number = request.GET.get('page', 1)
    # 원료관리: 식품구분 오름차순(가나다순), 식품유형 오름차순(가나다순), 원재료명 오름차순
    my_ingredients = MyIngredient.objects.filter(search_conditions).order_by(sort_field)
    paginator, page_obj, page_range = paginate_queryset(my_ingredients, page_number, items_per_page)
    querystring_without_page = get_querystring_without(request, ['page'])
    querydict_sort = request.GET.copy()
    querydict_sort.pop('sort', None)
    querydict_sort.pop('order', None)
    querystring_without_sort = querydict_sort.urlencode()

    context = {
        'page_obj': page_obj,
        'paginator': paginator,
        'page_range': page_range,
        'search_fields': [
            {'name': 'prdlst_nm', 'placeholder': '원재료명', 'value': search_values.get('prdlst_nm', '')},
            {'name': 'prdlst_report_no', 'placeholder': '품목제조번호', 'value': search_values.get('prdlst_report_no', '')},
            {'name': 'prdlst_dcnm', 'placeholder': '식품유형', 'value': search_values.get('prdlst_dcnm', '')},
            {'name': 'bssh_nm', 'placeholder': '제조사명', 'value': search_values.get('bssh_nm', '')},
            {'name': 'ingredient_display_name', 'placeholder': '원료 표시명', 'value': search_values.get('ingredient_display_name', '')},
        ],
        'items_per_page': items_per_page,
        'sort_field': sort_field.lstrip('-'),
        'sort_order': sort_order,
        'querystring_without_page': querystring_without_page,
        'querystring_without_sort': querystring_without_sort,
    }
    return render(request, 'label/my_ingredient_list.html', context)

@login_required
def my_ingredient_list_combined(request):
    label_id = request.GET.get('label_id')
    search_fields = {
        'prdlst_nm': 'prdlst_nm',
        'prdlst_report_no': 'prdlst_report_no',
        'prdlst_dcnm': 'prdlst_dcnm',
        'bssh_nm': 'bssh_nm',
        'ingredient_display_name': 'ingredient_display_name',
        'allergens': 'allergens',
        'gmo': 'gmo',
    }
    search_conditions, search_values = get_search_conditions(request, search_fields)
    search_conditions &= Q(delete_YN='N') & (Q(user_id=request.user) | Q(user_id__isnull=True))

    # 식품구분(카테고리) 검색 지원
    food_category = request.GET.get('food_category', '').strip()
    if food_category:
        search_conditions &= Q(food_category=food_category)

    sort_field, sort_order = process_sorting(request, 'prdlst_nm')
    items_per_page = int(request.GET.get('items_per_page', 10))
    page_number = request.GET.get('page', 1)

    if label_id:
        # 라벨에 연결된 원료만 조회
        ingredient_ids = LabelIngredientRelation.objects.filter(label_id=label_id).values_list('ingredient_id', flat=True)
        my_ingredients = MyIngredient.objects.filter(my_ingredient_id__in=ingredient_ids).filter(search_conditions).order_by(sort_field)
        # 라벨명 가져오기
        label_name = None
        try:
            label_obj = MyLabel.objects.get(my_label_id=label_id)
            label_name = label_obj.my_label_name
        except MyLabel.DoesNotExist:
            label_name = None
    else:
        my_ingredients = MyIngredient.objects.filter(search_conditions).order_by(sort_field)
        label_name = None

    total_count = my_ingredients.count()
    paginator, page_obj, page_range = paginate_queryset(my_ingredients, page_number, items_per_page)
    querystring_without_page = get_querystring_without(request, ['page'])
    querydict_sort = request.GET.copy()
    querydict_sort.pop('sort', None)
    querydict_sort.pop('order', None)
    querystring_without_sort = querydict_sort.urlencode()

    # 검색 조건이 있는지 확인
    has_search_conditions = any(search_values.values()) or food_category
    search_result_count = total_count if has_search_conditions else None

    context = {
        'page_obj': page_obj,
        'paginator': paginator,
        'page_range': page_range,
        'search_fields': [
            {'name': 'prdlst_nm', 'placeholder': '원재료명', 'value': search_values.get('prdlst_nm', '')},
            {'name': 'prdlst_report_no', 'placeholder': '품목제조번호', 'value': search_values.get('prdlst_report_no', '')},
            {'name': 'prdlst_dcnm', 'placeholder': '식품유형', 'value': search_values.get('prdlst_dcnm', '')},
            {'name': 'bssh_nm', 'placeholder': '제조사명', 'value': search_values.get('bssh_nm', '')},
            {'name': 'ingredient_display_name', 'placeholder': '원료 표시명', 'value': search_values.get('ingredient_display_name', '')},
            {'name': 'allergens', 'placeholder': '알레르기', 'value': search_values.get('allergens', '')},
            {'name': 'gmo', 'placeholder': 'GMO', 'value': search_values.get('gmo', '')},
        ],
        'items_per_page': items_per_page,
        'sort_field': sort_field.lstrip('-'),
        'sort_order': sort_order,
        'querystring_without_page': querystring_without_page,
        'querystring_without_sort': querystring_without_sort,
        'label_id': label_id,
        'label_name': label_name,
        'search_result_count': search_result_count,  # 검색 결과 건수 추가
    }
    return render(request, 'label/my_ingredient_list_combined.html', context)

@login_required
def my_ingredient_detail(request, ingredient_id=None):
    # POST일 때 my_ingredient_id 우선 활용
    # my_ingredient_id -> ingredient_id 로 수정하여 NameError 해결
    current_ingredient_id = request.POST.get('my_ingredient_id') or ingredient_id
    if current_ingredient_id:
        # user_id는 request.user 또는 None도 허용 (공용 원료 지원)
        ingredient = get_object_or_404(MyIngredient, my_ingredient_id=current_ingredient_id)
        mode = 'edit'
    else:
        ingredient = MyIngredient(user_id=request.user, delete_YN='N')
        mode = 'create'

    if request.method == 'POST':
        # 공용 원료(user_id=None)는 수정 불가
        if getattr(ingredient, 'user_id', None) is None:
            msg = '공용 원료는 수정할 수 없습니다.'
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': msg})
            else:
                messages.error(request, msg)
                return redirect('label:my_ingredient_detail', ingredient_id=ingredient.my_ingredient_id)
        # 중복 체크: 신규 등록(생성)일 때만 동일한 이름의 내 원료가 이미 존재하면 에러 반환
        prdlst_nm = request.POST.get('prdlst_nm', '').strip()
        ignore_duplicate = request.POST.get('ignore_duplicate', '').upper() == 'Y'
        if prdlst_nm and mode == 'create' and not ignore_duplicate:
            duplicate_qs = MyIngredient.objects.filter(user_id=request.user, prdlst_nm=prdlst_nm, delete_YN='N')
            if duplicate_qs.exists():
                msg = '동일한 이름의 원료가 이미 존재합니다.'
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': msg}, status=400)
                else:
                    messages.error(request, msg)
                    return redirect('label:my_ingredient_detail', ingredient_id=ingredient.my_ingredient_id)
        form = MyIngredientsForm(request.POST, instance=ingredient)
        if form.is_valid():
            new_ingredient = form.save(commit=False)
            new_ingredient.user_id = request.user
            new_ingredient.save()
            # AJAX 요청 시 ingredient_id 항상 반환
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': '내 원료가 성공적으로 저장되었습니다.',
                    'ingredient_id': new_ingredient.my_ingredient_id
                })
            # 메시지 제거
            return redirect('label:my_ingredient_detail', ingredient_id=new_ingredient.my_ingredient_id)
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                # 폼 에러를 문자열로 반환
                errors = {field: str(error) for field, error in form.errors.items()}
                return JsonResponse({'success': False, 'errors': errors, 'message': '입력값 오류'})
    else:
        form = MyIngredientsForm(instance=ingredient)

    context = {
        'ingredient': ingredient,
        'form': form,
        'mode': mode,
        'food_types': list(FoodType.objects.all().values('food_type', 'food_group').order_by('food_type')),
        'agricultural_products': list(AgriculturalProduct.objects.all().values(name_kr=F('rprsnt_rawmtrl_nm'))),
        'food_additives': list(FoodAdditive.objects.all().values('name_kr')),
        'allergen_list': ALLERGEN_LIST,
        'gmo_list': GMO_LIST,
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'label/my_ingredient_detail_partial.html', context)
    else:
        return render(request, 'label/my_ingredient_detail.html', context)

@login_required
@csrf_exempt
def save_ingredients_to_label(request, label_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    # 원재료 표로 입력 로깅
    log_user_activity(request, 'label', 'ingredient_table_input', label_id)
    
    try:
        data = json.loads(request.body)
        ingredients_data = data.get('ingredients', [])
        label = get_object_or_404(MyLabel, my_label_id=label_id, user_id=request.user)

        # 기존 연결 관계를 모두 삭제
        LabelIngredientRelation.objects.filter(label_id=label.my_label_id).delete()

        # 새로운 원재료 정보 저장
        for sequence, ingredient_data in enumerate(ingredients_data, start=1):
            # "정제수"는 공용 MyIngredient(user_id=None, prdlst_nm='정제수')와만 연결
            if ingredient_data.get('ingredient_name') == "정제수":
                ingredient = MyIngredient.objects.filter(user_id=None, prdlst_nm="정제수", delete_YN='N').first()
                if not ingredient:
                    # 공용 정제수 원료가 없으면 이 원료는 건너뜀
                    continue
                relation = LabelIngredientRelation(
                    label=label,
                    ingredient=ingredient,
                    relation_sequence=sequence
                )
                try:
                    ratio_value = ingredient_data.get('ratio', '')
                    if ratio_value and ratio_value.strip():
                        relation.ingredient_ratio = Decimal(ratio_value)
                except (ValueError, InvalidOperation):
                    pass
                relation.save()
                continue

            # 원재료 ID가 있는 경우 기존 원재료 사용
            my_ingredient_id = ingredient_data.get('my_ingredient_id')
            
            # 원재료 ID가 없거나 빈 문자열인 경우 새 원재료 생성
            if not my_ingredient_id:
                # 새 원재료 생성
                ingredient = MyIngredient.objects.create(
                    user_id=request.user,
                    prdlst_nm=ingredient_data.get('ingredient_name', ''),
                    prdlst_report_no=ingredient_data.get('prdlst_report_no', ''),
                    prdlst_dcnm=ingredient_data.get('food_type', ''),
                    ingredient_display_name=ingredient_data.get('display_name', ''),
                    bssh_nm=ingredient_data.get('manufacturer', ''),
                    allergens=ingredient_data.get('allergen', ''),
                    gmo=ingredient_data.get('gmo', ''),
                    summary_type_flag=ingredient_data.get('summary_type_flag', 'Y'),  # summary_type_flag 추가
                    delete_YN='N'
                )
            else:
                # 기존 원재료 사용
                try:
                    ingredient = MyIngredient.objects.get(my_ingredient_id=my_ingredient_id)
                    # 기존 원재료의 summary_type_flag 업데이트
                    ingredient.summary_type_flag = ingredient_data.get('summary_type_flag', 'Y')
                    ingredient.save()
                except MyIngredient.DoesNotExist:
                    # 원재료가 존재하지 않는 경우 새로 생성
                    ingredient = MyIngredient.objects.create(
                        user_id=request.user,
                        prdlst_nm=ingredient_data.get('ingredient_name', ''),
                        prdlst_report_no=ingredient_data.get('prdlst_report_no', ''),
                        prdlst_dcnm=ingredient_data.get('food_type', ''),
                        ingredient_display_name=ingredient_data.get('display_name', ''),
                        bssh_nm=ingredient_data.get('manufacturer', ''),
                        allergens=ingredient_data.get('allergen', ''),
                        gmo=ingredient_data.get('gmo', ''),
                        summary_type_flag=ingredient_data.get('summary_type_flag', 'Y'),  # summary_type_flag 추가
                        delete_YN='N'
                    )
            
            # 원재료와 라벨 사이의 관계 생성
            relation = LabelIngredientRelation(
                label=label,
                ingredient=ingredient,
                relation_sequence=sequence
            )
            
            # 비율 설정
            try:
                ratio_value = ingredient_data.get('ratio', '')
                if ratio_value and ratio_value.strip():
                    relation.ingredient_ratio = Decimal(ratio_value)
            except (ValueError, InvalidOperation):
                # 숫자가 아닌 값이 들어온 경우 비율 설정 안함
                pass
                
            # 관계 저장
            relation.save()
        
        # 원재료명 업데이트 (참고사항에 "원산지 표시대상" 포함된 항목만)
        origin_targets = []
        for ingredient_data in ingredients_data:
            notes = ingredient_data.get('notes', '')
            if notes and '원산지 표시대상' in notes:
                display_name = ingredient_data.get('display_name') or ingredient_data.get('ingredient_name')
                if display_name:
                    origin_targets.append(display_name)
        
        # 원산지 표시대상 원재료명을 라벨에 저장
        if origin_targets:
            label.country_of_origin = ', '.join(origin_targets)
            
        label.save()
        
        # 메시지 제거 - JSON 응답만 반환
        return JsonResponse({'success': True, 'message': '저장되었습니다.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@csrf_exempt
def delete_my_ingredient(request, ingredient_id):
    # 게스트 사용자는 삭제 불가
    if request.user.username == 'guest@labeasylabel.com':
        return JsonResponse({'success': False, 'error': '게스트 계정은 삭제 기능을 사용할 수 없습니다.'})
        
    if request.method == 'POST':
        # 원료 삭제 로깅
        log_user_activity(request, 'ingredient', 'ingredient_delete', ingredient_id)
        
        try:
            # 먼저 ID로 원료 조회
            ingredient = get_object_or_404(MyIngredient, my_ingredient_id=ingredient_id)
            
            # 공용 원료(user_id=None)는 삭제 불가
            if ingredient.user_id is None:
                return JsonResponse({'success': False, 'error': '정제수는 기본 원료로 삭제가 불가합니다.'})
            
            LabelIngredientRelation.objects.filter(ingredient_id=ingredient.my_ingredient_id).delete()
            ingredient.delete_YN = 'Y'
            ingredient.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
@csrf_exempt
def search_ingredient_add_row(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)
    
    try:
        data = json.loads(request.body)
        name = data.get('ingredient_name', '').strip()
        report = data.get('prdlst_report_no', '').strip()
        food_type = data.get('food_type', '').strip()
        manufacturer = data.get('manufacturer', '').strip()
        food_category = data.get('food_category', '').strip()
        allergen = data.get('allergen', '').strip()
        gmo = data.get('gmo', '').strip()
        
        # 기존: qs = MyIngredient.objects.filter(user_id=request.user, delete_YN='N')
        qs = MyIngredient.objects.filter(delete_YN='N').filter(Q(user_id=request.user) | Q(user_id__isnull=True))
        if name:
            qs = qs.filter(prdlst_nm__icontains=name)
        if report:
            qs = qs.filter(prdlst_report_no__icontains=report)
        if food_type:
            qs = qs.filter(prdlst_dcnm__icontains=food_type)
        if manufacturer:
            qs = qs.filter(bssh_nm__icontains=manufacturer)
        if food_category:
            qs = qs.filter(food_category=food_category)
        if allergen:
            qs = qs.filter(allergens__icontains=allergen)
        if gmo:
            qs = qs.filter(gmo__icontains=gmo)
        
        ingredients = list(qs.values(
            'prdlst_nm',
            'prdlst_report_no',
            'prdlst_dcnm',
            'bssh_nm',
            'ingredient_display_name',
            'my_ingredient_id',
            'food_category',  # 식품 구분 필드 추가
            'allergens',  # 알레르기 정보 추가
            'gmo',  # GMO 정보 추가
        ))

        if ingredients:
            return JsonResponse({'success': True, 'ingredients': ingredients})
        else:
            return JsonResponse({'success': False, 'error': '검색 결과가 없습니다.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@csrf_exempt
def quick_register_ingredient(request):
    """빠른 원료 등록 API"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)
    
    # 원재료 빠른 등록 로깅
    log_user_activity(request, 'label', 'ingredient_quick_register')
    
    try:
        data = json.loads(request.body)
        ingredient_name = data.get('ingredient_name', '').strip()
        display_name = data.get('display_name', '').strip()
        food_category = data.get('food_category', '').strip()
        food_type = data.get('food_type', '').strip()
        manufacturer = data.get('manufacturer', '').strip()
        report_no = data.get('report_no', '').strip()
        allergens = data.get('allergens', '').strip()
        gmo = data.get('gmo', '').strip()
        
        if not ingredient_name or not food_category or not food_type:
            return JsonResponse({'success': False, 'error': '필수 항목(원재료명, 식품구분, 식품유형)을 입력해주세요.'}, status=400)
        
        # display_name이 비어있으면 ingredient_name 사용
        if not display_name:
            display_name = ingredient_name
        
        # 새 원료 생성
        new_ingredient = MyIngredient.objects.create(
            user_id=request.user,
            prdlst_nm=ingredient_name,
            food_category=food_category,
            prdlst_dcnm=food_type,
            bssh_nm=manufacturer,
            prdlst_report_no=report_no,
            ingredient_display_name=display_name,
            summary_type_flag='Y',  # 기본값: 식품유형 요약
            allergens=allergens,
            gmo=gmo,
            delete_YN='N'
        )
        
        return JsonResponse({
            'success': True,
            'my_ingredient_id': new_ingredient.my_ingredient_id,
            'message': '원료가 성공적으로 등록되었습니다.'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@csrf_exempt
def verify_ingredients(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid request method"}, status=400)
    try:
        data = json.loads(request.body)
        ing_data = data.get("ingredient", {})
        if isinstance(ing_data, list):
            if len(ing_data) == 1:
                ing_data = ing_data[0]
            else:
                return JsonResponse({"success": False, "error": "Expected a single ingredient dictionary."}, status=400)
        
        results = []
        
        prdlst_report_no = str(ing_data.get("prdlst_report_no", "")).strip()
        
        if prdlst_report_no:
            qs = MyIngredient.objects.filter(
                user_id=request.user,
                prdlst_report_no=prdlst_report_no,
                delete_YN="N"
            )
            if qs.exists():
                record = qs.values(
                    "my_ingredient_id",
                    "prdlst_report_no",
                    "prdlst_nm",
                    "prdlst_dcnm",
                    "ingredient_display_name",
                    "delete_YN"
                ).first()
                results.append(record)
            else:
                results.append({})
        else:
            qs = MyIngredient.objects.filter(
                user_id=request.user,
                prdlst_nm=ing_data.get("ingredient_name", "").strip(),
                prdlst_dcnm=ing_data.get("food_type", "").strip(),
                ingredient_display_name=ing_data.get("display_name", "").strip(),
                delete_YN="N"
            )
            if qs.exists():
                record = qs.values(
                    "my_ingredient_id",
                    "prdlst_report_no",
                    "prdlst_nm",
                    "prdlst_dcnm",
                    "ingredient_display_name",
                    "delete_YN"
                ).first()
                results.append(record)
            else:
                results.append({})
        return JsonResponse({"success": True, "results": results})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)

@login_required
def food_items_count(request):
    total = FoodItem.objects.count()
    one_week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y%m%d")
    new_count = FoodItem.objects.filter(last_updt_dtm__gte=one_week_ago).count()
    # 쉼표 없는 정수로 반환
    return JsonResponse({'total': total, 'new': new_count})

@login_required
def my_labels_count(request):
    total = MyLabel.objects.filter(user_id=request.user).count()
    one_week_ago = datetime.now() - timedelta(days=7)
    new_count = MyLabel.objects.filter(user_id=request.user, update_datetime__gte=one_week_ago).count()
    total_formatted = f"{total:,}"
    return JsonResponse({'total': total_formatted, 'new': new_count})

@login_required
def my_ingredients_count(request):
    total = MyIngredient.objects.filter(user_id=request.user, delete_YN='N').count()
    one_week_ago = datetime.now() - timedelta(days=7)
    new_count = MyIngredient.objects.filter(user_id=request.user, delete_YN='N', update_datetime__gte=one_week_ago).count()
    total_formatted = f"{total:,}"
    return JsonResponse({'total': total_formatted, 'new': new_count})


@login_required
@csrf_exempt
def register_my_ingredient(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)
    
    try:
        data = json.loads(request.body)
        prdlst_nm = data.get('ingredient_name', '').strip()
        # 중복 체크: 동일한 이름의 내 원료가 이미 존재하면 에러 반환
        if MyIngredient.objects.filter(user_id=request.user, prdlst_nm=prdlst_nm, delete_YN='N').exists():
            return JsonResponse({'success': False, 'error': '동일한 이름의 원료가 이미 존재합니다.'}, status=400)
        MyIngredient.objects.create(
            user_id=request.user,
            prdlst_nm=prdlst_nm,
            prdlst_report_no=data.get('prdlst_report_no', ''),
            prdlst_dcnm=data.get('food_type', ''),
            ingredient_display_name=data.get('display_name', ''),
            bssh_nm=data.get('manufacturer', ''),
            delete_YN='N'
        )
        # 메시지 제거 - JSON 응답만 반환
        return JsonResponse({'success': True})
    # 중복된 except 블록을 하나로 합쳐서 올바르게 오류를 처리합니다.
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    


@login_required
def nutrition_calculator_popup(request):
    label_id = request.GET.get('label_id')
    
    nutrition_data = {
        'serving_size': '',
        'serving_size_unit': 'g',
        'units_per_package': '1',
        'display_unit': 'unit',
        'nutrients': {}
    }
    # 모든 영양성분 필드 정의 (추가 항목 포함)
    nutrition_fields = [
        ('calories', 'calories_unit'), ('natriums', 'natriums_unit'), ('carbohydrates', 'carbohydrates_unit'),
        ('sugars', 'sugars_unit'), ('fats', 'fats_unit'), ('trans_fats', 'trans_fats_unit'), ('saturated_fats', 'saturated_fats_unit'),
        ('cholesterols', 'cholesterols_unit'), ('proteins', 'proteins_unit'), ('dietary_fiber', 'dietary_fiber_unit'),
        ('calcium', 'calcium_unit'), ('iron', 'iron_unit'), ('magnesium', 'magnesium_unit'), ('phosphorus', 'phosphorus_unit'),
        ('potassium', 'potassium_unit'), ('zinc', 'zinc_unit'), ('vitamin_a', 'vitamin_a_unit'), ('vitamin_d', 'vitamin_d_unit'),
        ('vitamin_c', 'vitamin_c_unit'), ('thiamine', 'thiamine_unit'), ('riboflavin', 'riboflavin_unit'), ('niacin', 'niacin_unit'),
        ('vitamin_b6', 'vitamin_b6_unit'), ('folic_acid', 'folic_acid_unit'), ('vitamin_b12', 'vitamin_b12_unit'), ('selenium', 'selenium_unit'),
        ('iodine', 'iodine_unit'), ('copper', 'copper_unit'), ('manganese', 'manganese_unit'), ('chromium', 'chromium_unit'),
        ('molybdenum', 'molybdenum_unit'), ('vitamin_e', 'vitamin_e_unit'), ('vitamin_k', 'vitamin_k_unit'), ('biotin', 'biotin_unit'),
        ('pantothenic_acid', 'pantothenic_acid_unit')
    ]
    if label_id:
        try:
            label = get_object_or_404(MyLabel, my_label_id=label_id, user_id=request.user)
            nutrition_data['serving_size'] = label.serving_size
            nutrition_data['serving_size_unit'] = label.serving_size_unit or 'g'
            nutrition_data['units_per_package'] = label.units_per_package or '1'
            nutrition_data['display_unit'] = label.nutrition_display_unit or 'unit'
            nutrition_data['nutrients'] = {}
            for field, unit_field in nutrition_fields:
                value = getattr(label, field, '')
                unit = getattr(label, unit_field, '')
                # 값이 None, 0, '0'이면 빈값('') 처리
                if value is None or value == 0 or value == '0':
                    value = ''
                if unit is None:
                    unit = ''
                nutrition_data['nutrients'][field] = {
                    'value': value,
                    'unit': unit
                }
        except Exception as e:
            pass  # 영양성분 데이터 로딩 오류 무시
    # None 값 처리
    for key, value in nutrition_data.items():
        if value is None:
            nutrition_data[key] = ''
    for nutrient_name, nutrient_data in nutrition_data.get('nutrients', {}).items():
        for key, value in nutrient_data.items():
            if value is None:
                nutrient_data[key] = ''
    context = {
        'nutrition_data': json.dumps(nutrition_data)
    }
    return render(request, 'label/nutrition_calculator_popup.html', context)

def duplicate_label(request, label_id):
    original = get_object_or_404(MyLabel, my_label_id=label_id)  
    original.pk = None  
    original.my_label_name += " (복사본)"
    original.save()
    
    # 표시사항 복사 로깅
    log_user_activity(request, 'label', 'label_copy', original.my_label_id)
    
    return redirect('label:label_creation', label_id=original.my_label_id)

def delete_label(request, label_id):
    # 게스트 사용자는 삭제 불가
    if request.user.username == 'guest@labeasylabel.com':
        messages.error(request, '게스트 계정은 삭제 기능을 사용할 수 없습니다.')
        return redirect('label:my_label_list')
        
    label = get_object_or_404(MyLabel, my_label_id=label_id)
    
    # 표시사항 삭제 로깅
    log_user_activity(request, 'label', 'label_delete', label_id)
    
    label.delete()
    return redirect('label:my_label_list')

@login_required
@csrf_exempt
def bulk_copy_labels(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            ids = data.get("label_ids", [])
            for label_id in ids:
                original = get_object_or_404(MyLabel, my_label_id=label_id, user_id=request.user)
                new_label = MyLabel.objects.get(pk=original.pk)
                new_label.pk = None  
                new_label.my_label_name += " (복사본)"
                new_label.save()
                
                # 선택 복사 로깅
                log_user_activity(request, 'label', 'selection_copy', new_label.my_label_id)
            
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid request method"})

@login_required
@csrf_exempt
@login_required
@csrf_exempt
def bulk_delete_labels(request):
    # 게스트 사용자는 삭제 불가
    if request.user.username == 'guest@labeasylabel.com':
        return JsonResponse({"success": False, "error": "게스트 계정은 삭제 기능을 사용할 수 없습니다."})
        
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            ids = data.get("label_ids", [])
            MyLabel.objects.filter(my_label_id__in=ids, user_id=request.user).delete()
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid request method"})
    

@login_required
@csrf_exempt
def save_nutrition(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)
    
    try:
        data = json.loads(request.body)
        label_id = data.get('label_id')
        
        # 영양성분 저장 로깅
        log_user_activity(request, 'calculator', 'calculator_save', label_id)
        
        label = get_object_or_404(MyLabel, my_label_id=label_id, user_id=request.user)
        
        label.serving_size = data.get('serving_size', '')
        label.serving_size_unit = data.get('serving_size_unit', '')
        label.units_per_package = data.get('units_per_package', '')
        label.nutrition_display_unit = data.get('nutrition_display_unit', '')

        label.nutrition_text = data.get('nutritions', '')

        # 모든 영양성분 필드 저장
        nutrition_fields = [
            'calories', 'calories_unit', 'natriums', 'natriums_unit', 'carbohydrates', 'carbohydrates_unit',
            'sugars', 'sugars_unit', 'fats', 'fats_unit', 'trans_fats', 'trans_fats_unit', 'saturated_fats', 'saturated_fats_unit',
            'cholesterols', 'cholesterols_unit', 'proteins', 'proteins_unit',
            'dietary_fiber', 'dietary_fiber_unit', 'calcium', 'calcium_unit', 'iron', 'iron_unit', 'magnesium', 'magnesium_unit',
            'phosphorus', 'phosphorus_unit', 'potassium', 'potassium_unit', 'zinc', 'zinc_unit', 'vitamin_a', 'vitamin_a_unit',
            'vitamin_d', 'vitamin_d_unit', 'vitamin_c', 'vitamin_c_unit', 'thiamine', 'thiamine_unit', 'riboflavin', 'riboflavin_unit',
            'niacin', 'niacin_unit', 'vitamin_b6', 'vitamin_b6_unit', 'folic_acid', 'folic_acid_unit', 'vitamin_b12', 'vitamin_b12_unit',
            'selenium', 'selenium_unit', 'iodine', 'iodine_unit', 'copper', 'copper_unit', 'manganese', 'manganese_unit',
            'chromium', 'chromium_unit', 'molybdenum', 'molybdenum_unit', 'vitamin_e', 'vitamin_e_unit', 'vitamin_k', 'vitamin_k_unit',
            'biotin', 'biotin_unit', 'pantothenic_acid', 'pantothenic_acid_unit'
        ]
        nutrition_inputs = data.get('nutritionInputs', {})
        for field in nutrition_fields:
            value = ''
            # nutritionInputs에 값이 있으면 그 값을 사용
            if field in nutrition_inputs:
                field_data = nutrition_inputs.get(field)
                # nutritionInputs[field]가 dict일 경우 value 키 사용
                if isinstance(field_data, dict):
                    value = field_data.get('value', '')
                else:
                    value = field_data
            else:
                value = data.get(field, '')
            setattr(label, field, value)

        label.save()
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def food_types_by_group(request):
    """식품 대분류별 소분류 목록을 반환하는 API"""
    try:
        group = request.GET.get('group', '')
        
        # 모델 임포트 확인
        from .models import FoodType, FoodAdditive
        
        # 식품첨가물 또는 혼합제제인 경우 FoodAdditive 테이블에서 데이터 가져오기
        if group == '식품첨가물':
            additives = FoodAdditive.objects.filter(category='식품첨가물').order_by('name_kr')
            food_types_data = [{'food_type': add.name_kr, 'food_group': '식품첨가물'} for add in additives]
            return JsonResponse({
                'success': True,
                'food_types': food_types_data
            })
        elif group == '혼합제제':
            # 혼합제제는 FoodAdditive에서 category가 '혼합제제류'인 것들
            additives = FoodAdditive.objects.filter(category='혼합제제류').order_by('name_kr')
            food_types_data = [{'food_type': add.name_kr, 'food_group': '혼합제제'} for add in additives]
            return JsonResponse({
                'success': True,
                'food_types': food_types_data
            })
        elif group == '농수축산물':
            # 농수축산물 소분류 하드코딩
            food_types_data = [
                {'food_type': '농산물', 'food_group': '농수축산물'},
                {'food_type': '수산물', 'food_group': '농수축산물'},
                {'food_type': '축산물', 'food_group': '농수축산물'}
            ]
            return JsonResponse({
                'success': True,
                'food_types': food_types_data
            })
        
        # 일반 식품유형인 경우
        # 대분류가 지정된 경우 해당 대분류의 소분류만, 아니면 모든 소분류 반환
        if group:
            food_types = FoodType.objects.filter(food_group=group).order_by('food_type')
        else:
            # 모든 소분류 반환 (대분류 선택과 무관하게)
            food_types = FoodType.objects.all().order_by('food_group', 'food_type')
        
        # JSON 응답용 데이터 구성
        food_types_data = []
        for ft in food_types:
            food_types_data.append({
                'food_type': ft.food_type,
                'food_group': ft.food_group
            })
        
        return JsonResponse({
            'success': True,
            'food_types': food_types_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def preview_popup(request):
    """표시사항 미리보기 팝업"""
    label_id = request.GET.get('label_id')
    
    # 미리보기 조회 로깅
    log_user_activity(request, 'preview', 'preview_view', label_id)
    
    if not label_id:
        return JsonResponse({'success': False, 'error': '라벨 ID가 제공되지 않았습니다.'})
    
    try:
        label = get_object_or_404(MyLabel, my_label_id=label_id, user_id=request.user)
        
        # 미리보기 항목 구성
        preview_items = []
        field_mappings = [
            ('my_label_name', '라벨명'),
            ('prdlst_dcnm', '식품유형'),
            ('prdlst_nm', '제품명'),
            ('ingredient_info', '성분명 및 함량'),
           
            ('content_weight', '내용량'),
            ('weight_calorie', '내용량(열량)'),
            ('prdlst_report_no', '품목보고번호'),
            ('country_of_origin', '원산지'),
            ('storage_method', '보관방법'),
            ('frmlc_mtrqlt', '포장재질'),
            ('bssh_nm', '제조원 소재지'),
            ('distributor_address', '유통전문판매원'),
            ('repacker_address', '소분원'),
            ('importer_address', '수입원'),
            ('pog_daycnt', '소비기한'),
            ('rawmtrl_nm_display', '원재료명'),
            ('rawmtrl_nm', '원재료명(참고)'),
            ('cautions', '주의사항'),
            ('additional_info', '기타표시사항')
        ]

        for field, label_text in field_mappings:
            value = getattr(label, field)
            if value:
                # 주의사항과 기타표시는 공통 함수로 변환
                if field in ['cautions', 'additional_info']:
                    value = format_cautions_text(value)
                
                preview_items.append({
                    'id': len(preview_items) + 1,
                    'label': label_text,
                    'value': value
                })

        # 영양성분 정보 구성
        nutrition_items = []
        if label.nutrition_text:
            nutrition_fields = [
                ('calories', '열량', 'kcal'),
                ('natriums', '나트륨', 'mg'),
                ('carbohydrates', '탄수화물', 'g'),
                ('sugars', '당류', 'g'),
                ('fats', '지방', 'g'),
                ('trans_fats', '트랜스지방', 'g'),
                ('saturated_fats', '포화지방', 'g'),
                ('cholesterols', '콜레스테롤', 'mg'),
                               ('proteins', '단백질', 'g')
            ]
            
            for field, label_text, unit in nutrition_fields:
                value = getattr(label, field)
                if value:
                    unit_value = getattr(label, f'{field}_unit', unit)
                    nutrition_items.append({
                        'label': label_text,
                        'value': f'{value}{unit_value}',
                        'dv': ''  # 영양성분 기준치 대비 비율 (필요한 경우 계산)
                    })

        # 알레르기 유발물질과 원산지 표시대상 목록
        allergens = []  # 알레르기 유발물질 목록
        origins = []    # 원산지 표시대상 목록
        
        # 알레르기 관리에서 설정된 알레르기 성분 가져오기
        if label.allergens:
            allergens = [a.strip() for a in label.allergens.split(',') if a.strip()]

        # 영양성분 데이터 nutrition_data (계산기와 동일 구조)
        nutrition_data = {
            'serving_size': label.serving_size or '',
            'serving_size_unit': label.serving_size_unit or 'g',
            'units_per_package': label.units_per_package or '1',
            'display_unit': label.nutrition_display_unit or 'unit',
            'nutrients': {
                'calorie': {
                    'value': label.calories,
                    'unit': label.calories_unit or 'kcal'
                },
                'natrium': {
                    'value': label.natriums,
                    'unit': label.natriums_unit or 'mg'
                },
                'carbohydrate': {
                    'value': label.carbohydrates,
                    'unit': label.carbohydrates_unit or 'g'
                },
                'sugar': {
                    'value': label.sugars,
                    'unit': label.sugars_unit or 'g'
                },
                'afat': {
                    'value': label.fats,
                    'unit': label.fats_unit or 'g'
                },
                'transfat': {
                    'value': label.trans_fats,
                    'unit': label.trans_fats_unit or 'g'
                },
                'satufat': {
                    'value': label.saturated_fats,
                    'unit': label.saturated_fats_unit or 'g'
                },
                'cholesterol': {
                    'value': label.cholesterols,
                    'unit': label.cholesterols_unit or 'mg'
                },
                'protein': {
                    'value': label.proteins,
                    'unit': label.proteins_unit or 'g'
                }
            }
        }

        # 국가명 목록 추가 - 안전한 JSON 변환
        country_list = list(CountryList.objects.all().values_list('country_name_ko', flat=True))
        # None 값 제거 및 빈 문자열 제거
        country_list = [country for country in country_list if country and country.strip()]
        
        # 국가 코드 매핑 데이터 추가 (country_code2 -> 한글명)
        country_mapping = {}
        country_data = CountryList.objects.all().values('country_code2', 'country_name_ko')
        for country in country_data:
            if country['country_code2'] and country['country_name_ko']:
                country_mapping[country['country_code2']] = country['country_name_ko']
        
        # 맞춤항목 데이터 추가
        custom_fields = label.custom_fields if label.custom_fields else []
        
        context = {
            'label': label,  # label 객체를 context에 추가
            'preview_items': preview_items,
            'nutrition_items': nutrition_items,
            'allergens': json.dumps(list(set(allergens)), ensure_ascii=False),  # JSON 직렬화 추가
            'origins': list(set(origins)),       # 중복 제거
            'nutrition_data': json.dumps(nutrition_data, ensure_ascii=False),
            'country_list': json.dumps(country_list, ensure_ascii=False),  # JSON 직렬화
            'country_mapping': json.dumps(country_mapping, ensure_ascii=False),  # 국가 코드 매핑 추가
            'expiry_recommendation_json': json.dumps(get_expiry_recommendations(), ensure_ascii=False),  # 소비기한 권장 데이터 추가
            'custom_fields': json.dumps(custom_fields, ensure_ascii=False),  # 맞춤항목 추가
            # 프론트엔드 상수들은 /static/js/constants.js 파일에서 직접 로드됨
        }
        
        return render(request, 'label/label_preview.html', context)
        
    except MyLabel.DoesNotExist:
        return JsonResponse({'success': False, 'error': '라벨을 찾을 수 없습니다.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def my_ingredient_table_partial(request):
    search_fields = {
        'prdlst_nm': 'prdlst_nm',
        'prdlst_report_no': 'prdlst_report_no',
        'prdlst_dcnm': 'prdlst_dcnm',
        'bssh_nm': 'bssh_nm',
        'ingredient_display_name': 'ingredient_display_name',
    }
    search_conditions, search_values = get_search_conditions(request, search_fields)
    # 기존: search_conditions &= Q(delete_YN='N') & Q(user_id=request.user)
    search_conditions &= Q(delete_YN='N') & (Q(user_id=request.user) | Q(user_id__isnull=True))
    sort_field, sort_order = process_sorting(request, 'prdlst_nm')
    my_ingredients = MyIngredient.objects.filter(search_conditions).order_by(sort_field)
    paginator, page_obj, page_range = paginate_queryset(my_ingredients, request.GET.get('page', 1), request.GET.get('items_per_page', 10))
    context = {
        'page_obj': page_obj,
        'paginator': paginator,
        'page_range': page_range,
    }
    return render(request, 'label/my_ingredient_table.html', context)

@require_POST
@login_required
def save_preview_settings(request):
    """미리보기 설정(prv_*) 저장"""
    try:
        data = json.loads(request.body)
        
        label_id = data.get('label_id')
        label = get_object_or_404(MyLabel, my_label_id=label_id, user_id=request.user)
        
        # 미리보기 설정 변경 로깅
        if data.get('font_size'):
            log_user_activity(request, 'preview', 'preview_size', label_id)
        if data.get('font'):
            log_user_activity(request, 'preview', 'preview_font', label_id)
        
        # 기본 미리보기 설정 저장
        label.prv_layout = data.get('layout')
        label.prv_width = data.get('width')
        label.prv_length = data.get('length')
        label.prv_font = data.get('font')
        label.prv_font_size = data.get('font_size')
        label.prv_letter_spacing = data.get('letter_spacing')
        label.prv_line_spacing = data.get('line_spacing')
        
        # 분리배출마크 정보 저장 (첫 번째 마크만)
        recycling_mark = data.get('recycling_mark', {})
        if recycling_mark:
            label.prv_recycling_mark_enabled = 'Y' if recycling_mark.get('enabled') else 'N'
            
            # marks 배열이 있으면 첫 번째 마크만 저장
            if recycling_mark.get('marks') and isinstance(recycling_mark['marks'], list) and len(recycling_mark['marks']) > 0:
                first_mark = recycling_mark['marks'][0]
                label.prv_recycling_mark_type = first_mark.get('type', '')
                label.prv_recycling_mark_position_x = str(first_mark.get('position_x', ''))
                label.prv_recycling_mark_position_y = str(first_mark.get('position_y', ''))
                label.prv_recycling_mark_text = first_mark.get('text', '')
            else:
                # 구버전 형식 지원
                label.prv_recycling_mark_type = recycling_mark.get('type', '')
                label.prv_recycling_mark_position_x = str(recycling_mark.get('position_x', ''))
                label.prv_recycling_mark_position_y = str(recycling_mark.get('position_y', ''))
                label.prv_recycling_mark_text = recycling_mark.get('text', '')
        else:
            label.prv_recycling_mark_enabled = 'N'
        
        label.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def log_validation(request):
    """규정 검증 로깅 API"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            label_id = data.get('label_id')
            
            # 규정 검증 로깅
            log_user_activity(request, 'validation', 'validation_nutrition', label_id)
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'POST required'})


@login_required
def log_mode_switch(request):
    """간편/상세 모드 전환 로깅 API"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            label_id = data.get('label_id')
            
            # 모드 전환 로깅
            log_user_activity(request, 'label', 'mode_switch', label_id)
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'POST required'})


@login_required
def log_quick_text(request):
    """주의문구/기타문구 빠른 등록 로깅 API"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            label_id = data.get('label_id')
            field_type = data.get('field_type')  # 'cautions' or 'additional_info'
            
            if label_id and field_type:
                action = 'caution_quick_add' if field_type == 'cautions' else 'other_text_quick_add'
                log_user_activity(request, 'label', action, label_id)
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'POST required'})


@login_required
def log_custom_field(request):
    """맞춤항목 등록 로깅 API"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            label_id = data.get('label_id')
            
            if label_id:
                log_user_activity(request, 'label', 'custom_field_use', label_id)
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'POST required'})


@login_required
def log_pdf_save(request):
    """PDF 저장 로깅 API"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            label_id = data.get('label_id')
            source = data.get('source', 'preview')  # 'preview' 또는 'calculator'
            
            if source == 'calculator':
                # 영양성분 계산기 PDF 저장 로깅
                log_user_activity(request, 'calculator', 'calculator_calc', label_id)
            else:
                # 미리보기 PDF 저장 로깅
                log_user_activity(request, 'preview', 'preview_pdf_save', label_id)
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'POST required'})


@login_required
def log_preview_action(request):
    """미리보기 팝업 각 기능별 로깅 API"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            label_id = data.get('label_id')
            action = data.get('action')  # 'preview_view', 'preview_table_settings', 'preview_order', 'preview_text_transform', 'preview_recycling_mark', 'preview_settings_save'
            
            if label_id and action:
                # action이 유효한지 확인
                valid_actions = [
                    'preview_view', 'preview_table_settings', 'preview_order', 
                    'preview_text_transform', 'preview_recycling_mark', 'preview_settings_save'
                ]
                if action in valid_actions:
                    log_user_activity(request, 'preview', action, label_id)
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'POST required'})

@login_required
def imported_food_count(request):
    total = ImportedFood.objects.count()
    one_week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y%m%d")
    # expirde_dtm, expirde_end_dtm, procs_dtm 중 하나라도 최근 1주일 이내인 경우 신규로 간주
    new_count = ImportedFood.objects.filter(
        Q(expirde_dtm__gte=one_week_ago) |
        Q(expirde_end_dtm__gte=one_week_ago) |
        Q(procs_dtm__gte=one_week_ago)
    ).count()
    # 쉼표 없는 정수로 반환
    return JsonResponse({'total': total, 'new': new_count})

@login_required
@csrf_exempt
def linked_labels_count(request, ingredient_id):
    """
    특정 내원료(ingredient_id)와 연결된 표시사항(LabelIngredientRelation) 개수 반환
    """
    count = LabelIngredientRelation.objects.filter(ingredient_id=ingredient_id).count()
    return JsonResponse({'count': count})

@login_required
@csrf_exempt
def linked_ingredient_count(request, label_id):
    """
    Returns the count of ingredients linked to a given label (MyLabel).
    """
    try:
        count = LabelIngredientRelation.objects.filter(label_id=label_id).count()
        return JsonResponse({'count': count})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_POST
def verify_report_no(request):
    data = json.loads(request.body)
    label_id = data.get('label_id')
    prdlst_report_no = data.get('prdlst_report_no', '').strip()
    
    # 품목보고번호 검증 로깅
    log_user_activity(request, 'validation', 'validation_report', label_id)
    
    # 품목보고번호가 없으면 오류
    if not prdlst_report_no:
        return JsonResponse({
            'verified': False, 
            'status': 'error',
            'error_type': 'missing_data',
            'message': '품목보고번호를 입력해주세요.'
        })
    
    # label_id가 있고 로그인된 경우에만 라벨 조회
    label = None
    if label_id and request.user.is_authenticated:
        try:
            label = MyLabel.objects.get(pk=label_id, user_id=request.user)
        except MyLabel.DoesNotExist:
            return JsonResponse({
                'verified': False, 
                'status': 'error',
                'error_type': 'label_not_found',
                'message': '라벨을 찾을 수 없습니다.'
            })
    
    # 1. 품목보고번호 형식 검증 (13~15자리 숫자)
    if not re.match(r'^\d{13,15}$', prdlst_report_no):
        # 형식 오류 상태 저장
        if label:
            label.report_no_verify_YN = 'F'
            label.save(update_fields=['report_no_verify_YN'])
        
        return JsonResponse({
            'verified': False,
            'status': 'format_error',
            'error_type': 'format',
            'message': '품목보고번호는 13~15자리 숫자여야 합니다.'
        })
    
    # 2. 기본 규칙 검증 (첫 4자리가 연도 형식인지 확인)
    year_part = prdlst_report_no[:4]
    try:
        year = int(year_part)
        current_year = 2025
        if year < 1960 or year > current_year + 1:
            # 규칙 오류 상태 저장
            if label:
                label.report_no_verify_YN = 'R'
                label.save(update_fields=['report_no_verify_YN'])
            
            return JsonResponse({
                'verified': False,
                'status': 'rule_error',
                'error_type': 'invalid_year',
                'message': f'품목보고번호의 연도({year})가 유효하지 않습니다. (1960-{current_year + 1})'
            })
    except ValueError:
        # ValueError도 형식 오류로 처리
        if label:
            label.report_no_verify_YN = 'F'
            label.save(update_fields=['report_no_verify_YN'])
        
        return JsonResponse({
            'verified': False,
            'status': 'format_error',
            'error_type': 'format',
            'message': '품목보고번호 형식이 올바르지 않습니다.'
        })
    
    # 3. 중복 검증
    food_item = FoodItem.objects.filter(prdlst_report_no=prdlst_report_no).first()
    if food_item:
        # 중복된 번호 (이미 신고되어 있음) - 해당 제품 정보 반환
        # rawmtrl_nm_sorted가 있으면 우선 사용, 없으면 rawmtrl_nm 사용
        rawmtrl_nm = food_item.rawmtrl_nm_sorted or food_item.rawmtrl_nm or ''
        
        # label이 있는 경우 검증 상태 업데이트
        if label:
            # 기존 품목보고번호와 검색한 번호가 다른 경우 상태를 'N'으로 설정
            if label.prdlst_report_no != prdlst_report_no:
                label.prdlst_report_no = prdlst_report_no
                label.report_no_verify_YN = 'N'
                label.save(update_fields=['prdlst_report_no', 'report_no_verify_YN'])
            else:
                # 번호가 동일한 경우에만 'Y'로 설정
                label.report_no_verify_YN = 'Y'
                label.save(update_fields=['report_no_verify_YN'])
        
        return JsonResponse({
            'verified': True,
            'status': 'completed',
            'message': '식품안전나라에 등록된 제품정보를 가져왔습니다.',
            'product_data': {
                'prdlst_nm': food_item.prdlst_nm or '',
                'prdlst_dcnm': food_item.prdlst_dcnm or '',
                'packaging_material': food_item.frmlc_mtrqlt or '',
                'manufacturer': food_item.bssh_nm or '',
                'rawmtrl_nm': rawmtrl_nm
            }
        })
    else:
        # 등록되지 않은 번호 (신고 가능)
        # FoodItem DB에 없으므로 '미확인' 상태로 설정
        if label:
            label.prdlst_report_no = prdlst_report_no
            label.report_no_verify_YN = 'N'
            label.save(update_fields=['prdlst_report_no', 'report_no_verify_YN'])
        
        return JsonResponse({
            'verified': True,
            'status': 'available',
            'message': '품목보고신고 가능한 번호입니다.'
        })

@login_required
def phrases_data_api(request):
    """내 문구 데이터 JSON 반환 (AJAX)"""
    phrases_data = {}
    categories = CATEGORY_CHOICES
    for category_code, _ in categories:
        phrases_data[category_code] = []
    user_phrases = MyPhrase.objects.filter(
        user_id=request.user,
        delete_YN='N'
    ).order_by('category_name', 'display_order')
    for phrase in user_phrases:
        category = phrase.category_name
        if category not in phrases_data:
            continue
        phrases_data[category].append({
            'id': phrase.my_phrase_id,
            'name': phrase.my_phrase_name,
            'content': phrase.comment_content,
            'note': phrase.note or '',
            'order': phrase.display_order
        })
    return JsonResponse({'success': True, 'phrases': phrases_data})

@login_required
def my_ingredient_calculate_page(request):
    """신규 등록된 원료의 페이지 위치 계산"""
    ingredient_id = request.GET.get('ingredient_id')
    if not ingredient_id:
        return JsonResponse({'success': False, 'error': '원료 ID가 필요합니다.'})
    
    try:
        # 검색 조건 가져오기
        search_fields = {
            'prdlst_nm': 'prdlst_nm',
            'prdlst_report_no': 'prdlst_report_no',
            'prdlst_dcnm': 'prdlst_dcnm',
            'bssh_nm': 'bssh_nm',
            'ingredient_display_name': 'ingredient_display_name',
        }
        search_conditions, search_values = get_search_conditions(request, search_fields)
        search_conditions &= Q(delete_YN='N') & (Q(user_id=request.user) | Q(user_id__isnull=True))

        # 식품구분(카테고리) 검색 지원
        food_category = request.GET.get('food_category', '').strip()
        if food_category:
            search_conditions &= Q(food_category=food_category)

        # 알레르기, GMO 검색조건 추가
        allergens = request.GET.get('allergens', '').strip()
        if allergens:
            search_conditions &= Q(allergens__icontains=allergens)
        gmo = request.GET.get('gmo', '').strip()
        if gmo:
            search_conditions &= Q(gmo__icontains=gmo)
        
        items_per_page = int(request.GET.get('items_per_page', 10))
        
        # 전체 원료 리스트에서 해당 원료의 위치 찾기
        my_ingredients = MyIngredient.objects.filter(search_conditions).order_by('-prms_dt', 'food_category', 'prdlst_nm')
        
        # 해당 원료의 인덱스 찾기
        ingredient_index = None
        for index, ingredient in enumerate(my_ingredients):
            if str(ingredient.my_ingredient_id) == str(ingredient_id):
                ingredient_index = index
                break
        
        if ingredient_index is not None:
            # 페이지 계산 (0부터 시작하는 인덱스를 1부터 시작하는 페이지로 변환)
            target_page = (ingredient_index // items_per_page) + 1
            return JsonResponse({
                'success': True, 
                'page': target_page,
                'index': ingredient_index,
                'total_items': my_ingredients.count()
            })
        else:
            return JsonResponse({'success': False, 'error': '해당 원료를 찾을 수 없습니다.'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def my_ingredient_pagination_info(request):
    """페이지네이션 정보 반환"""
    try:
        search_fields = {
            'prdlst_nm': 'prdlst_nm',
            'prdlst_report_no': 'prdlst_report_no',
            'prdlst_dcnm': 'prdlst_dcnm',
            'bssh_nm': 'bssh_nm',
            'ingredient_display_name': 'ingredient_display_name',
        }
        search_conditions, search_values = get_search_conditions(request, search_fields)
        search_conditions &= Q(delete_YN='N') & (Q(user_id=request.user) | Q(user_id__isnull=True))

        # 식품구분(카테고리) 검색 지원
        food_category = request.GET.get('food_category', '').strip()
        if food_category:
            search_conditions &= Q(food_category=food_category)

        # 알레르기, GMO 검색조건 추가
        allergens = request.GET.get('allergens', '').strip()
        if allergens:
            search_conditions &= Q(allergens__icontains=allergens)
        gmo = request.GET.get('gmo', '').strip()
        if gmo:
            search_conditions &= Q(gmo__icontains=gmo)

        items_per_page = int(request.GET.get('items_per_page', 10))
        page_number = request.GET.get('page', 1)
        sort_field, sort_order = process_sorting(request, 'prdlst_nm')
        if sort_field.lstrip('-') == 'report_no_verify_yn':
            sort_field = sort_field.replace('report_no_verify_yn', 'report_no_verify_YN')

        my_ingredients = MyIngredient.objects.filter(search_conditions).order_by(sort_field)
        paginator, page_obj, page_range = paginate_queryset(my_ingredients, page_number, items_per_page)

        return JsonResponse({
            'success': True,
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'item_count': len(page_obj.object_list),
            'total_items': paginator.count
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def _parse_short_names(short_name_value):
    """간략명 파싱 헬퍼 함수"""
    if not short_name_value:
        return []
    short_name_str = str(short_name_value).strip()
    return [s.strip() for s in short_name_str.split(',') 
            if s.strip() and s.strip() not in {"Y", "N", "-"}]


@login_required
@csrf_exempt
def get_additive_regulation(request):
    """식품첨가물의 표시규정 정보 조회 - 구조화된 데이터 반환"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

    try:
        data = json.loads(request.body)
        food_type = data.get('food_type', '').strip()
        prdlst_nm = data.get('prdlst_nm', '').strip()

        if not food_type:
            return JsonResponse({'success': True, 'has_regulation': False})

        # 1. 향료, 천연향료, 합성향료 추천 로직
        if food_type in ["향료", "천연향료", "합성향료"] and "향료제제" not in food_type:
            m = re.search(r'([가-힣\s]+)\s*향', prdlst_nm)
            flavor_name = m.group(1).strip() if m else ""
            
            buttons = []
            header = ""
            
            if food_type == "향료":
                if flavor_name:
                    buttons = [
                        {"value": "향료"},
                        {"value": f"향료({flavor_name}향)"}
                    ]
                else:
                    buttons = [
                        {"value": "향료"},
                        {"value": "향료(OO향)"}
                    ]
                header = "※ 식품첨가물의 유형이 \"향료\"인 경우에는 선택하여 입력하세요."
            else:
                if flavor_name:
                    buttons = [
                        {"value": food_type},
                        {"value": f"{food_type}({flavor_name}향)"}
                    ]
                else:
                    buttons = [{"value": food_type}]
                header = f"※ 식품첨가물의 유형이 \"{food_type}\"인 경우에는 선택하여 입력하세요."
            
            return JsonResponse({
                'success': True,
                'has_regulation': True,
                'header': header,
                'buttons': buttons
            })

        # 2. 기타 식품첨가물 (표 정보 포함)
        try:
            additive = FoodAdditive.objects.get(name_kr=food_type)
            
            buttons = []
            applicable_tables = []
            table_info = {}
            
            # 표 4: 명칭+용도 (명칭(용도1), 명칭(용도2) 형태만 생성)
            # 표 4: 명칭+용도 (명칭(용도1), 명칭(용도2) 형태만 생성)
            if str(additive.alias_4).strip().upper() == 'Y':
                applicable_tables.append("4")
                table_info['4'] = '명칭+용도'
                if additive.name_kr:
                    # 각 용도 필드를 체크하여 'Y'로 설정된 용도 추출
                    purpose_fields = [
                        ('color_agent', '착색료'),
                        ('sweetener', '감미료'),
                        ('nutrient_enhancer', '영양강화제'),
                        ('preservative', '보존료'),
                        ('antioxidant', '산화방지제'),
                        ('bleaching_agent', '표백제'),
                        ('color_fixative', '발색제'),
                        ('stabilizer', '안정제'),
                        ('emulsifier', '유화제'),
                        ('thickener', '증점제'),
                        ('coagulant', '응고제'),
                        ('leavening_agent', '팽창제'),
                        ('sterilizer', '살균제'),
                        ('coating_agent', '피막제'),
                    ]
                    
                    purposes = []
                    for field_name, purpose_label in purpose_fields:
                        field_value = getattr(additive, field_name, None)
                        if field_value and str(field_value).strip().upper() == 'Y':
                            purposes.append(purpose_label)
                    
                    # "명칭(용도)" 형태의 버튼 생성
                    for purpose in purposes:
                        buttons.append({"value": f"{additive.name_kr}({purpose})"})

            # 표 5: 명칭or간략명 (명칭, 간략명1, 간략명2... 각각 별도 버튼)
            if str(additive.alias_5).strip().upper() == 'Y':
                applicable_tables.append("5")
                table_info['5'] = '명칭or간략명'
                # 1. 명칭 버튼 추가
                if additive.name_kr:
                    buttons.append({"value": additive.name_kr})
                # 2. 간략명 처리
                for short_name in _parse_short_names(additive.short_name):
                    buttons.append({"value": short_name})

            # 표 6: 명칭or간략명or용도 (명칭, 간략명1, 간략명2, 용도1, 용도2... 각각 별도 버튼)
            if str(additive.alias_6).strip().upper() == 'Y':
                applicable_tables.append("6")
                table_info['6'] = '명칭or간략명or용도'
                
                # 1. 명칭 버튼 추가
                if additive.name_kr:
                    buttons.append({"value": additive.name_kr})
                
                # 2. 간략명 처리
                for short_name in _parse_short_names(additive.short_name):
                    buttons.append({"value": short_name})
                
                # 3. 용도 처리 (각 용도 필드를 체크)
                purpose_fields = [
                    ('color_agent', '착색료'),
                    ('sweetener', '감미료'),
                    ('nutrient_enhancer', '영양강화제'),
                    ('preservative', '보존료'),
                    ('antioxidant', '산화방지제'),
                    ('bleaching_agent', '표백제'),
                    ('color_fixative', '발색제'),
                    ('stabilizer', '안정제'),
                    ('emulsifier', '유화제'),
                    ('thickener', '증점제'),
                    ('coagulant', '응고제'),
                    ('leavening_agent', '팽창제'),
                    ('sterilizer', '살균제'),
                    ('coating_agent', '피막제'),
                ]
                
                for field_name, purpose_label in purpose_fields:
                    field_value = getattr(additive, field_name, None)
                    if field_value and str(field_value).strip().upper() == 'Y':
                        buttons.append({"value": purpose_label})

            # 표 정보가 있으면 반환
            if buttons:
                table_numbers = ", ".join(applicable_tables)
                # 가장 큰 표 번호의 타입을 헤더에 표시
                last_table = applicable_tables[-1] if applicable_tables else None
                table_type = f"({table_info[last_table]})" if last_table and last_table in table_info else ""
                header = f"※ 식품첨가물공전 표 {table_numbers}{table_type}에 해당되어 선택하여 입력하세요."
                
                return JsonResponse({
                    'success': True,
                    'has_regulation': True,
                    'header': header,
                    'buttons': buttons
                })
            
            # 표 정보가 없으면 빈 반환
            return JsonResponse({'success': True, 'has_regulation': False})

        except FoodAdditive.DoesNotExist:
            return JsonResponse({'success': True, 'has_regulation': False})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@csrf_exempt
def export_labels_excel(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)
    try:
        data = json.loads(request.body)
        label_ids = data.get('label_ids', [])
        if not label_ids:
            return JsonResponse({'success': False, 'error': '선택된 라벨이 없습니다.'}, status=400)
        labels = MyLabel.objects.filter(my_label_id__in=label_ids, user_id=request.user).order_by('my_label_id')
        if not labels.exists():
            return JsonResponse({'success': False, 'error': '다운로드할 데이터가 없습니다.'}, status=400)

        # --- [수정] 원산지 영문/코드 -> 한글 변환을 위한 통합 매핑 생성 ---
        country_map = {}
        for country in CountryList.objects.all():
            if country.country_name_ko:
                # 영문명 -> 한글명
                if hasattr(country, 'country_name_en') and country.country_name_en:
                    country_map[country.country_name_en] = country.country_name_ko
                # 2자리 코드 -> 한글명
                if hasattr(country, 'country_code2') and country.country_code2:
                    country_map[country.country_code2] = country.country_name_ko

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "표시사항"

        headers = [
            '번호', '품목보고번호', '제품명', '라벨명', '식품유형', '제조사명', '최종수정일자',
            '성분명 및 함량', '내용량', '내용량(열량)', '원산지', '보관방법', '포장재질',
            '유통전문판매원', '소분원', '수입원', '소비기한', '원재료명(표시)', '원재료명(참고)', '주의사항', '기타표시사항', '영양성분'
        ]
        ws.append(headers)

        # --- 디자인 서식 적용 ---
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        font_size_9 = Font(size=9)
        ws.freeze_panes = 'E2' # 1행 및 A-D열 고정

        for cell in ws[1]:
            cell.fill = header_fill

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['R'].width = 25
        ws.column_dimensions['S'].width = 25
        ws.column_dimensions['T'].width = 25
        ws.column_dimensions['U'].width = 25

        for idx, label in enumerate(labels, start=1):
            update_dt = ''
            if label.update_datetime:
                update_dt = label.update_datetime.strftime('%y-%m-%d %H:%M')

            # --- [수정] 원산지 한글 변환 로직 ---
            country_of_origin_ko = []
            if label.country_of_origin:
                origin_list = [origin.strip() for origin in label.country_of_origin.split(',')]
                for origin_item in origin_list:
                    # 통합된 country_map을 사용하여 한글명 조회, 없으면 원본 사용
                    country_of_origin_ko.append(country_map.get(origin_item, origin_item))
            country_of_origin_display = ', '.join(country_of_origin_ko)

            row = [
                idx, label.prdlst_report_no, label.prdlst_nm, label.my_label_name,
                label.prdlst_dcnm, label.bssh_nm, update_dt, label.ingredient_info,
                label.content_weight, label.weight_calorie, country_of_origin_display,
                label.storage_method, label.frmlc_mtrqlt, label.distributor_address,
                label.repacker_address, label.importer_address, label.pog_daycnt,
                label.rawmtrl_nm_display, label.rawmtrl_nm, label.cautions,
                label.additional_info, label.nutrition_text,
            ]
            ws.append(row)

        # 모든 셀에 폰트 크기 적용
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font_size_9

        # 파일명에서 사용자 ID 제거
        today_str = timezone.now().strftime('%y%m%d')
        filename = f'LabelData_표시사항_{today_str}.xlsx'
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        encoded_filename = quote(filename)
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
        
        wb.save(response)
        return response
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
       
@login_required
def download_my_ingredients_excel(request):
    """
    현재 사용자의 '내 원료' 목록을 엑셀 파일로 다운로드합니다.
    """
    # ... (기존의 queryset 필터링 로직은 그대로 유지) ...
    search_fields = {
        'prdlst_nm': 'prdlst_nm',
        'prdlst_report_no': 'prdlst_report_no',
        'prdlst_dcnm': 'prdlst_dcnm',
        'bssh_nm': 'bssh_nm',
        'ingredient_display_name': 'ingredient_display_name',
    }
    search_conditions, _ = get_search_conditions(request, search_fields)
    search_conditions &= Q(delete_YN='N') & (Q(user_id=request.user.id) | Q(user_id__isnull=True))
    food_category = request.GET.get('food_category', '').strip()
    if food_category:
        search_conditions &= Q(food_category=food_category)
    queryset = MyIngredient.objects.filter(search_conditions).order_by('prdlst_nm')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = '내 원료 목록'

    headers = ['원료명', '품목보고번호', '식품구분', '식품유형', '원료표시명', '제조사'] + ALLERGEN_LIST + GMO_LIST
    sheet.append(headers)

    # --- 디자인 서식 적용 ---
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    font_size_9 = Font(size=9)
    sheet.freeze_panes = 'A2'

    for cell in sheet[1]:
        cell.fill = header_fill

    # 열 너비 설정
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 13
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 13
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20

    # 식품구분 드롭다운 목록 설정
    food_category_options = '"가공식품,식품첨가물,농수산물,정제수"'
    food_category_dv = DataValidation(type="list", formula1=food_category_options, allow_blank=True)
    sheet.add_data_validation(food_category_dv)
    food_category_dv.add('C2:C1000')

    # --- [수정] 알레르기/GMO 드롭다운 목록 설정 ---
    allergen_gmo_dv = DataValidation(type="list", formula1='"O"', allow_blank=True)
    sheet.add_data_validation(allergen_gmo_dv)

    # 알레르기/GMO 열 너비 설정 및 드롭다운 적용
    start_col = 7 # G열부터 시작
    for i in range(len(ALLERGEN_LIST) + len(GMO_LIST)):
        col_letter = get_column_letter(start_col + i)
        sheet.column_dimensions[col_letter].width = 3
        allergen_gmo_dv.add(f'{col_letter}2:{col_letter}1000') # 드롭다운 적용

    # 데이터 추가
    food_category_map = {'processed': '가공식품', 'additive': '식품첨가물', 'agricultural': '농수산물', 'water': '정제수'}
    for ingredient in queryset:
        current_allergens = [a.strip() for a in (ingredient.allergens or "").split(',') if a.strip()]
        current_gmos = [g.strip() for g in (ingredient.gmo or "").split(',') if g.strip()]
        
        row_data = [
            ingredient.prdlst_nm,
            ingredient.prdlst_report_no,
            food_category_map.get(ingredient.food_category, ingredient.food_category),
            ingredient.prdlst_dcnm,
            ingredient.ingredient_display_name,
            ingredient.bssh_nm,
        ]
        for allergen in ALLERGEN_LIST:
            row_data.append('O' if allergen in current_allergens else '')
        for gmo in GMO_LIST:
            row_data.append('O' if gmo in current_gmos else '')
        
        sheet.append(row_data)

    # 모든 셀에 폰트 크기 적용
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = font_size_9

    # 파일명에서 사용자 ID 제거
    today_str = timezone.now().strftime('%y%m%d')
    filename = f'LabelData_내원료_{today_str}.xlsx'
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    encoded_filename = quote(filename)
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"

    workbook.save(response)
    return response

@login_required
@transaction.atomic
def upload_my_ingredients_excel(request):
    """
    업로드된 엑셀 파일을 파싱하여 '내 원료'를 일괄 생성/수정합니다.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST 요청만 가능합니다.'})

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return JsonResponse({'success': False, 'message': '엑셀 파일이 없습니다.'})

    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        success_count = 0
        failure_count = 0
        failure_details = {} # 실패 사유별 건수 저장
        
        headers = [cell.value for cell in sheet[1]]
        reverse_food_category_map = {'가공식품': 'processed', '식품첨가물': 'additive', '농수산물': 'agricultural', '정제수': 'water'}
        
        for index, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
            row = [cell.value for cell in row_cells]
            if not any(row):
                continue

            row_dict = dict(zip(headers, row))
            prdlst_nm = row_dict.get('원료명')

            if not prdlst_nm:
                failure_count += 1
                failure_details['원료명 누락'] = failure_details.get('원료명 누락', 0) + 1
                continue

            try:
                food_category_kr = row_dict.get('식품구분', '')
                food_category_code = reverse_food_category_map.get(food_category_kr, food_category_kr)
                prdlst_dcnm = row_dict.get('식품유형', '')
                prdlst_report_no = row_dict.get('품목보고번호', '')
                bssh_nm = row_dict.get('제조사', '')

                filter_conditions = {
                    'user_id': request.user,
                    'prdlst_nm': prdlst_nm,
                    'food_category': food_category_code or '',
                    'prdlst_dcnm': prdlst_dcnm or '',
                    'prdlst_report_no': prdlst_report_no or '',
                    'bssh_nm': bssh_nm or ''
                }
                
                if MyIngredient.objects.filter(**filter_conditions).exists():
                    failure_count += 1
                    failure_details['동일원료'] = failure_details.get('동일원료', 0) + 1
                    continue

                allergens = [h for h in ALLERGEN_LIST if str(row_dict.get(h, '')).strip().upper() == 'O']
                gmos = [h for h in GMO_LIST if str(row_dict.get(h, '')).strip().upper() == 'O']

                MyIngredient.objects.create(
                    user_id=request.user,
                    prdlst_nm=prdlst_nm,
                    prdlst_report_no=prdlst_report_no or '',
                    food_category=food_category_code or '',
                    prdlst_dcnm=prdlst_dcnm or '',
                    ingredient_display_name=row_dict.get('원료표시명', prdlst_nm),
                    bssh_nm=bssh_nm or '',
                    allergens=','.join(allergens),
                    gmo=','.join(gmos),
                    delete_YN='N',
                    update_datetime=timezone.now()
                )
                success_count += 1

            except Exception as e:
                failure_count += 1
                failure_details['처리 오류'] = failure_details.get('처리 오류', 0) + 1
                continue

        # --- [수정] 최종 결과 메시지 생성 로직 ---
        message = f"성공 {success_count}건, 실패 {failure_count}건"
        if failure_count > 0:
            detail_parts = [f"{reason} {count}건" for reason, count in failure_details.items()]
            # 상세 사유를 기본 메시지에 추가
            message += f" ({', '.join(detail_parts)})"

        # 처리 오류가 있었다면 전체 롤백
        if failure_details.get('처리 오류', 0) > 0:
            transaction.set_rollback(True)
            return JsonResponse({
                'success': False, 
                'message': f"오류가 발생하여 전체 업로드가 취소되었습니다. ({message})"
            })

        return JsonResponse({
            'success': True, 
            'message': message
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': f'파일 처리 중 심각한 오류가 발생했습니다: {str(e)}'})

@login_required
def get_recent_usage_api(request):
    """
    최근 사용한 항목 API - 사용자의 최근 라벨에서 해당 필드값 추출
    """
    try:
        field_name = request.GET.get('field')
        limit = int(request.GET.get('limit', 5))
        
        if not field_name:
            return JsonResponse({
                'success': False,
                'error': 'field 파라미터가 필요합니다.'
            }, status=400)
        
        # 현재 사용자의 최근 라벨에서 해당 필드값 추출
        recent_labels = MyLabel.objects.filter(
            user_id=request.user
        ).order_by('-update_datetime')[:50]  # 최근 50개 라벨
        
        recent_items = []
        seen_contents = set()  # 중복 제거용
        
        for label in recent_labels:
            field_value = getattr(label, field_name, None)
            if field_value and field_value.strip() and field_value not in seen_contents:
                recent_items.append({
                    'content': field_value.strip(),
                    'field': field_name,
                    'last_used': label.update_datetime.isoformat() if label.update_datetime else None,
                    'label_name': label.my_label_name or '무제'
                })
                seen_contents.add(field_value)
                
                # limit 도달하면 중단
                if len(recent_items) >= limit:
                    break
        
        return JsonResponse({
            'success': True,
            'field_name': field_name,
            'recent_items': recent_items,
            'total_count': len(recent_items),
            'message': f'{field_name}에 대한 {len(recent_items)}개 최근 사용 항목을 찾았습니다.'
        }, json_dumps_params={'ensure_ascii': False})
        
    except ValueError as e:
        return JsonResponse({
            'success': False,
            'error': f'잘못된 파라미터: {str(e)}'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'최근 사용 항목 조회 중 오류: {str(e)}'
        }, status=500)

@login_required
@require_GET
def auto_fill_api(request):
    """
    기본 추천 시스템용 자동 채우기 API
    스마트 추천 기능을 시뮬레이션하여 기본적인 추천 제공
    """
    try:
        input_field = request.GET.get('input_field', '')
        input_value = request.GET.get('input_value', '')
        category = request.GET.get('category', 'general')
        priority = request.GET.get('priority', 'medium')
        
        # 필드별 기본 추천 데이터 (recommendation_system.js와 동일)
        basic_recommendations = {
            'prdlst_nm': [
                "진한 사골곰탕", "100% 착즙 사과주스", "즉석 발아현미밥",
                "바삭한 통밀 쿠키", "담백한 살코기참치", "글루텐프리 쌀 파스타"
            ],
            'ingredient_info': [  # 상세 입력 영역
                "딸기농축액 1.5 % (고형분 65 %, 딸기 100 %)", "국산 돼지고기 92.5 %",
                "유기농 토마토 85 %", "고등어 50 %, 정제수, 토마토소스",
                "칼슘 110 mg, 비타민D 5 μg", "페닐알라닌 함유",
                "과량 섭취 시 설사를 일으킬 수 있습니다."
            ],
            'prdlst_dcnm': [  # 상세 입력 영역
                "과·채주스 (살균제품)", "식육함유가공품 (비살균제품 / 가열하여 섭취하는 냉동식품)",
                "건강기능식품 (홍삼제품)", "특수용도식품 (체중조절용 조제식품)",
                "초콜릿가공품 (과자)", "어육소시지 (멸균제품)"
            ],
            'prdlst_report_no': [
                "20240001234", "20240005678", "20240009012", "20240003456",
                "20240007890", "20240002468", "20240008024", "20240004680",
                "20240006802", "20240001357"
            ],
            'frmlc_mtrqlt': [
                "폴리에틸렌(PE)", "용기: 폴리프로필렌(PP), 리드필름(뚜껑): OTHER(복합재질)",
                "용기: 페트(PET), 뚜껑: 폴리에틸렌(PE)", "알루미늄",
                "종이, 폴리에틸렌(내면)"
            ],
            'cautions': [
                "부정·불량식품신고는 국번없이 1399",
                "이 제품은 알류(가금류), 우유, 메밀, 땅콩, 대두, 밀, 게, 새우, 돼지고기, 복숭아, 토마토, 아황산류, 호두, 닭고기, 쇠고기, 오징어, 조개류(굴, 전복, 홍합 포함), 잣을 사용한 제품과 같은 제조시설에서 제조하고 있습니다.",
                "개봉 후 냉장보관하시고 가급적 빠른 시일 내에 섭취하시기 바랍니다.",
                "어린이, 임산부, 카페인 민감자는 섭취에 주의해 주세요.",
                "직사광선을 피하여 보관하시기 바랍니다.",
                "흔들어 드세요"
            ],
            'additional_info': [
                "반품 및 교환장소: 구입처 또는 제조원",
                "본 제품은 소비자분쟁해결기준에 의거 교환 또는 보상을 받을 수 있습니다.",
                "무료소비자 상담실: 080-***-**** (평일 오전9시~오후6시)",
                "개봉 전·후 주의사항을 반드시 확인하세요"
            ],
            'storage_method': [
                "냉장 보관 (0~10℃)", "냉동 보관 (-18℃ 이하)",
                "실온 보관 (1~35℃)", "상온 보관 (15~25℃)",
                "직사광선을 피하고 서늘한 곳에 보관하십시오.",
                "개봉 후에는 냉장 보관하시고, 가급적 빨리 드시기 바랍니다."
            ],
            'content_weight': [
                "100g", "200g", "250g", "300g", "500g", 
            ],
            'bssh_nm': [
                "(주)한국식품", "대한제과", "맛있는식품(주)", "우리농장",
            ],
            'pog_daycnt': [
                "제조일로부터 12개월", "제조일로부터 18개월", 
                "제조일로부터 24개월", "제조일로부터 6개월",
            ],
            'processing_condition': [
                "85℃에서 15분간 살균",
                "121℃에서 4분간 멸균", 
                "냉동 -18℃ 이하 보관",
                "상온 유통 가능",
                "65℃에서 30분간 저온살균",
                "100℃에서 10분간 끓임 살균",
                "자외선 살균 처리",
                "고압 살균 처리"
            ]
        }
        
        # 해당 필드의 추천 데이터 가져오기
        suggestions = basic_recommendations.get(input_field, [])
        
        # 입력값이 있으면 관련성 높은 순서로 정렬
        if input_value and suggestions:
            # 간단한 키워드 매칭으로 관련성 계산
            input_lower = input_value.lower()
            scored_suggestions = []
            
            for suggestion in suggestions:
                score = 0
                suggestion_lower = suggestion.lower()
                
                # 완전 일치 시 높은 점수
                if input_lower in suggestion_lower:
                    score += 10
                
                # 첫 글자 일치 시 점수 추가
                if suggestion_lower.startswith(input_lower[:1]):
                    score += 5
                    
                scored_suggestions.append((score, suggestion))
            
            # 점수순으로 정렬하고 상위 3개만 선택
            scored_suggestions.sort(key=lambda x: x[0], reverse=True)
            suggestions = [item[1] for item in scored_suggestions[:3]]
        else:
            # 입력값이 없으면 상위 3개만
            suggestions = suggestions[:3]
        
        return JsonResponse({
            'success': True,
            'suggestions': suggestions,
            'field': input_field,
            'category': category,
            'priority': priority,
            'message': f'{input_field}에 대한 {len(suggestions)}개 추천을 제공합니다.'
        }, json_dumps_params={'ensure_ascii': False})
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'자동 채우기 API 오류: {str(e)}'
        }, status=500)

@login_required  
@require_GET
def phrases_api(request):
    """
    사용자 정의 문구 API
    MyPhrase 모델에서 사용자의 저장된 문구들을 반환
    """
    try:
        category = request.GET.get('category', 'general')
        field_name = request.GET.get('field', '')
        
        # 카테고리별 문구 조회
        user_phrases = MyPhrase.objects.filter(
            user_id=request.user,
            delete_YN='N'
        )
        
        if category != 'all':
            user_phrases = user_phrases.filter(category_name=category)
            
        user_phrases = user_phrases.order_by('display_order', 'my_phrase_id')
        
        phrases = []
        for phrase in user_phrases:
            phrases.append({
                'id': phrase.my_phrase_id,
                'name': phrase.my_phrase_name,
                'text': phrase.comment_content,  # recommendation_system.js에서 text 속성 사용
                'content': phrase.comment_content,
                'note': phrase.note or '',
                'category': phrase.category_name,
                'order': phrase.display_order
            })
        
        return JsonResponse({
            'success': True,
            'phrases': phrases,
            'data': phrases,  # recommendation_system.js에서 data 속성도 확인
            'category': category,
            'field': field_name,
            'total_count': len(phrases),
            'message': f'{category} 카테고리의 {len(phrases)}개 문구를 제공합니다.'
        }, json_dumps_params={'ensure_ascii': False})
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'문구 API 오류: {str(e)}'
        }, status=500)


# 제품 조회 - 국내 제품
@login_required
def food_item_list_domestic(request):
    """국내 제품 조회"""
    # 기존 food_item_list에 food_category=domestic 파라미터 추가
    from django.http import QueryDict
    query_dict = request.GET.copy()
    query_dict['food_category'] = 'domestic'
    request.GET = query_dict
    return food_item_list(request)


# 제품 조회 - 수입 제품
@login_required
def food_item_list_imported(request):
    """수입 제품 조회"""
    # 기존 food_item_list에 food_category=imported 파라미터 추가
    from django.http import QueryDict
    query_dict = request.GET.copy()
    query_dict['food_category'] = 'imported'
    request.GET = query_dict
    return food_item_list(request)


# 식품첨가물 검색
@login_required
def food_additive_search(request):
    """식품첨가물 번호 검색"""
    # 식품첨가물 조회 로깅
    log_user_activity(request, 'search', 'search_additive')
    
    name_kr = request.GET.get('name_kr', '').strip()
    name_en = request.GET.get('name_en', '').strip()
    alias_name = request.GET.get('alias_name', '').strip()
    ins_no = request.GET.get('ins_no', '').strip()
    e_no = request.GET.get('e_no', '').strip()
    cas_no = request.GET.get('cas_no', '').strip()
    sort_by = request.GET.get('sort', 'name_kr')
    order = request.GET.get('order', 'asc')
    
    items_per_page = int(request.GET.get("items_per_page", 10))
    page_number = request.GET.get("page", 1)
    
    # 혼합제제류는 제외 (내원료 관리 등에서만 사용)
    additives_qs = FoodAdditive.objects.exclude(category='혼합제제류')
    
    # 각 필드별 검색 조건 적용 (LIKE 검색 - 부분 일치)
    if name_kr:
        # 쉼표로 구분된 경우 OR 검색
        if ',' in name_kr:
            from django.db.models import Q
            q = Q()
            for term in name_kr.split(','):
                term = term.strip()
                if term:
                    q |= Q(name_kr__icontains=term)
            additives_qs = additives_qs.filter(q)
        else:
            additives_qs = additives_qs.filter(name_kr__icontains=name_kr)
    
    if name_en:
        if ',' in name_en:
            from django.db.models import Q
            q = Q()
            for term in name_en.split(','):
                term = term.strip()
                if term:
                    q |= Q(name_en__icontains=term)
            additives_qs = additives_qs.filter(q)
        else:
            additives_qs = additives_qs.filter(name_en__icontains=name_en)
    
    if alias_name:
        if ',' in alias_name:
            from django.db.models import Q
            q = Q()
            for term in alias_name.split(','):
                term = term.strip()
                if term:
                    q |= Q(alias_name__icontains=term)
            additives_qs = additives_qs.filter(q)
        else:
            additives_qs = additives_qs.filter(alias_name__icontains=alias_name)
    
    if ins_no:
        additives_qs = additives_qs.filter(ins_no__icontains=ins_no)
    if e_no:
        additives_qs = additives_qs.filter(e_no__icontains=e_no)
    if cas_no:
        additives_qs = additives_qs.filter(cas_no__icontains=cas_no)
    
    # 정렬
    if order == 'desc':
        sort_by = f'-{sort_by}'
    additives_qs = additives_qs.order_by(sort_by)
    
    total_count = additives_qs.count()
    
    paginator, page_obj, page_range = paginate_queryset(additives_qs, page_number, items_per_page)
    querystring_without_page = get_querystring_without(request, ["page"])
    querystring_without_sort = get_querystring_without(request, ["sort", "order"])
    
    # search_values 딕셔너리 생성
    search_values = {
        'name_kr': name_kr,
        'name_en': name_en,
        'alias_name': alias_name,
        'ins_no': ins_no,
        'e_no': e_no,
        'cas_no': cas_no,
    }
    
    # 검색 조건이 있으면 검색 결과 수 전달
    has_search = any([name_kr, name_en, alias_name, ins_no, e_no, cas_no])
    search_result_count = total_count if has_search else None
    
    context = {
        "page_obj": page_obj,
        "paginator": paginator,
        "page_range": page_range,
        "search_values": search_values,
        "items_per_page": items_per_page,
        "total_count": total_count,
        "search_result_count": search_result_count,
        "querystring_without_page": querystring_without_page,
        "querystring_without_sort": querystring_without_sort,
    }
    
    return render(request, "label/food_additive_search.html", context)


@login_required
@require_POST
def copy_additives_to_ingredients(request):
    """선택한 식품첨가물을 내 원료로 복사"""
    try:
        data = json.loads(request.body)
        additives = data.get('additives', [])
        
        if not additives:
            return JsonResponse({'success': False, 'error': '선택한 항목이 없습니다.'})
        
        created_count = 0
        skipped_count = 0
        
        for additive_data in additives:
            name_kr = additive_data.get('name_kr')
            
            # 이미 존재하는지 확인
            existing = MyIngredient.objects.filter(
                user_id=request.user,
                prdlst_nm=name_kr
            ).first()
            
            if existing:
                skipped_count += 1
                continue
            
            # 새 원료 생성 (식품첨가물명을 모든 필드에 입력)
            MyIngredient.objects.create(
                user_id=request.user,
                prdlst_nm=name_kr,  # 원료명
                food_category='additive',  # 식품구분 (영문 코드)
                prdlst_dcnm=name_kr,  # 식품유형
                rawmtrl_nm=name_kr,  # 원재료명
                ingredient_display_name=name_kr,  # 원재료 표시명
                delete_YN='N'  # 삭제되지 않은 상태
            )
            created_count += 1
        
        message = f'{created_count}개의 식품첨가물이 내 원료로 복사되었습니다.'
        if skipped_count > 0:
            message += f' ({skipped_count}개는 이미 존재하여 건너뛰었습니다.)'
        
        return JsonResponse({'success': True, 'message': message})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def request_additive_correction(request):
    """식품첨가물 데이터 수정 요청"""
    try:
        data = json.loads(request.body)
        corrections = data.get('corrections', [])
        reason = data.get('reason', '')
        
        if not corrections or not reason:
            return JsonResponse({'success': False, 'error': '수정 내용과 사유를 모두 입력해주세요.'})
        
        # 게시판에 수정 요청 글 작성
        from v1.board.models import Board
        
        # 작성자 이름 마스킹 (앞 2글자만 표시)
        user_email = request.user.email
        if len(user_email) >= 2:
            masked_email = user_email[:2] + '*' * (len(user_email.split('@')[0]) - 2) + '@' + user_email.split('@')[1]
        else:
            masked_email = '*' * len(user_email)
        
        # 수정 요청 내용 포맷팅
        content_lines = ['[식품첨가물 데이터 수정 요청]\n']
        content_lines.append(f'요청자: {masked_email}\n')
        content_lines.append(f'요청일시: {timezone.now().strftime("%Y-%m-%d %H:%M")}\n\n')
        content_lines.append('=== 수정 내용 ===\n')
        
        for idx, correction in enumerate(corrections, 1):
            name_kr = correction.get('name_kr')
            fields = correction.get('fields', {})
            
            try:
                additive = FoodAdditive.objects.get(name_kr=name_kr)
                content_lines.append(f'\n{idx}. {additive.name_kr}\n')
                content_lines.append(f'   [현재 데이터]\n')
                content_lines.append(f'   - 영문명: {additive.name_en or "(없음)"}\n')
                content_lines.append(f'   - 이명: {additive.alias_name or "(없음)"}\n')
                content_lines.append(f'   - INS No.: {additive.ins_no or "(없음)"}\n')
                content_lines.append(f'   - E No.: {additive.e_no or "(없음)"}\n')
                content_lines.append(f'   - CAS No.: {additive.cas_no or "(없음)"}\n')
                content_lines.append(f'   [수정 요청 데이터]\n')
                content_lines.append(f'   - 영문명: {fields.get("name_en", "")}\n')
                content_lines.append(f'   - 이명: {fields.get("alias_name", "")}\n')
                content_lines.append(f'   - INS No.: {fields.get("ins_no", "")}\n')
                content_lines.append(f'   - E No.: {fields.get("e_no", "")}\n')
                content_lines.append(f'   - CAS No.: {fields.get("cas_no", "")}\n')
            except FoodAdditive.DoesNotExist:
                content_lines.append(f'\n{idx}. [식품첨가물명: {name_kr}] - 항목을 찾을 수 없음\n')
        
        content_lines.append('\n=== 수정 사유 및 근거 ===\n')
        content_lines.append(reason)
        
        content = ''.join(content_lines)
        
        # 게시판에 비밀글로 작성
        Board.objects.create(
            author=request.user,
            title=f'[데이터 수정 요청] 식품첨가물 {len(corrections)}건',
            content=content,
            is_notice=False,
            is_hidden=True  # 비밀글로 설정
        )
        
        return JsonResponse({
            'success': True,
            'message': '수정 요청이 게시판에 등록되었습니다. 관리자가 확인 후 반영하겠습니다.'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def get_additive_field_settings(request):
    """식품첨가물/혼합제제/농수축산물 선택 시 필드 표시 규칙을 반환하는 API"""
    try:
        food_group = request.GET.get('food_group', '')
        food_type = request.GET.get('food_type', '')  # 농수축산물 소분류 판별용
        
        # 식품첨가물 또는 혼합제제인 경우
        if food_group in ['식품첨가물', '혼합제제']:
            settings = {
                'prdlst_nm': 'Y',  # 제품명: 필수
                'ingredient_info': 'N',  # 특정성분 함량: 선택
                'prdlst_dcnm': 'Y',  # 식품유형명: 필수
                'content_weight': 'Y',  # 내용량: 필수
                'weight_calorie': 'Y',  # 내용량(열량): 필수
                'prdlst_report_no': 'Y',  # 품목보고번호: 필수
                'country_of_origin': 'N',  # 원산지: 선택
                'frmlc_mtrqlt': 'Y',  # 포장재질: 필수
                'pog_daycnt': 'Y',  # 소비기한: 필수
                'rawmtrl_nm': 'Y',  # 원재료명: 필수
                'storage_method': 'Y',  # 보관방법: 필수
                'bssh_nm': 'Y',  # 제조원: 필수
                'nutritions': 'D',  # 영양성분: 조회불가
                'cautions': 'Y',  # 주의사항: 필수
            }
            
            return JsonResponse({
                'success': True,
                'settings': settings
            })
        
        # 농수축산물인 경우
        elif food_group == '농수축산물':
            # 기본 설정 (공통 비활성화 항목)
            settings = {
                'prdlst_nm': 'Y',  # 제품명: 필수
                'ingredient_info': 'D',  # 특정성분 함량: 비활성화
                'prdlst_dcnm': 'D',  # 식품유형명: 비활성화
                'content_weight': 'Y',  # 내용량: 필수
                'weight_calorie': 'D',  # 내용량(열량): 비활성화
                'prdlst_report_no': 'D',  # 품목보고번호: 비활성화
                'country_of_origin': 'Y',  # 원산지: 필수
                'frmlc_mtrqlt': 'D',  # 포장재질: 비활성화
                'pog_daycnt': 'N',  # 소비기한: 기본 선택
                'rawmtrl_nm': 'N',  # 원재료명: 기본 선택
                'storage_method': 'N',  # 보관방법: 기본 선택
                'bssh_nm': 'Y',  # 제조원/생산자: 필수
                'nutritions': 'D',  # 영양성분: 비활성화
                'cautions': 'D',  # 주의사항: 비활성화
            }
            
            custom_fields = []
            
            # 농산물 (사과, 채소 등)
            if food_type == '농산물':
                settings.update({
                    'pog_daycnt': 'N',  # 소비기한: 선택
                    'rawmtrl_nm': 'D',  # 원재료명: 비활성화
                    'storage_method': 'N',  # 보관방법: 선택
                })
                custom_fields = [
                    {'label': '생산연도 (또는 생산연월일)', 'value': ''},
                    {'label': '포장일', 'value': ''},
                    {'label': '품종', 'value': ''},
                    {'label': '등급 (표준규격품인 경우)', 'value': ''},
                ]
            
            # 수산물 (생선, 조개 등)
            elif food_type == '수산물':
                settings.update({
                    'pog_daycnt': 'N',  # 소비기한: 선택
                    'rawmtrl_nm': 'D',  # 원재료명: 비활성화
                    'storage_method': 'N',  # 보관방법: 선택
                })
                custom_fields = [
                    {'label': '생산연월일', 'value': ''},
                    {'label': '포장일', 'value': ''},
                    {'label': '등급', 'value': ''},
                    {'label': '마릿수', 'value': ''},
                ]
            
            # 축산물 (생고기 등)
            elif food_type == '축산물':
                settings.update({
                    'pog_daycnt': 'Y',  # 소비기한: 필수
                    'rawmtrl_nm': 'Y',  # 원재료명: 필수
                    'storage_method': 'N',  # 보관방법: 선택 (냉장/냉동)
                })
                custom_fields = [
                    {'label': '이력관리번호', 'value': ''},
                    {'label': '등급', 'value': ''},
                    {'label': '부위', 'value': ''},
                    {'label': '도축일', 'value': ''},
                    {'label': '포장일', 'value': ''},
                    {'label': '도축장명', 'value': ''},
                    {'label': '보관방법 (냉장/냉동)', 'value': ''},
                ]
            
            return JsonResponse({
                'success': True,
                'settings': settings,
                'custom_fields': custom_fields  # 맞춤항목 추가
            })
        
        else:
            return JsonResponse({
                'success': False,
                'error': '식품첨가물, 혼합제제 또는 농수축산물이 아닙니다.'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
